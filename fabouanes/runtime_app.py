from __future__ import annotations

import json
import re
import os
import shutil
import sqlite3
import traceback
import atexit
import ast
import secrets
from collections import defaultdict
from contextlib import contextmanager
from datetime import date, datetime
from functools import wraps
from pathlib import Path
from typing import Any

from flask import Flask, abort, flash, g, jsonify, redirect, render_template, request, session, url_for, send_file, send_from_directory
from werkzeug.exceptions import HTTPException
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
from io import BytesIO

from fabouanes.config import APP_DATA_DIR, BASE_DIR, BUNDLED_DB_PATH, DATABASE_URL, DEFAULT_ADMIN_PASSWORD, DEFAULT_ADMIN_USERNAME, SESSION_COOKIE_SECURE
from fabouanes.db import connect_database, list_columns
from fabouanes.core.permissions import has_permission, normalize_role, permission_for_endpoint, permission_denied_response
from fabouanes.security import csrf_protect as shared_csrf_protect, ensure_csrf as shared_ensure_csrf, security_headers, validate_password_strength
from fabouanes.services.backup_service import start_background_services

try:
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image as RLImage
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

APP_DATA_DIR.mkdir(parents=True, exist_ok=True)

load_dotenv(BASE_DIR / '.env')
load_dotenv(APP_DATA_DIR / '.env', override=False)
DB_PATH = APP_DATA_DIR / 'database.db'
BACKUP_DIR = APP_DATA_DIR / 'backups'
LOCAL_BACKUP_DIR = BACKUP_DIR / 'local'
LOG_DIR = APP_DATA_DIR / 'logs'
REPORT_DIR = APP_DATA_DIR / 'reports_generated'
NOTES_DIR = APP_DATA_DIR / 'notes'
PDF_READER_DIR = APP_DATA_DIR / 'pdf_reader'
IMPORT_DIR = APP_DATA_DIR / 'imports'


app = Flask(__name__, template_folder=str(BASE_DIR / 'templates'), static_folder=str(BASE_DIR / 'static'))
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

SECRET_KEY = os.environ.get('SECRET_KEY') or get_setting('secret_key', '') if 'get_setting' in globals() else os.environ.get('SECRET_KEY')
if not SECRET_KEY:
    SECRET_KEY = secrets.token_hex(32)
    env_path = APP_DATA_DIR / '.env'
    try:
        existing = env_path.read_text(encoding='utf-8') if env_path.exists() else ''
        if 'SECRET_KEY=' not in existing:
            with env_path.open('a', encoding='utf-8') as fh:
                if existing and not existing.endswith('\n'):
                    fh.write('\n')
                fh.write(f'SECRET_KEY={SECRET_KEY}\n')
    except Exception:
        pass
app.config['SECRET_KEY'] = SECRET_KEY

def ensure_runtime_dirs() -> None:
    APP_DATA_DIR.mkdir(parents=True, exist_ok=True)
    LOCAL_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    NOTES_DIR.mkdir(parents=True, exist_ok=True)
    PDF_READER_DIR.mkdir(parents=True, exist_ok=True)
    IMPORT_DIR.mkdir(parents=True, exist_ok=True)

    if not DB_PATH.exists() and BUNDLED_DB_PATH.exists():
        shutil.copy2(BUNDLED_DB_PATH, DB_PATH)


def now_str() -> str:
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def write_text_log(filename: str, message: str) -> None:
    ensure_runtime_dirs()
    with (LOG_DIR / filename).open('a', encoding='utf-8') as f:
        f.write(f"[{now_str()}] {message}\n")


def safe_username() -> str:
    try:
        if getattr(g, 'user', None):
            return g.user['username']
    except Exception:
        pass
    return 'system'




def notes_file_path() -> Path:
    ensure_runtime_dirs()
    return NOTES_DIR / 'bloc_note.txt'


def read_app_notes() -> str:
    path = notes_file_path()
    if path.exists():
        return path.read_text(encoding='utf-8')
    return ''


def write_app_notes(content: str) -> None:
    """Sauvegarde la note courante et archive une version horodatée."""
    ensure_runtime_dirs()
    path = notes_file_path()
    # Archiver l'ancienne version si différente
    old_content = path.read_text(encoding='utf-8') if path.exists() else ''
    if old_content.strip() and old_content.strip() != (content or '').strip():
        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        archive = NOTES_DIR / f'history_{stamp}.txt'
        archive.write_text(old_content, encoding='utf-8')
        # Garder seulement les 20 dernières versions
        hist_files = sorted(NOTES_DIR.glob('history_*.txt'), reverse=True)
        for old_f in hist_files[20:]:
            try:
                old_f.unlink()
            except Exception:
                pass
    path.write_text(content or '', encoding='utf-8')


def list_notes_history() -> list[dict]:
    """Retourne la liste des versions archivées."""
    ensure_runtime_dirs()
    versions = []
    for p in sorted(NOTES_DIR.glob('history_*.txt'), reverse=True)[:20]:
        try:
            stamp = p.stem.replace('history_', '')
            dt = datetime.strptime(stamp, '%Y%m%d_%H%M%S')
            label = dt.strftime('%d/%m/%Y %H:%M:%S')
        except Exception:
            label = p.stem
        versions.append({'filename': p.name, 'label': label})
    return versions


def read_notes_version(filename: str) -> str:
    safe = filename.replace('..', '').replace('/', '').replace('\\', '')
    path = NOTES_DIR / safe
    if path.exists() and path.suffix == '.txt':
        return path.read_text(encoding='utf-8')
    return ''



def parse_flexible_date(value) -> str:
    """Convertit n'importe quel format de date en ISO YYYY-MM-DD."""
    from datetime import datetime as _dt, date as _date
    if isinstance(value, _dt):
        return value.date().isoformat()
    if isinstance(value, _date):
        return value.isoformat()
    s = str(value).strip() if value is not None else ''
    if not s or s.lower() in {'none', 'nan'}:
        return _date.today().isoformat()
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%y', '%Y/%m/%d'):
        try:
            return _dt.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    try:
        return _dt.fromisoformat(s).date().isoformat()
    except Exception:
        return _date.today().isoformat()


def parse_flexible_amount(value) -> float:
    if value in (None, ''):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    import re as _re
    s = str(value).strip().replace('\xa0', '').replace(' ', '').replace('DA', '').replace('da', '')
    s = s.replace(',', '.') if s.count(',') == 1 and s.count('.') == 0 else s.replace(',', '')
    m = _re.search(r'-?\d+(?:\.\d+)?', s)
    return float(m.group(0)) if m else 0.0


def parse_excel_client_history(file_path) -> dict:
    """Extrait uniquement la dernière date et le dernier reste à payer de la fiche Excel."""
    try:
        import openpyxl
    except Exception as exc:
        raise RuntimeError("Le module openpyxl est requis.") from exc
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    def cell_str(v):
        return str(v).strip() if v is not None else ''
    # Trouver la ligne d'en-tête
    header_row = None
    for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [cell_str(v).lower() for v in row]
        joined = ' | '.join(vals)
        if 'designation' in joined and ('montant' in joined or 'reste' in joined or 'versement' in joined):
            header_row = ridx
            break
    if not header_row:
        return {'last_date': None, 'last_balance': 0.0}
    last_date = None
    last_balance = 0.0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        cells = (list(row) + [None] * 6)[:6]
        d, desig, montant, versement, reste = cells[:5]
        # Ignorer les lignes vides
        if not any(v is not None and str(v).strip() and str(v).strip().lower() not in {'none', 'nan'}
                   for v in [d, desig, montant, versement, reste]):
            continue
        # Mettre à jour la dernière date et le dernier solde trouvés
        if d is not None and str(d).strip() and str(d).strip().lower() not in {'none', 'nan'}:
            last_date = parse_flexible_date(d)
        reste_val = parse_flexible_amount(reste)
        if reste_val > 0:
            last_balance = round(reste_val, 2)
    return {'last_date': last_date, 'last_balance': last_balance}


def parse_excel_client_file(file_path) -> dict:
    """Extrait les métadonnées du client (nom, téléphone, crédit initial)."""

    try:
        import openpyxl
    except Exception as exc:
        raise RuntimeError("Le module openpyxl est requis pour l'import Excel.") from exc

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    def cell_str(value: Any) -> str:
        return str(value).strip() if value is not None else ''

    top_rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(6, ws.max_row), values_only=True):
        top_rows.append([cell_str(v) for v in row])

    client_name = ''
    phone = ''
    for row in top_rows:
        cleaned = [v for v in row if v]
        if len(cleaned) >= 2 and not client_name:
            for val in cleaned:
                upper = val.upper()
                if val and 'TEL' not in upper and 'DATE' not in upper and 'DESIGNATION' not in upper:
                    client_name = val
                    break
        for i, val in enumerate(row):
            upper = val.upper()
            if 'TEL' in upper:
                inline = val.split(':', 1)[1].strip() if ':' in val else ''
                if inline:
                    phone = inline
                elif i + 1 < len(row):
                    phone = row[i + 1].strip()

    header_row = None
    for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [cell_str(v).lower() for v in row]
        joined = ' | '.join(vals)
        if 'designation' in joined and ('montant a payer' in joined or 'reste a payer' in joined):
            header_row = ridx
            break

    history_count = 0
    opening_credit = 0.0
    final_balance = 0.0
    if header_row:
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            d, designation, montant, versement, reste = (list(row) + [None] * 5)[:5]
            designation_text = cell_str(designation)
            if not any(v is not None and str(v).strip() for v in [d, designation, montant, versement, reste]):
                continue
            history_count += 1
            reste_num = 0.0
            montant_num = 0.0
            try:
                reste_num = float(reste) if reste not in (None, '') else 0.0
            except Exception:
                pass
            try:
                montant_num = float(montant) if montant not in (None, '') else 0.0
            except Exception:
                pass
            final_balance = reste_num or final_balance
            if designation_text and 'ancien solde' in designation_text.lower():
                opening_credit = reste_num or montant_num or opening_credit

    if opening_credit <= 0 and final_balance > 0:
        opening_credit = final_balance

    if not client_name:
        client_name = file_path.stem.replace('_', ' ').strip()

    notes = f"Importé depuis Excel ({file_path.name}). Lignes détectées: {history_count}."
    return {
        'name': client_name.strip(),
        'phone': phone.strip(),
        'address': '',
        'notes': notes,
        'opening_credit': round(float(opening_credit or 0.0), 2),
        'history_count': history_count,
        'source_file': file_path.name,
    }



def get_db():
    if 'db' not in g:
        g.db = connect_database(DATABASE_URL, DB_PATH)
    return g.db


@app.teardown_appcontext
def close_db(_: Any) -> None:
    db = g.pop('db', None)
    if db is not None:
        db.close()


def query_db(query: str, params: tuple = (), one: bool = False):
    cur = get_db().execute(query, params)
    rows = cur.fetchall()
    cur.close()
    return (rows[0] if rows else None) if one else rows


def execute_db(query: str, params: tuple = ()) -> int:
    db = get_db()
    cur = db.execute(query, params)
    if int(getattr(g, '_db_tx_depth', 0) or 0) == 0:
        db.commit()
    last_id = cur.lastrowid
    cur.close()
    return int(last_id or 0)


@contextmanager
def db_transaction():
    db = get_db()
    previous_depth = int(getattr(g, '_db_tx_depth', 0) or 0)
    g._db_tx_depth = previous_depth + 1
    try:
        yield db
    except Exception:
        if previous_depth == 0:
            try:
                db.rollback()
            except Exception:
                pass
        raise
    else:
        if previous_depth == 0:
            db.commit()
    finally:
        g._db_tx_depth = previous_depth


def get_setting(key: str, default: str = '') -> str:
    try:
        row = query_db('SELECT value FROM app_settings WHERE key = ?', (key,), one=True)
        return row['value'] if row and row['value'] is not None else default
    except Exception:
        return default


def set_setting(key: str, value: str) -> None:
    execute_db('INSERT INTO app_settings (key, value, updated_at) VALUES (?, ?, CURRENT_TIMESTAMP) ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=CURRENT_TIMESTAMP', (key, value))


app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=SESSION_COOKIE_SECURE,
    PERMANENT_SESSION_LIFETIME=60 * 60 * 12,
)


@app.before_request
def ensure_csrf_token() -> None:
    shared_ensure_csrf()

@app.before_request
def csrf_protect() -> None:
    return shared_csrf_protect()

@app.after_request
def apply_security_headers(response):
    return security_headers(response)


@app.context_processor
def inject_csrf_token():
    return {
        'csrf_token': session.get('csrf_token', ''),
        'is_desktop_app': os.environ.get('FAB_DESKTOP', '0') == '1',
    }


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if getattr(g, 'user', None) is None:
            return redirect(url_for('login'))
        return view(*args, **kwargs)

    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        user = getattr(g, 'user', None)
        if user is None:
            return redirect(url_for('login'))
        if user['role'] != 'admin':
            flash('Accès réservé à l’administrateur.', 'danger')
            return redirect(url_for('index'))
        return view(*args, **kwargs)

    return wrapped


def log_activity(action: str, entity_type: str = '', entity_id: int | None = None, details: str = '') -> None:
    username = safe_username()
    execute_db('INSERT INTO activity_logs (username, action, entity_type, entity_id, details, created_at) VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)', (username, action, entity_type, entity_id, details))
    write_text_log('activity.log', f"{username} | {action} | {entity_type}#{entity_id or '-'} | {details}")


def log_system(level: str, message: str) -> None:
    execute_db('INSERT INTO system_logs (level, message, created_at) VALUES (?, ?, CURRENT_TIMESTAMP)', (level, message))
    write_text_log('system.log', f"{level.upper()} | {message}")


def log_error(exc: Exception, route: str = '') -> None:
    username = safe_username()
    tb = traceback.format_exc()
    execute_db('INSERT INTO error_logs (username, route, error_type, message, traceback, created_at) VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)', (username, route or request.path if request else '', type(exc).__name__, str(exc), tb))
    write_text_log('errors.log', f"{username} | {route or (request.path if request else '')} | {type(exc).__name__}: {exc}\n{tb}")


def backup_database(reason: str = 'manual') -> Path:
    ensure_runtime_dirs()
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    suffix = 'sql' if DATABASE_URL.lower().startswith('postgres') else 'db'
    filename = f"database_{stamp}_{reason.replace(' ', '_')}.{suffix}"
    target = LOCAL_BACKUP_DIR / filename
    db = g.get('db')
    if db is not None:
        db.commit()
    if DATABASE_URL.lower().startswith('postgres'):
        target.write_text('-- Backup PostgreSQL logique non implémenté automatiquement dans cette version.\n', encoding='utf-8')
    else:
        shutil.copy2(DB_PATH, target)
    cloud_path = get_setting('gdrive_backup_dir', '').strip()
    if cloud_path:
        try:
            cloud_dir = Path(cloud_path)
            cloud_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy2(target, cloud_dir / filename)
        except Exception as cloud_exc:
            write_text_log('errors.log', f"cloud backup failed: {cloud_exc}")
    return target


def restore_database_from(path_str: str) -> None:
    db = g.pop('db', None)
    if db is not None:
        db.close()
    shutil.copy2(path_str, DB_PATH)


def get_cloud_backup_dir() -> Path | None:
    cloud_path = get_setting('gdrive_backup_dir', '').strip()
    if not cloud_path:
        return None
    return Path(cloud_path)


def list_restore_backups() -> list[dict[str, str]]:
    backups: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for p in sorted(LOCAL_BACKUP_DIR.glob('*.db'), reverse=True):
        item = {'value': f'local:{p.name}', 'name': p.name, 'source': 'local', 'label': f'Local · {p.name}'}
        backups.append(item)
        seen.add(('local', p.name))
    cloud_dir = get_cloud_backup_dir()
    if cloud_dir and cloud_dir.exists():
        for p in sorted(cloud_dir.glob('*.db'), reverse=True):
            if ('cloud', p.name) in seen:
                continue
            backups.append({'value': f'cloud:{p.name}', 'name': p.name, 'source': 'cloud', 'label': f'Google Drive · {p.name}'})
    return backups


def resolve_backup_path(backup_value: str) -> Path | None:
    raw = (backup_value or '').strip()
    if not raw:
        return None
    if ':' in raw:
        source, name = raw.split(':', 1)
    else:
        source, name = 'local', raw
    if source == 'cloud':
        cloud_dir = get_cloud_backup_dir()
        if cloud_dir:
            path = cloud_dir / name
            if path.exists():
                return path
    path = LOCAL_BACKUP_DIR / name
    return path if path.exists() else None


def to_float(value: str | None, default: float = 0.0) -> float:
    try:
        return float(str(value).replace(',', '.'))
    except (TypeError, ValueError):
        return default


def qty_to_kg(quantity: float, unit: str | None) -> float:
    unit = (unit or 'kg').strip().lower()
    if unit == 'sac':
        return quantity * 50
    if unit in {'qt', 'quintal'}:
        return quantity * 100
    return quantity


def unit_price_to_kg(unit_price: float, unit: str | None) -> float:
    unit = (unit or 'kg').strip().lower()
    if unit == 'sac':
        return unit_price / 50
    if unit in {'qt', 'quintal'}:
        return unit_price / 100
    return unit_price


def kg_to_display(quantity_kg: float, unit: str | None) -> float:
    unit = (unit or 'kg').strip().lower()
    if unit == 'sac':
        return quantity_kg / 50
    if unit in {'qt', 'quintal'}:
        return quantity_kg / 100
    return quantity_kg


def unit_display_factor(unit: str | None) -> float:
    unit = (unit or 'kg').strip().lower()
    if unit == 'sac':
        return 50.0
    if unit in {'qt', 'quintal'}:
        return 100.0
    return 1.0


def unit_price_to_display(unit_price_kg: float, unit: str | None) -> float:
    return float(unit_price_kg or 0) * unit_display_factor(unit)


def wants_print_after_submit() -> bool:
    return (request.form.get('print_after') or '').strip().lower() in {'1', 'true', 'on', 'yes'}


def recalc_raw_material_avg_cost(material_id: int) -> None:
    material = query_db('SELECT id, stock_qty, avg_cost FROM raw_materials WHERE id = ?', (material_id,), one=True)
    if not material:
        return
    stock_qty = float(material['stock_qty'])
    purchase_rows = query_db('SELECT quantity, unit_price FROM purchases WHERE raw_material_id = ? ORDER BY purchase_date, id', (material_id,))
    purchased_qty = sum(float(r['quantity']) for r in purchase_rows)
    purchased_value = sum(float(r['quantity']) * float(r['unit_price']) for r in purchase_rows)
    base_qty = max(0.0, stock_qty - purchased_qty)
    base_value = base_qty * float(material['avg_cost'])
    total_qty = base_qty
    total_value = base_value
    for row in purchase_rows:
        total_qty += float(row['quantity'])
        total_value += float(row['quantity']) * float(row['unit_price'])
    new_avg = (total_value / total_qty) if total_qty > 0 else 0.0
    execute_db('UPDATE raw_materials SET avg_cost = ? WHERE id = ?', (new_avg, material_id))


def recalc_finished_product_avg_cost(product_id: int) -> None:
    product = query_db('SELECT id, stock_qty, avg_cost FROM finished_products WHERE id = ?', (product_id,), one=True)
    if not product:
        return
    stock_qty = float(product['stock_qty'])
    prod_rows = query_db('SELECT output_quantity, production_cost FROM production_batches WHERE finished_product_id = ? ORDER BY production_date, id', (product_id,))
    produced_qty = sum(float(r['output_quantity']) for r in prod_rows)
    produced_value = sum(float(r['production_cost']) for r in prod_rows)
    base_qty = max(0.0, stock_qty - produced_qty)
    base_value = base_qty * float(product['avg_cost'])
    total_qty = base_qty
    total_value = base_value
    for row in prod_rows:
        total_qty += float(row['output_quantity'])
        total_value += float(row['production_cost'])
    new_avg = (total_value / total_qty) if total_qty > 0 else 0.0
    execute_db('UPDATE finished_products SET avg_cost = ? WHERE id = ?', (new_avg, product_id))


def create_purchase_document_record(supplier_id, purchase_date: str, notes: str) -> int:
    return execute_db(
        'INSERT INTO purchase_documents (supplier_id, total, purchase_date, notes) VALUES (?, 0, ?, ?)',
        (supplier_id, purchase_date, notes),
    )


def ensure_purchase_document_record(document_id: int, supplier_id, purchase_date: str, notes: str) -> int:
    existing = query_db('SELECT id FROM purchase_documents WHERE id = ?', (document_id,), one=True)
    if existing:
        return int(existing['id'])
    execute_db(
        'INSERT INTO purchase_documents (id, supplier_id, total, purchase_date, notes) VALUES (?, ?, 0, ?, ?)',
        (document_id, supplier_id, purchase_date, notes),
    )
    return document_id


def recalc_purchase_document_totals(document_id: int | None) -> None:
    if not document_id:
        return
    totals = query_db(
        'SELECT COUNT(*) AS line_count, COALESCE(SUM(total), 0) AS total_amount FROM purchases WHERE document_id = ?',
        (document_id,),
        one=True,
    )
    if not totals or int(totals['line_count'] or 0) <= 0:
        execute_db('DELETE FROM purchase_documents WHERE id = ?', (document_id,))
        return
    execute_db(
        'UPDATE purchase_documents SET total = ? WHERE id = ?',
        (float(totals['total_amount'] or 0), document_id),
    )


def create_sale_document_record(client_id, sale_type: str, sale_date: str, notes: str) -> int:
    return execute_db(
        '''
        INSERT INTO sale_documents (client_id, sale_type, total, amount_paid, balance_due, sale_date, notes)
        VALUES (?, ?, 0, 0, 0, ?, ?)
        ''',
        (client_id, sale_type, sale_date, notes),
    )


def ensure_sale_document_record(document_id: int, client_id, sale_type: str, sale_date: str, notes: str) -> int:
    existing = query_db('SELECT id FROM sale_documents WHERE id = ?', (document_id,), one=True)
    if existing:
        return int(existing['id'])
    execute_db(
        '''
        INSERT INTO sale_documents (id, client_id, sale_type, total, amount_paid, balance_due, sale_date, notes)
        VALUES (?, ?, ?, 0, 0, 0, ?, ?)
        ''',
        (document_id, client_id, sale_type, sale_date, notes),
    )
    return document_id


def recalc_sale_document_totals(document_id: int | None) -> None:
    if not document_id:
        return
    finished_totals = query_db(
        'SELECT COUNT(*) AS line_count, COALESCE(SUM(total), 0) AS total_amount, COALESCE(SUM(amount_paid), 0) AS paid_amount, COALESCE(SUM(balance_due), 0) AS due_amount FROM sales WHERE document_id = ?',
        (document_id,),
        one=True,
    )
    raw_totals = query_db(
        'SELECT COUNT(*) AS line_count, COALESCE(SUM(total), 0) AS total_amount, COALESCE(SUM(amount_paid), 0) AS paid_amount, COALESCE(SUM(balance_due), 0) AS due_amount FROM raw_sales WHERE document_id = ?',
        (document_id,),
        one=True,
    )
    line_count = int((finished_totals['line_count'] if finished_totals else 0) or 0) + int((raw_totals['line_count'] if raw_totals else 0) or 0)
    if line_count <= 0:
        execute_db('DELETE FROM sale_documents WHERE id = ?', (document_id,))
        return
    total_amount = float((finished_totals['total_amount'] if finished_totals else 0) or 0) + float((raw_totals['total_amount'] if raw_totals else 0) or 0)
    paid_amount = float((finished_totals['paid_amount'] if finished_totals else 0) or 0) + float((raw_totals['paid_amount'] if raw_totals else 0) or 0)
    due_amount = float((finished_totals['due_amount'] if finished_totals else 0) or 0) + float((raw_totals['due_amount'] if raw_totals else 0) or 0)
    execute_db(
        'UPDATE sale_documents SET total = ?, amount_paid = ?, balance_due = ? WHERE id = ?',
        (total_amount, paid_amount, due_amount, document_id),
    )


def fmt_money(v: float | int | None) -> str:
    try:
        if v is None:
            return '0,00 DA'
        amount = f"{float(v):,.2f}".replace(',', ' ').replace('.', ',')
        return f'{amount} DA'
    except Exception:
        return '0,00 DA'




def fmt_qty(v):
    try:
        if v is None:
            return '0'
        n=float(v)
        if abs(n-round(n)) < 1e-9:
            return str(int(round(n)))
        s=f'{n:,.2f}'.replace(',', ' ').replace('.', ',')
        return s.rstrip('0').rstrip(',')
    except Exception:
        return str(v or '0')


app.jinja_env.filters['money'] = fmt_money
app.jinja_env.filters['qty'] = fmt_qty


@app.get('/api/item-info')
def api_item_info():
    if not session.get('user_id'):
        return {'ok': False}, 401
    kind = (request.args.get('kind') or '').strip()
    item_id = request.args.get('id', type=int)
    if not item_id or kind not in {'raw', 'finished'}:
        return {'ok': False}, 400
    if kind == 'raw':
        row = query_db('SELECT id, name, unit, stock_qty, sale_price, avg_cost FROM raw_materials WHERE id = ?', (item_id,), one=True)
        if not row:
            return {'ok': False}, 404
        return {'ok': True, 'item': {'unit': row['unit'], 'stock_qty': float(row['stock_qty']), 'sale_price': float(row['sale_price']), 'avg_cost': float(row['avg_cost'])}}
    row = query_db('SELECT id, name, default_unit AS unit, stock_qty, sale_price, avg_cost FROM finished_products WHERE id = ?', (item_id,), one=True)
    if not row:
        return {'ok': False}, 404
    return {'ok': True, 'item': {'unit': row['unit'], 'stock_qty': float(row['stock_qty']), 'sale_price': float(row['sale_price']), 'avg_cost': float(row['avg_cost'])}}


@app.get('/api/recipe/<int:recipe_id>')
@login_required
def api_recipe(recipe_id: int):
    recipe = query_db('''
        SELECT sr.id, sr.finished_product_id, sr.name, COALESCE(sr.notes,'') AS notes, fp.name AS finished_name
        FROM saved_recipes sr
        JOIN finished_products fp ON fp.id = sr.finished_product_id
        WHERE sr.id = ?
    ''', (recipe_id,), one=True)
    if not recipe:
        return {'ok': False}, 404
    items = [dict(r) for r in query_db('''
        SELECT sri.raw_material_id, sri.quantity, sri.position,
               rm.name AS material_name, rm.stock_qty, rm.unit
        FROM saved_recipe_items sri
        JOIN raw_materials rm ON rm.id = sri.raw_material_id
        WHERE sri.recipe_id = ?
        ORDER BY sri.position, sri.id
    ''', (recipe_id,))]
    return {'ok': True, 'recipe': {
        'id': int(recipe['id']),
        'finished_product_id': int(recipe['finished_product_id']),
        'name': recipe['name'],
        'notes': recipe['notes'],
        'finished_name': recipe['finished_name'],
        'items': [
            {
                'raw_material_id': int(r['raw_material_id']),
                'quantity': float(r['quantity']),
                'material_name': r['material_name'],
                'stock_qty': float(r['stock_qty']),
                'unit': r['unit'],
            } for r in items
        ]
    }}



def refresh_sale_profits_for_item(item_kind: str, item_id: int, avg_cost: float, sale_price: float | None = None) -> None:
    if item_kind == 'raw':
        rows = query_db('SELECT id, quantity, unit, unit_price FROM raw_sales WHERE raw_material_id = ?', (item_id,))
        for r in rows:
            qty = float(r['quantity'])
            kg = qty_to_kg(qty, r['unit'])
            total = qty * float(r['unit_price'])
            cost = kg * avg_cost
            execute_db('UPDATE raw_sales SET cost_price_snapshot = ?, profit_amount = ? WHERE id = ?', (avg_cost, total - cost, r['id']))
    else:
        rows = query_db('SELECT id, quantity, unit, unit_price FROM sales WHERE finished_product_id = ?', (item_id,))
        for r in rows:
            qty = float(r['quantity'])
            kg = qty_to_kg(qty, r['unit'])
            total = qty * float(r['unit_price'])
            cost = kg * avg_cost
            execute_db('UPDATE sales SET cost_price_snapshot = ?, profit_amount = ? WHERE id = ?', (avg_cost, total - cost, r['id']))


def table_columns(conn, table: str) -> set[str]:
    return set(list_columns(conn, table))


def migrate_db(conn) -> None:
    conn.execute("CREATE TABLE IF NOT EXISTS schema_migrations (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE, applied_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP)")

    def has_table(name: str) -> bool:
        if getattr(conn, 'dialect', 'sqlite') == 'postgres':
            row = conn.execute(
                "SELECT table_name FROM information_schema.tables WHERE table_schema = 'public' AND table_name = ?",
                (name,),
            ).fetchone()
        else:
            row = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (name,)).fetchone()
        return row is not None

    def add_column_if_missing(table: str, column: str, ddl: str) -> None:
        if has_table(table) and column not in table_columns(conn, table):
            conn.execute(ddl)

    def migrate_users_roles() -> None:
        if not has_table('users'):
            return
        columns = table_columns(conn, 'users')
        if getattr(conn, 'dialect', 'sqlite') == 'sqlite':
            schema_row = conn.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name=?", ('users',)).fetchone()
            schema_sql = (schema_row[0] if schema_row else '') or ''
            if all(column in columns for column in {'is_active', 'last_login_at', 'last_password_change_at'}) and 'manager' in schema_sql and 'operator' in schema_sql:
                conn.execute("UPDATE users SET role = CASE WHEN lower(COALESCE(role, '')) = 'admin' THEN 'admin' WHEN lower(COALESCE(role, '')) = 'manager' THEN 'manager' ELSE 'operator' END")
                return
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS users_v11 (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT NOT NULL UNIQUE,
                    password_hash TEXT NOT NULL,
                    role TEXT NOT NULL DEFAULT 'operator' CHECK(role IN ('admin','manager','operator')),
                    must_change_password INTEGER NOT NULL DEFAULT 0,
                    is_active INTEGER NOT NULL DEFAULT 1,
                    last_login_at TEXT,
                    last_password_change_at TEXT,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            conn.execute(
                """
                INSERT OR REPLACE INTO users_v11 (
                    id, username, password_hash, role, must_change_password, is_active, last_login_at, last_password_change_at, created_at
                )
                SELECT
                    id,
                    username,
                    password_hash,
                    CASE
                        WHEN lower(COALESCE(role, '')) = 'admin' THEN 'admin'
                        WHEN lower(COALESCE(role, '')) = 'manager' THEN 'manager'
                        ELSE 'operator'
                    END,
                    COALESCE(must_change_password, 0),
                    COALESCE(is_active, 1),
                    last_login_at,
                    last_password_change_at,
                    COALESCE(created_at, CURRENT_TIMESTAMP)
                FROM users
                """
            )
            conn.execute("DROP TABLE users")
            conn.execute("ALTER TABLE users_v11 RENAME TO users")
            return
        if columns:
            try:
                conn.execute("ALTER TABLE users DROP CONSTRAINT IF EXISTS users_role_check")
            except Exception:
                pass
            try:
                conn.execute("ALTER TABLE users ADD CONSTRAINT users_role_check CHECK (role IN ('admin','manager','operator'))")
            except Exception:
                pass
            conn.execute("UPDATE users SET role = CASE WHEN lower(role) = 'admin' THEN 'admin' WHEN lower(role) = 'manager' THEN 'manager' ELSE 'operator' END")

    add_column_if_missing('clients', 'opening_credit', "ALTER TABLE clients ADD COLUMN opening_credit REAL NOT NULL DEFAULT 0")
    add_column_if_missing('users', 'must_change_password', "ALTER TABLE users ADD COLUMN must_change_password INTEGER NOT NULL DEFAULT 0")
    add_column_if_missing('users', 'is_active', "ALTER TABLE users ADD COLUMN is_active INTEGER NOT NULL DEFAULT 1")
    add_column_if_missing('users', 'last_login_at', "ALTER TABLE users ADD COLUMN last_login_at TEXT")
    add_column_if_missing('users', 'last_password_change_at', "ALTER TABLE users ADD COLUMN last_password_change_at TEXT")
    add_column_if_missing('raw_materials', 'sale_price', "ALTER TABLE raw_materials ADD COLUMN sale_price REAL NOT NULL DEFAULT 0")
    add_column_if_missing('raw_materials', 'alert_threshold', "ALTER TABLE raw_materials ADD COLUMN alert_threshold REAL NOT NULL DEFAULT 0")
    add_column_if_missing('raw_materials', 'threshold_qty', "ALTER TABLE raw_materials ADD COLUMN threshold_qty REAL NOT NULL DEFAULT 0")
    add_column_if_missing('finished_products', 'avg_cost', "ALTER TABLE finished_products ADD COLUMN avg_cost REAL NOT NULL DEFAULT 0")
    add_column_if_missing('sales', 'cost_price_snapshot', "ALTER TABLE sales ADD COLUMN cost_price_snapshot REAL NOT NULL DEFAULT 0")
    add_column_if_missing('sales', 'profit_amount', "ALTER TABLE sales ADD COLUMN profit_amount REAL NOT NULL DEFAULT 0")
    add_column_if_missing('sales', 'document_id', "ALTER TABLE sales ADD COLUMN document_id INTEGER")
    add_column_if_missing('raw_sales', 'cost_price_snapshot', "ALTER TABLE raw_sales ADD COLUMN cost_price_snapshot REAL NOT NULL DEFAULT 0")
    add_column_if_missing('raw_sales', 'profit_amount', "ALTER TABLE raw_sales ADD COLUMN profit_amount REAL NOT NULL DEFAULT 0")
    add_column_if_missing('raw_sales', 'document_id', "ALTER TABLE raw_sales ADD COLUMN document_id INTEGER")
    add_column_if_missing('raw_sales', 'custom_item_name', "ALTER TABLE raw_sales ADD COLUMN custom_item_name TEXT")
    add_column_if_missing('purchases', 'unit', "ALTER TABLE purchases ADD COLUMN unit TEXT NOT NULL DEFAULT 'kg'")
    add_column_if_missing('purchases', 'document_id', "ALTER TABLE purchases ADD COLUMN document_id INTEGER")
    add_column_if_missing('purchases', 'custom_item_name', "ALTER TABLE purchases ADD COLUMN custom_item_name TEXT")
    add_column_if_missing('payments', 'raw_sale_id', "ALTER TABLE payments ADD COLUMN raw_sale_id INTEGER")
    add_column_if_missing('payments', 'sale_kind', "ALTER TABLE payments ADD COLUMN sale_kind TEXT")
    add_column_if_missing('payments', 'payment_type', "ALTER TABLE payments ADD COLUMN payment_type TEXT NOT NULL DEFAULT 'versement'")
    add_column_if_missing('payments', 'allocation_meta', "ALTER TABLE payments ADD COLUMN allocation_meta TEXT")

    if has_table('raw_materials'):
        cols = table_columns(conn, 'raw_materials')
        if 'threshold_qty' in cols and 'alert_threshold' in cols:
            conn.execute("UPDATE raw_materials SET alert_threshold = COALESCE(NULLIF(alert_threshold,0), threshold_qty, 0)")
            conn.execute("UPDATE raw_materials SET threshold_qty = COALESCE(NULLIF(threshold_qty,0), alert_threshold, 0)")

    migrate_users_roles()

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS purchase_documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            supplier_id INTEGER,
            total REAL NOT NULL DEFAULT 0,
            purchase_date TEXT NOT NULL,
            notes TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS sale_documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER,
            sale_type TEXT NOT NULL DEFAULT 'cash',
            total REAL NOT NULL DEFAULT 0,
            amount_paid REAL NOT NULL DEFAULT 0,
            balance_due REAL NOT NULL DEFAULT 0,
            sale_date TEXT NOT NULL,
            notes TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            actor_user_id INTEGER,
            actor_username TEXT NOT NULL,
            actor_role TEXT NOT NULL,
            source TEXT NOT NULL DEFAULT 'web',
            action TEXT NOT NULL,
            entity_type TEXT,
            entity_id TEXT,
            status TEXT NOT NULL DEFAULT 'success',
            ip_address TEXT,
            user_agent TEXT,
            request_id TEXT,
            before_json TEXT,
            after_json TEXT,
            meta_json TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (actor_user_id) REFERENCES users(id) ON DELETE SET NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS backup_jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reason TEXT NOT NULL,
            backup_type TEXT NOT NULL DEFAULT 'event',
            local_path TEXT NOT NULL,
            requested_by_user_id INTEGER,
            status TEXT NOT NULL DEFAULT 'pending',
            context_json TEXT,
            cloud_file_id TEXT,
            cloud_file_name TEXT,
            error_message TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            started_at TEXT,
            finished_at TEXT,
            FOREIGN KEY (requested_by_user_id) REFERENCES users(id) ON DELETE SET NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS backup_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id INTEGER NOT NULL,
            status TEXT NOT NULL,
            cloud_file_id TEXT,
            cloud_file_name TEXT,
            details_json TEXT,
            started_at TEXT,
            finished_at TEXT,
            FOREIGN KEY (job_id) REFERENCES backup_jobs(id) ON DELETE CASCADE
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS api_refresh_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            token_hash TEXT NOT NULL UNIQUE,
            token_hint TEXT,
            created_ip TEXT,
            user_agent TEXT,
            expires_at TEXT NOT NULL,
            revoked_at TEXT,
            last_used_at TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        )
        """
    )


    # ── Indexes de performance (v37 fix) ─────────────────────────────────────
    _indexes = [
        "CREATE INDEX IF NOT EXISTS idx_sales_client_id           ON sales(client_id)",
        "CREATE INDEX IF NOT EXISTS idx_sales_sale_date           ON sales(sale_date)",
        "CREATE INDEX IF NOT EXISTS idx_sales_document_id         ON sales(document_id)",
        "CREATE INDEX IF NOT EXISTS idx_sales_client_type         ON sales(client_id, sale_type)",
        "CREATE INDEX IF NOT EXISTS idx_raw_sales_client_id       ON raw_sales(client_id)",
        "CREATE INDEX IF NOT EXISTS idx_raw_sales_sale_date       ON raw_sales(sale_date)",
        "CREATE INDEX IF NOT EXISTS idx_raw_sales_document_id     ON raw_sales(document_id)",
        "CREATE INDEX IF NOT EXISTS idx_raw_sales_client_type     ON raw_sales(client_id, sale_type)",
        "CREATE INDEX IF NOT EXISTS idx_raw_sales_material_date   ON raw_sales(raw_material_id, sale_date)",
        "CREATE INDEX IF NOT EXISTS idx_payments_client_id        ON payments(client_id)",
        "CREATE INDEX IF NOT EXISTS idx_payments_payment_date     ON payments(payment_date)",
        "CREATE INDEX IF NOT EXISTS idx_payments_client_type      ON payments(client_id, payment_type)",
        "CREATE INDEX IF NOT EXISTS idx_payments_sale_id          ON payments(sale_id)",
        "CREATE INDEX IF NOT EXISTS idx_payments_raw_sale_id      ON payments(raw_sale_id)",
        "CREATE INDEX IF NOT EXISTS idx_purchases_raw_material_id ON purchases(raw_material_id)",
        "CREATE INDEX IF NOT EXISTS idx_purchases_supplier_id     ON purchases(supplier_id)",
        "CREATE INDEX IF NOT EXISTS idx_purchases_purchase_date   ON purchases(purchase_date)",
        "CREATE INDEX IF NOT EXISTS idx_purchases_document_id     ON purchases(document_id)",
        "CREATE INDEX IF NOT EXISTS idx_prod_batch_product_id     ON production_batches(finished_product_id)",
        "CREATE INDEX IF NOT EXISTS idx_prod_batch_date           ON production_batches(production_date)",
        "CREATE INDEX IF NOT EXISTS idx_prod_items_batch_id       ON production_batch_items(batch_id)",
        "CREATE INDEX IF NOT EXISTS idx_prod_items_material_id    ON production_batch_items(raw_material_id)",
        "CREATE INDEX IF NOT EXISTS idx_saved_recipes_product     ON saved_recipes(finished_product_id)",
        "CREATE INDEX IF NOT EXISTS idx_saved_recipe_items_recipe ON saved_recipe_items(recipe_id)",
        "CREATE INDEX IF NOT EXISTS idx_activity_logs_action      ON activity_logs(action)",
        "CREATE INDEX IF NOT EXISTS idx_activity_logs_username    ON activity_logs(username)",
        "CREATE INDEX IF NOT EXISTS idx_users_role                ON users(role)",
        "CREATE INDEX IF NOT EXISTS idx_audit_logs_created_at     ON audit_logs(created_at)",
        "CREATE INDEX IF NOT EXISTS idx_audit_logs_actor          ON audit_logs(actor_username)",
        "CREATE INDEX IF NOT EXISTS idx_audit_logs_action         ON audit_logs(action)",
        "CREATE INDEX IF NOT EXISTS idx_backup_jobs_status        ON backup_jobs(status)",
        "CREATE INDEX IF NOT EXISTS idx_backup_runs_job           ON backup_runs(job_id)",
        "CREATE INDEX IF NOT EXISTS idx_api_refresh_tokens_user   ON api_refresh_tokens(user_id)",
    ]
    for _ddl in _indexes:
        try:
            conn.execute(_ddl)
        except Exception:
            pass

    conn.execute("CREATE TABLE IF NOT EXISTS imported_client_history (id INTEGER PRIMARY KEY AUTOINCREMENT, client_id INTEGER NOT NULL, source_file TEXT, entry_date TEXT NOT NULL, designation TEXT, debit_amount REAL NOT NULL DEFAULT 0, credit_amount REAL NOT NULL DEFAULT 0, running_balance REAL NOT NULL DEFAULT 0, imported_by_user_id INTEGER, created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP, FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE CASCADE, FOREIGN KEY (imported_by_user_id) REFERENCES users(id) ON DELETE SET NULL)")
    conn.execute("INSERT INTO schema_migrations (name) VALUES (?) ON CONFLICT(name) DO NOTHING", ('baseline_v21',))
    try:
        conn.execute("DELETE FROM app_settings WHERE key IN ('ai_provider', 'ai_base_url', 'ai_api_key', 'ai_model')")
    except Exception:
        pass
    try:
        conn.execute("DROP TABLE IF EXISTS ia_chat_history")
    except Exception:
        pass
    try:
        conn.execute("UPDATE users SET role = 'operator' WHERE lower(COALESCE(role, '')) = 'user'")
    except Exception:
        pass
    conn.execute("INSERT INTO schema_migrations (name) VALUES (?) ON CONFLICT(name) DO NOTHING", ('remove_ai_v38',))
    conn.execute("INSERT INTO schema_migrations (name) VALUES (?) ON CONFLICT(name) DO NOTHING", ('indexes_v37fix',))
    conn.execute("INSERT INTO schema_migrations (name) VALUES (?) ON CONFLICT(name) DO NOTHING", ('multiuser_audit_api_v41',))
    conn.commit()


def init_db() -> None:
    conn = connect_database(DATABASE_URL, DB_PATH)
    conn.executescript(
        '''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'operator' CHECK(role IN ('admin','manager','operator')),
            must_change_password INTEGER NOT NULL DEFAULT 0,
            is_active INTEGER NOT NULL DEFAULT 1,
            last_login_at TEXT,
            last_password_change_at TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT,
            address TEXT,
            notes TEXT,
            opening_credit REAL NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS suppliers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT,
            address TEXT,
            notes TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS raw_materials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            unit TEXT NOT NULL DEFAULT 'kg',
            stock_qty REAL NOT NULL DEFAULT 0,
            avg_cost REAL NOT NULL DEFAULT 0,
            sale_price REAL NOT NULL DEFAULT 0,
            alert_threshold REAL NOT NULL DEFAULT 0,
            threshold_qty REAL NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS finished_products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            default_unit TEXT NOT NULL DEFAULT 'kg',
            stock_qty REAL NOT NULL DEFAULT 0,
            sale_price REAL NOT NULL DEFAULT 0,
            avg_cost REAL NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS purchase_documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            supplier_id INTEGER,
            total REAL NOT NULL DEFAULT 0,
            purchase_date TEXT NOT NULL,
            notes TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS sale_documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER,
            sale_type TEXT NOT NULL CHECK(sale_type IN ('cash','credit')),
            total REAL NOT NULL DEFAULT 0,
            amount_paid REAL NOT NULL DEFAULT 0,
            balance_due REAL NOT NULL DEFAULT 0,
            sale_date TEXT NOT NULL,
            notes TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS purchases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            supplier_id INTEGER,
            document_id INTEGER,
            raw_material_id INTEGER NOT NULL,
            quantity REAL NOT NULL,
            unit TEXT NOT NULL DEFAULT 'kg',
            unit_price REAL NOT NULL,
            total REAL NOT NULL,
            purchase_date TEXT NOT NULL,
            notes TEXT,
            custom_item_name TEXT,
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE SET NULL,
            FOREIGN KEY (raw_material_id) REFERENCES raw_materials(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS production_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            finished_product_id INTEGER NOT NULL,
            output_quantity REAL NOT NULL,
            production_cost REAL NOT NULL DEFAULT 0,
            unit_cost REAL NOT NULL DEFAULT 0,
            production_date TEXT NOT NULL,
            notes TEXT,
            FOREIGN KEY (finished_product_id) REFERENCES finished_products(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS production_batch_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id INTEGER NOT NULL,
            raw_material_id INTEGER NOT NULL,
            quantity REAL NOT NULL,
            unit_cost_snapshot REAL NOT NULL DEFAULT 0,
            line_cost REAL NOT NULL DEFAULT 0,
            FOREIGN KEY (batch_id) REFERENCES production_batches(id) ON DELETE CASCADE,
            FOREIGN KEY (raw_material_id) REFERENCES raw_materials(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS saved_recipes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            finished_product_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            notes TEXT,
            created_by_user_id INTEGER,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (finished_product_id) REFERENCES finished_products(id) ON DELETE CASCADE,
            FOREIGN KEY (created_by_user_id) REFERENCES users(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS saved_recipe_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            recipe_id INTEGER NOT NULL,
            raw_material_id INTEGER NOT NULL,
            quantity REAL NOT NULL,
            position INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY (recipe_id) REFERENCES saved_recipes(id) ON DELETE CASCADE,
            FOREIGN KEY (raw_material_id) REFERENCES raw_materials(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER,
            document_id INTEGER,
            finished_product_id INTEGER NOT NULL,
            quantity REAL NOT NULL,
            unit TEXT NOT NULL,
            unit_price REAL NOT NULL,
            total REAL NOT NULL,
            sale_type TEXT NOT NULL CHECK(sale_type IN ('cash','credit')),
            amount_paid REAL NOT NULL DEFAULT 0,
            balance_due REAL NOT NULL DEFAULT 0,
            cost_price_snapshot REAL NOT NULL DEFAULT 0,
            profit_amount REAL NOT NULL DEFAULT 0,
            sale_date TEXT NOT NULL,
            notes TEXT,
            FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE SET NULL,
            FOREIGN KEY (finished_product_id) REFERENCES finished_products(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS raw_sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER,
            document_id INTEGER,
            raw_material_id INTEGER NOT NULL,
            quantity REAL NOT NULL,
            unit TEXT NOT NULL,
            unit_price REAL NOT NULL,
            total REAL NOT NULL,
            sale_type TEXT NOT NULL CHECK(sale_type IN ('cash','credit')),
            amount_paid REAL NOT NULL DEFAULT 0,
            balance_due REAL NOT NULL DEFAULT 0,
            cost_price_snapshot REAL NOT NULL DEFAULT 0,
            profit_amount REAL NOT NULL DEFAULT 0,
            sale_date TEXT NOT NULL,
            notes TEXT,
            custom_item_name TEXT,
            FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE SET NULL,
            FOREIGN KEY (raw_material_id) REFERENCES raw_materials(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER NOT NULL,
            sale_id INTEGER,
            raw_sale_id INTEGER,
            sale_kind TEXT,
            payment_type TEXT NOT NULL DEFAULT 'versement',
            allocation_meta TEXT,
            amount REAL NOT NULL,
            payment_date TEXT NOT NULL,
            notes TEXT,
            FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE CASCADE,
            FOREIGN KEY (sale_id) REFERENCES sales(id) ON DELETE SET NULL,
            FOREIGN KEY (raw_sale_id) REFERENCES raw_sales(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS activity_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            action TEXT NOT NULL,
            entity_type TEXT,
            entity_id INTEGER,
            details TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS error_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            route TEXT,
            error_type TEXT,
            message TEXT,
            traceback TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS system_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            level TEXT NOT NULL,
            message TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            actor_user_id INTEGER,
            actor_username TEXT NOT NULL,
            actor_role TEXT NOT NULL,
            source TEXT NOT NULL DEFAULT 'web',
            action TEXT NOT NULL,
            entity_type TEXT,
            entity_id TEXT,
            status TEXT NOT NULL DEFAULT 'success',
            ip_address TEXT,
            user_agent TEXT,
            request_id TEXT,
            before_json TEXT,
            after_json TEXT,
            meta_json TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (actor_user_id) REFERENCES users(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS backup_jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reason TEXT NOT NULL,
            backup_type TEXT NOT NULL DEFAULT 'event',
            local_path TEXT NOT NULL,
            requested_by_user_id INTEGER,
            status TEXT NOT NULL DEFAULT 'pending',
            context_json TEXT,
            cloud_file_id TEXT,
            cloud_file_name TEXT,
            error_message TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            started_at TEXT,
            finished_at TEXT,
            FOREIGN KEY (requested_by_user_id) REFERENCES users(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS backup_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id INTEGER NOT NULL,
            status TEXT NOT NULL,
            cloud_file_id TEXT,
            cloud_file_name TEXT,
            details_json TEXT,
            started_at TEXT,
            finished_at TEXT,
            FOREIGN KEY (job_id) REFERENCES backup_jobs(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS api_refresh_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            token_hash TEXT NOT NULL UNIQUE,
            token_hint TEXT,
            created_ip TEXT,
            user_agent TEXT,
            expires_at TEXT NOT NULL,
            revoked_at TEXT,
            last_used_at TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS imported_client_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER NOT NULL,
            source_file TEXT,
            entry_date TEXT NOT NULL,
            designation TEXT,
            debit_amount REAL NOT NULL DEFAULT 0,
            credit_amount REAL NOT NULL DEFAULT 0,
            running_balance REAL NOT NULL DEFAULT 0,
            imported_by_user_id INTEGER,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE CASCADE,
            FOREIGN KEY (imported_by_user_id) REFERENCES users(id) ON DELETE SET NULL
        );

        '''
    )

    migrate_db(conn)

    admin_exists = conn.execute('SELECT id FROM users WHERE username = ?', (DEFAULT_ADMIN_USERNAME,)).fetchone()
    if not admin_exists:
        conn.execute(
            'INSERT INTO users (username, password_hash, role, must_change_password, is_active, last_password_change_at) VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)',
            (DEFAULT_ADMIN_USERNAME, generate_password_hash(DEFAULT_ADMIN_PASSWORD), 'admin', 1, 1),
        )
    try:
        conn.execute("INSERT INTO app_settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO NOTHING", ('gdrive_backup_dir', ''))
        conn.execute("INSERT INTO app_settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO NOTHING", ('backup_snapshot_time', '02:00'))
        conn.execute("INSERT INTO app_settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO NOTHING", ('backup_local_retention', '30'))
        conn.execute("INSERT INTO app_settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO NOTHING", ('backup_event_retention', '100'))
        conn.execute("INSERT INTO app_settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO NOTHING", ('backup_last_nightly_date', ''))
    except Exception:
        pass

    other_raw = conn.execute(
        "SELECT id FROM raw_materials WHERE lower(trim(name)) = lower(trim(?)) ORDER BY id LIMIT 1",
        (OTHER_OPERATION_NAME,),
    ).fetchone()
    if other_raw:
        conn.execute(
            "UPDATE raw_materials SET name = ?, unit = ? WHERE id = ?",
            (OTHER_OPERATION_NAME, OTHER_OPERATION_UNIT, int(other_raw["id"])),
        )
    else:
        conn.execute(
            """
            INSERT INTO raw_materials (name, unit, stock_qty, avg_cost, sale_price, alert_threshold, threshold_qty)
            VALUES (?, ?, 0, 0, 0, 0, 0)
            """,
            (OTHER_OPERATION_NAME, OTHER_OPERATION_UNIT),
        )


    conn.commit()
    conn.close()


@app.before_request
def load_current_user() -> None:
    g.request_id = secrets.token_hex(12)
    g.audit_source = 'api' if request.path.startswith('/api/v1/') else 'web'
    start_background_services(app)
    user_id = session.get('user_id')
    g.user = None
    if user_id:
        user = query_db(
            'SELECT id, username, role, is_active, must_change_password, last_login_at, last_password_change_at FROM users WHERE id = ?',
            (user_id,),
            one=True,
        )
        if user and int(user['is_active'] or 0):
            user = dict(user)
            user['role'] = normalize_role(user.get('role'))
            g.user = user
        else:
            session.clear()


@app.before_request
def enforce_endpoint_permission():
    endpoint = request.endpoint
    permission = permission_for_endpoint(endpoint, request.method)
    if not permission:
        return None
    user = getattr(g, 'user', None)
    if user is None:
        return permission_denied_response(permission)
    if not has_permission(user, permission):
        return permission_denied_response(permission)
    return None



def client_balance(client_id: int) -> float:
    row = query_db(
        '''
        SELECT c.opening_credit
             + COALESCE((SELECT SUM(total) FROM sales WHERE client_id = c.id AND sale_type = 'credit'), 0)
             + COALESCE((SELECT SUM(total) FROM raw_sales WHERE client_id = c.id AND sale_type = 'credit'), 0)
             - COALESCE((SELECT SUM(amount) FROM payments WHERE client_id = c.id AND payment_type = 'versement'), 0)
             + COALESCE((SELECT SUM(amount) FROM payments WHERE client_id = c.id AND payment_type = 'avance'), 0) AS balance
        FROM clients c
        WHERE c.id = ?
        ''',
        (client_id,),
        one=True,
    )
    return float(row['balance']) if row else 0.0


def get_open_credit_entries(client_id: int | None = None):
    params: list[Any] = []
    where_sales = 'WHERE s.balance_due > 0'
    where_raw = 'WHERE rs.balance_due > 0'
    if client_id is not None:
        where_sales += ' AND s.client_id = ?'
        where_raw += ' AND rs.client_id = ?'
        params.append(client_id)
        params.append(client_id)
    rows = query_db(
        f'''
        SELECT * FROM (
            SELECT 'finished' AS item_kind, s.id, s.client_id, c.name AS client_name, f.name AS item_name,
                   s.balance_due, s.sale_date, s.total
            FROM sales s
            JOIN clients c ON c.id = s.client_id
            JOIN finished_products f ON f.id = s.finished_product_id
            {where_sales}
            UNION ALL
            SELECT 'raw' AS item_kind, rs.id, rs.client_id, c.name AS client_name, COALESCE(NULLIF(rs.custom_item_name, ''), r.name) AS item_name,
                   rs.balance_due, rs.sale_date, rs.total
            FROM raw_sales rs
            JOIN clients c ON c.id = rs.client_id
            JOIN raw_materials r ON r.id = rs.raw_material_id
            {where_raw}
        ) x
        ORDER BY sale_date ASC, id ASC
        ''',
        tuple(params),
    )
    return rows


def reverse_payment_allocations(payment_row) -> None:
    with db_transaction():
        meta_raw = payment_row['allocation_meta'] if 'allocation_meta' in payment_row.keys() else None
        if meta_raw:
            try:
                allocations = json.loads(meta_raw)
            except Exception:
                allocations = []
            for alloc in allocations:
                kind = alloc.get('kind')
                row_id = int(alloc.get('id'))
                amount = float(alloc.get('amount', 0) or 0)
                if amount <= 0:
                    continue
                if kind == 'finished':
                    doc_row = query_db('SELECT document_id FROM sales WHERE id = ?', (row_id,), one=True)
                    execute_db('UPDATE sales SET amount_paid = amount_paid - ?, balance_due = balance_due + ? WHERE id = ?', (amount, amount, row_id))
                    recalc_sale_document_totals(int(doc_row['document_id'])) if doc_row and doc_row['document_id'] else None
                elif kind == 'raw':
                    doc_row = query_db('SELECT document_id FROM raw_sales WHERE id = ?', (row_id,), one=True)
                    execute_db('UPDATE raw_sales SET amount_paid = amount_paid - ?, balance_due = balance_due + ? WHERE id = ?', (amount, amount, row_id))
                    recalc_sale_document_totals(int(doc_row['document_id'])) if doc_row and doc_row['document_id'] else None
            return
        if payment_row['payment_type'] != 'versement':
            return
        if payment_row['sale_kind'] == 'finished' and payment_row['sale_id']:
            doc_row = query_db('SELECT document_id FROM sales WHERE id = ?', (payment_row['sale_id'],), one=True)
            execute_db('UPDATE sales SET amount_paid = amount_paid - ?, balance_due = balance_due + ? WHERE id = ?', (payment_row['amount'], payment_row['amount'], payment_row['sale_id']))
            recalc_sale_document_totals(int(doc_row['document_id'])) if doc_row and doc_row['document_id'] else None
        elif payment_row['sale_kind'] == 'raw' and payment_row['raw_sale_id']:
            doc_row = query_db('SELECT document_id FROM raw_sales WHERE id = ?', (payment_row['raw_sale_id'],), one=True)
            execute_db('UPDATE raw_sales SET amount_paid = amount_paid - ?, balance_due = balance_due + ? WHERE id = ?', (payment_row['amount'], payment_row['amount'], payment_row['raw_sale_id']))
            recalc_sale_document_totals(int(doc_row['document_id'])) if doc_row and doc_row['document_id'] else None


def apply_payment_to_entry(kind: str, row_id: int, amount: float) -> float:
    if amount <= 0:
        return 0.0
    if kind == 'finished':
        sale = query_db('SELECT balance_due, document_id FROM sales WHERE id = ?', (row_id,), one=True)
        if not sale:
            return 0.0
        paid = min(amount, float(sale['balance_due']))
        execute_db('UPDATE sales SET balance_due = balance_due - ?, amount_paid = amount_paid + ? WHERE id = ?', (paid, paid, row_id))
        recalc_sale_document_totals(int(sale['document_id'])) if sale['document_id'] else None
        return paid
    sale = query_db('SELECT balance_due, document_id FROM raw_sales WHERE id = ?', (row_id,), one=True)
    if not sale:
        return 0.0
    paid = min(amount, float(sale['balance_due']))
    execute_db('UPDATE raw_sales SET balance_due = balance_due - ?, amount_paid = amount_paid + ? WHERE id = ?', (paid, paid, row_id))
    recalc_sale_document_totals(int(sale['document_id'])) if sale['document_id'] else None
    return paid


def allocate_payment_fifo(client_id: int, amount: float) -> float:
    remaining = amount
    for entry in get_open_credit_entries(client_id):
        if remaining <= 0:
            break
        applied = apply_payment_to_entry(entry['item_kind'], entry['id'], remaining)
        remaining -= applied
    return amount - remaining


def reverse_purchase(purchase_id: int) -> bool:
    with db_transaction():
        row = query_db('SELECT * FROM purchases WHERE id = ?', (purchase_id,), one=True)
        if not row:
            return False
        material = query_db('SELECT * FROM raw_materials WHERE id = ?', (row['raw_material_id'],), one=True)
        if not material or float(material['stock_qty']) < float(row['quantity']):
            return False
        execute_db('UPDATE raw_materials SET stock_qty = stock_qty - ? WHERE id = ?', (row['quantity'], row['raw_material_id']))
        execute_db('DELETE FROM purchases WHERE id = ?', (purchase_id,))
        recalc_raw_material_avg_cost(int(row['raw_material_id']))
        recalc_purchase_document_totals(int(row['document_id'])) if row['document_id'] else None
        return True


def reverse_sale(kind: str, row_id: int) -> bool:
    with db_transaction():
        if kind == 'finished':
            row = query_db('SELECT * FROM sales WHERE id = ?', (row_id,), one=True)
            if not row:
                return False
            restore_qty = qty_to_kg(float(row['quantity']), row['unit'])
            execute_db('UPDATE finished_products SET stock_qty = stock_qty + ? WHERE id = ?', (restore_qty, row['finished_product_id']))
            execute_db('DELETE FROM payments WHERE sale_kind = ? AND sale_id = ?', ('finished', row_id))
            execute_db('DELETE FROM sales WHERE id = ?', (row_id,))
            recalc_sale_document_totals(int(row['document_id'])) if row['document_id'] else None
            return True
        row = query_db('SELECT * FROM raw_sales WHERE id = ?', (row_id,), one=True)
        if not row:
            return False
        restore_qty = qty_to_kg(float(row['quantity']), row['unit'])
        execute_db('UPDATE raw_materials SET stock_qty = stock_qty + ? WHERE id = ?', (restore_qty, row['raw_material_id']))
        execute_db('DELETE FROM payments WHERE sale_kind = ? AND raw_sale_id = ?', ('raw', row_id))
        execute_db('DELETE FROM raw_sales WHERE id = ?', (row_id,))
        recalc_sale_document_totals(int(row['document_id'])) if row['document_id'] else None
        return True


def reverse_production(batch_id: int) -> bool:
    with db_transaction():
        batch = query_db('SELECT * FROM production_batches WHERE id = ?', (batch_id,), one=True)
        if not batch:
            return False
        product = query_db('SELECT * FROM finished_products WHERE id = ?', (batch['finished_product_id'],), one=True)
        if not product or float(product['stock_qty']) < float(batch['output_quantity']):
            return False
        items = query_db('SELECT * FROM production_batch_items WHERE batch_id = ?', (batch_id,))
        for item in items:
            execute_db('UPDATE raw_materials SET stock_qty = stock_qty + ? WHERE id = ?', (item['quantity'], item['raw_material_id']))
            recalc_raw_material_avg_cost(int(item['raw_material_id']))
        execute_db('UPDATE finished_products SET stock_qty = stock_qty - ? WHERE id = ?', (batch['output_quantity'], batch['finished_product_id']))
        execute_db('DELETE FROM production_batches WHERE id = ?', (batch_id,))
        recalc_finished_product_avg_cost(int(batch['finished_product_id']))
        return True


def unit_choices() -> list[str]:
    return ['kg', 'sac', 'Qt', 'unite']


OTHER_OPERATION_NAME = 'AUTRE'
OTHER_OPERATION_UNIT = 'unite'


def is_other_operation_name(name: str | None) -> bool:
    return str(name or '').strip().casefold() == OTHER_OPERATION_NAME.casefold()


OTHER_CATEGORY_VALUE = '__other__'
CATALOG_NAME_PRESETS = {
    'raw': ['Ble dur', 'Semoule', 'Farine', 'Mais', 'Son', 'Soja', 'Sel', 'Additif'],
    'finished': ['Aliment BV', 'Aliment OV', 'Aliment AZ', 'Couscous'],
}


def catalog_name_presets(kind: str, current_name: str = '') -> list[str]:
    presets = list(CATALOG_NAME_PRESETS.get(kind, ()))
    seen = {preset.casefold() for preset in presets}
    current_clean = current_name.strip()
    if current_clean and current_clean.casefold() not in seen:
        presets.append(current_clean)
    return presets


def _matching_catalog_preset(presets: list[str], value: str) -> str | None:
    target = value.strip().casefold()
    if not target:
        return None
    for preset in presets:
        if preset.casefold() == target:
            return preset
    return None


def catalog_name_form_context(kind: str, current_name: str = '', form_data=None) -> dict[str, object]:
    presets = catalog_name_presets(kind, current_name=current_name)
    selected_category = OTHER_CATEGORY_VALUE
    custom_name = current_name.strip()
    if form_data is not None:
        selected_category = (form_data.get('category_name') or '').strip()
        custom_name = (form_data.get('custom_name') or '').strip()
        legacy_name = (form_data.get('name') or '').strip()
        if not selected_category:
            preset_match = _matching_catalog_preset(presets, legacy_name)
            if preset_match:
                selected_category = preset_match
                custom_name = ''
            else:
                selected_category = OTHER_CATEGORY_VALUE
                custom_name = legacy_name or custom_name
        elif selected_category != OTHER_CATEGORY_VALUE:
            selected_category = _matching_catalog_preset(presets, selected_category) or selected_category
    else:
        preset_match = _matching_catalog_preset(presets, current_name)
        if preset_match:
            selected_category = preset_match
            custom_name = ''
    return {
        'name_presets': presets,
        'selected_category': selected_category,
        'custom_name_value': custom_name,
        'other_category_value': OTHER_CATEGORY_VALUE,
    }


def resolve_catalog_item_name(form_data) -> str:
    selected_category = (form_data.get('category_name') or '').strip()
    custom_name = (form_data.get('custom_name') or '').strip()
    legacy_name = (form_data.get('name') or '').strip()
    if selected_category and selected_category != OTHER_CATEGORY_VALUE:
        return selected_category
    resolved_name = custom_name or legacy_name
    if not resolved_name:
        raise ValueError("Le nom du produit est obligatoire.")
    return resolved_name


def smart_profit_for_sale(item_kind: str, item_id: int, qty_kg: float, total: float) -> tuple[float, float]:
    """Return (cost_snapshot_per_kg, profit_amount) using moving weighted cost."""
    if item_kind == 'finished':
        item = query_db('SELECT avg_cost FROM finished_products WHERE id = ?', (item_id,), one=True)
    else:
        item = query_db('SELECT avg_cost FROM raw_materials WHERE id = ?', (item_id,), one=True)
    cost_snapshot = float(item['avg_cost']) if item else 0.0
    cogs = qty_kg * cost_snapshot
    return cost_snapshot, round(total - cogs, 2)


@app.route('/sw.js')
def service_worker():
    return send_from_directory(BASE_DIR / 'static', 'sw.js', mimetype='application/javascript')

@app.route('/clients', methods=['GET', 'POST'])
@login_required
def clients():
    if request.method == 'POST':
        execute_db(
            'INSERT INTO clients (name, phone, address, notes, opening_credit) VALUES (?, ?, ?, ?, ?)',
            (
                request.form['name'].strip(),
                request.form.get('phone', '').strip(),
                request.form.get('address', '').strip(),
                request.form.get('notes', '').strip(),
                to_float(request.form.get('opening_credit')),
            ),
        )
        log_activity('create_client', 'client', None, request.form['name'].strip()); backup_database('create_client'); flash('Client ajouté avec succès.', 'success')
        return redirect(url_for('clients'))

    rows = query_db(
        '''
        SELECT c.*, 
               c.opening_credit
               + COALESCE((SELECT SUM(total) FROM sales s WHERE s.client_id = c.id AND s.sale_type = 'credit'), 0)
               + COALESCE((SELECT SUM(total) FROM raw_sales rs WHERE rs.client_id = c.id AND rs.sale_type = 'credit'), 0)
               - COALESCE((SELECT SUM(amount) FROM payments p WHERE p.client_id = c.id AND p.payment_type='versement'), 0)
               + COALESCE((SELECT SUM(amount) FROM payments p WHERE p.client_id = c.id AND p.payment_type='avance'), 0) AS current_debt,
               COALESCE((SELECT SUM(total) FROM sales s WHERE s.client_id = c.id), 0)
               + COALESCE((SELECT SUM(total) FROM raw_sales rs WHERE rs.client_id = c.id), 0) AS total_sales,
               COALESCE((SELECT SUM(amount) FROM payments p WHERE p.client_id = c.id AND p.payment_type='versement'), 0) AS total_payments
        FROM clients c
        ORDER BY c.name
        '''
    )
    return render_template('clients.html', clients=rows)




@login_required
def notes_page():
    if request.method == 'POST':
        action = request.form.get('action', 'save')
        if action == 'restore':
            filename = request.form.get('version_file', '')
            if filename:
                old_content = read_notes_version(filename)
                if old_content:
                    write_app_notes(old_content)
                    flash('Version restaurée avec succès.', 'success')
                else:
                    flash('Version introuvable.', 'danger')
        else:
            content = request.form.get('content', '')
            write_app_notes(content)
            flash('Bloc-note enregistré.', 'success')
        return redirect(url_for('notes_page'))

    view_version = request.args.get('v', '')
    viewing_content = read_notes_version(view_version) if view_version else read_app_notes()
    path = notes_file_path()
    updated_at = datetime.fromtimestamp(path.stat().st_mtime).strftime('%d/%m/%Y %H:%M') if path.exists() else None
    history = list_notes_history()
    return render_template('notes.html',
        content=viewing_content,
        current_content=read_app_notes(),
        updated_at=updated_at,
        history=history,
        view_version=view_version)


@login_required
def pdf_reader():
    ensure_runtime_dirs()
    if request.method == 'POST':
        action = request.form.get('action', 'upload')
        if action == 'delete':
            fname = secure_filename(request.form.get('filename', ''))
            if fname:
                target = PDF_READER_DIR / fname
                if target.exists():
                    target.unlink()
                    flash(f'PDF supprimé : {fname}', 'success')
                else:
                    flash('Fichier introuvable.', 'warning')
            return redirect(url_for('pdf_reader'))
        # upload
        uploaded = request.files.get('pdf_file')
        if not uploaded or not uploaded.filename:
            flash('Choisis un fichier PDF.', 'warning')
            return redirect(url_for('pdf_reader'))
        filename = secure_filename(uploaded.filename)
        if not filename.lower().endswith('.pdf'):
            flash('Seuls les fichiers PDF sont acceptés.', 'danger')
            return redirect(url_for('pdf_reader'))
        target = PDF_READER_DIR / filename
        uploaded.save(target)
        flash(f'PDF ajouté : {filename}', 'success')
        return redirect(url_for('pdf_reader', file=filename))
    files = sorted([p.name for p in PDF_READER_DIR.glob('*.pdf')], key=str.lower)
    selected = request.args.get('file', '').strip()
    if selected and selected not in files:
        selected = ''
    return render_template('pdf_reader.html', files=files, selected=selected)


@login_required
def pdf_reader_file(filename: str):
    safe_name = secure_filename(filename)
    path = PDF_READER_DIR / safe_name
    if not path.exists():
        abort(404)
    return send_from_directory(PDF_READER_DIR, safe_name, mimetype='application/pdf', as_attachment=False)


@app.route('/clients/import-excel', methods=['GET', 'POST'])
@login_required
def import_clients_excel():
    ensure_runtime_dirs()
    if request.method == 'POST':
        files = request.files.getlist('excel_files')
        if not files:
            flash('Ajoute au moins un fichier Excel.', 'warning')
            return redirect(url_for('import_clients_excel'))

        created = 0
        updated = 0
        errors = []
        for uploaded in files:
            if not uploaded or not uploaded.filename:
                continue
            filename = secure_filename(uploaded.filename)
            if not filename.lower().endswith(('.xlsx', '.xlsm')):
                continue
            temp_path = IMPORT_DIR / filename
            try:
                uploaded.save(temp_path)
                parsed = parse_excel_client_file(temp_path)
                last = parse_excel_client_history(temp_path)

                # Seul le dernier "reste à payer" est repris comme dette d'ouverture
                opening = last['last_balance'] if last['last_balance'] > 0 else parsed['opening_credit']

                existing = query_db('SELECT id FROM clients WHERE lower(trim(name)) = lower(trim(?))', (parsed['name'],), one=True)
                if existing:
                    client_id = int(existing['id'])
                    execute_db(
                        """UPDATE clients
                           SET phone = CASE WHEN COALESCE(phone,'')='' THEN ? ELSE phone END,
                               opening_credit = ?
                           WHERE id = ?""",
                        (parsed['phone'], opening, client_id),
                    )
                    updated += 1
                else:
                    client_id = execute_db(
                        'INSERT INTO clients (name, phone, address, notes, opening_credit) VALUES (?, ?, ?, ?, ?)',
                        (parsed['name'], parsed['phone'], parsed['address'], '', opening),
                    )
                    created += 1

            except Exception as exc:
                errors.append(f"{filename}: {exc}")
            finally:
                try:
                    if temp_path.exists():
                        temp_path.unlink()
                except Exception:
                    pass

        log_activity('import_clients_excel', 'client_import', None, f'{created} créés, {updated} mis à jour')
        backup_database('import_excel')
        if errors:
            for err in errors[:5]:
                flash(err, 'danger')
        flash(f'Import terminé : {created} client(s) créé(s), {updated} mis à jour avec dernier solde.', 'success' if (created or updated) else 'warning')
        return redirect(url_for('clients'))

    return render_template('client_import.html')


@app.route('/clients/<int:client_id>')
@login_required
def client_detail(client_id: int):
    client = query_db('SELECT * FROM clients WHERE id = ?', (client_id,), one=True)
    if not client:
        flash('Client introuvable.', 'danger')
        return redirect(url_for('clients'))

    finished_sales = query_db(
        '''
        SELECT s.id AS row_id, s.sale_date AS event_date, f.name || ' - ' || printf('%.2f', s.quantity) || ' ' || s.unit AS designation, s.total AS purchase_amount,
               0 AS payment_amount, 'sale_finished' AS event_type
        FROM sales s JOIN finished_products f ON f.id = s.finished_product_id
        WHERE s.client_id = ?
        ''',
        (client_id,),
    )
    raw_sales = query_db(
        '''
        SELECT rs.id AS row_id, rs.sale_date AS event_date, r.name || ' (matière première) - ' || printf('%.2f', rs.quantity) || ' ' || rs.unit AS designation, rs.total AS purchase_amount,
               0 AS payment_amount, 'sale_raw' AS event_type
        FROM raw_sales rs JOIN raw_materials r ON r.id = rs.raw_material_id
        WHERE rs.client_id = ?
        ''',
        (client_id,),
    )
    payments = query_db(
        '''
        SELECT p.id AS row_id, p.payment_date AS event_date,
               CASE
                   WHEN p.sale_kind = 'raw' THEN 'Versement lié à vente matière'
                   WHEN p.sale_kind = 'finished' THEN 'Versement lié à vente produit'
                   ELSE COALESCE(NULLIF(p.notes,''), CASE WHEN p.payment_type='avance' THEN 'Avance client' ELSE 'Versement client' END)
               END AS designation,
               CASE WHEN p.payment_type='avance' THEN p.amount ELSE 0 END AS purchase_amount,
               CASE WHEN p.payment_type='versement' THEN p.amount ELSE 0 END AS payment_amount,
               CASE WHEN p.payment_type='avance' THEN 'advance' ELSE 'payment' END AS event_type
        FROM payments p WHERE p.client_id = ?
        ''',
        (client_id,),
    )

    timeline = []
    if float(client['opening_credit']) > 0:
        timeline.append({'row_id': None, 'event_date': client['created_at'][:10], 'designation': 'Crédit initial (reprise Excel)', 'purchase_amount': float(client['opening_credit']), 'payment_amount': 0.0, 'event_type': 'opening'})
    timeline.extend([dict(x) for x in finished_sales])
    timeline.extend([dict(x) for x in raw_sales])
    timeline.extend([dict(x) for x in payments])
    timeline.sort(key=lambda x: (x['event_date'], 0 if x['event_type'] in ('opening', 'sale_finished', 'sale_raw') else 1))

    running = 0.0
    for item in timeline:
        running += float(item.get('purchase_amount', 0) or 0)
        running -= float(item.get('payment_amount', 0) or 0)
        item['running_balance'] = running

    stats = {
        'opening_credit': float(client['opening_credit']),
        'credit_sales_total': sum(float(i['purchase_amount']) for i in timeline if i['event_type'] in ('opening', 'sale_finished', 'sale_raw')), 
        'total_paid': sum(float(i['payment_amount']) for i in timeline if i['event_type'] == 'payment'),
    }
    balance = client_balance(client_id)
    return render_template('client_detail.html', client=client, timeline=timeline, stats=stats, balance=balance)


@app.route('/clients/<int:client_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_client(client_id: int):
    client = query_db('SELECT * FROM clients WHERE id = ?', (client_id,), one=True)
    if not client:
        flash('Client introuvable.', 'danger')
        return redirect(url_for('clients'))
    if request.method == 'POST':
        execute_db(
            'UPDATE clients SET name=?, phone=?, address=?, notes=?, opening_credit=? WHERE id=?',
            (
                request.form['name'].strip(),
                request.form.get('phone', '').strip(),
                request.form.get('address', '').strip(),
                request.form.get('notes', '').strip(),
                to_float(request.form.get('opening_credit')),
                client_id,
            ),
        )
        flash('Client modifié avec succès.', 'success')
        return redirect(url_for('client_detail', client_id=client_id))
    return render_template('client_edit.html', client=client)


@app.route('/suppliers', methods=['GET', 'POST'])
@login_required
def suppliers():
    if request.method == 'POST':
        execute_db(
            'INSERT INTO suppliers (name, phone, address, notes) VALUES (?, ?, ?, ?)',
            (
                request.form['name'].strip(),
                request.form.get('phone', '').strip(),
                request.form.get('address', '').strip(),
                request.form.get('notes', '').strip(),
            ),
        )
        log_activity('create_supplier', 'supplier', None, request.form['name'].strip()); backup_database('create_supplier'); flash('Fournisseur ajouté avec succès.', 'success')
        return redirect(url_for('suppliers'))
    return render_template('suppliers.html', suppliers=query_db('SELECT * FROM suppliers ORDER BY name'))


@app.route('/suppliers/<int:supplier_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_supplier(supplier_id: int):
    supplier = query_db('SELECT * FROM suppliers WHERE id = ?', (supplier_id,), one=True)
    if not supplier:
        flash('Fournisseur introuvable.', 'danger')
        return redirect(url_for('suppliers'))
    if request.method == 'POST':
        execute_db('UPDATE suppliers SET name=?, phone=?, address=?, notes=? WHERE id=?', (
            request.form['name'].strip(), request.form.get('phone', '').strip(), request.form.get('address', '').strip(), request.form.get('notes', '').strip(), supplier_id))
        log_activity('update_supplier', 'supplier', supplier_id, request.form['name'].strip()); backup_database('update_supplier'); flash('Fournisseur modifié.', 'success')
        return redirect(url_for('suppliers'))
    return render_template('supplier_edit.html', supplier=supplier)


@app.route('/raw-materials', methods=['GET', 'POST'])
@login_required
def raw_materials():
    return redirect(url_for('catalog'))


@app.route('/raw-materials/<int:material_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_raw_material(material_id: int):
    material = query_db('SELECT * FROM raw_materials WHERE id = ?', (material_id,), one=True)
    if not material:
        flash('Matière introuvable.', 'danger')
        return redirect(url_for('raw_materials'))
    if request.method == 'POST':
        item_name = resolve_catalog_item_name(request.form)
        avg_cost = to_float(request.form.get('avg_cost'))
        sale_price = to_float(request.form.get('sale_price'))
        execute_db('UPDATE raw_materials SET name=?, unit=?, stock_qty=?, avg_cost=?, sale_price=?, alert_threshold=? WHERE id=?', (
            item_name, request.form['unit'].strip(), to_float(request.form.get('stock_qty')), avg_cost, sale_price, to_float(request.form.get('alert_threshold')), material_id))
        refresh_sale_profits_for_item('raw', material_id, avg_cost, sale_price)
        log_activity('update_price', 'raw_material', material_id, f"{item_name} | achat={avg_cost} | vente={sale_price}"); backup_database('update_raw_material'); flash('Matière première modifiée.', 'success')
        return redirect(url_for('raw_materials'))
    return render_template(
        'raw_material_edit.html',
        material=material,
        units=unit_choices(),
        **catalog_name_form_context('raw', current_name=str(material['name'] or '')),
    )


@app.route('/products', methods=['GET', 'POST'])
@login_required
def products():
    return redirect(url_for('catalog'))


@app.route('/products/<int:product_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_product(product_id: int):
    product = query_db('SELECT * FROM finished_products WHERE id = ?', (product_id,), one=True)
    if not product:
        flash('Produit introuvable.', 'danger')
        return redirect(url_for('products'))
    if request.method == 'POST':
        item_name = resolve_catalog_item_name(request.form)
        avg_cost = to_float(request.form.get('avg_cost'))
        sale_price = to_float(request.form.get('sale_price'))
        execute_db('UPDATE finished_products SET name=?, default_unit=?, stock_qty=?, sale_price=?, avg_cost=? WHERE id=?', (
            item_name, request.form['default_unit'].strip(), to_float(request.form.get('stock_qty')), sale_price, avg_cost, product_id))
        refresh_sale_profits_for_item('finished', product_id, avg_cost, sale_price)
        log_activity('update_price', 'finished_product', product_id, f"{item_name} | revient={avg_cost} | vente={sale_price}"); backup_database('update_product'); flash('Produit modifié.', 'success')
        return redirect(url_for('products'))
    return render_template(
        'product_edit.html',
        product=product,
        units=unit_choices(),
        **catalog_name_form_context('finished', current_name=str(product['name'] or '')),
    )



@app.route('/quick-add')
@login_required
def quick_add():
    default_target = request.args.get('target', 'client')
    options = [
        ('client', 'Client', url_for('new_client')),
        ('supplier', 'Fournisseur', url_for('new_supplier')),
        ('product_raw', 'Matière première', url_for('new_catalog_item', kind='raw')),
        ('product_finished', 'Produit fini', url_for('new_catalog_item', kind='finished')),
        ('purchase', 'Achat', url_for('new_purchase')),
        ('sale', 'Vente', url_for('new_sale')),
        ('production', 'Production', url_for('new_production')),
        ('payment', 'Versement', url_for('new_payment')),
        ('advance', 'Avance', url_for('new_payment')),
    ]
    return render_template('quick_add.html', options=options, default_target=default_target)


@app.route('/clients/new', methods=['GET', 'POST'])
@login_required
def new_client():
    if request.method == 'POST':
        execute_db(
            'INSERT INTO clients (name, phone, address, notes, opening_credit) VALUES (?, ?, ?, ?, ?)',
            (
                request.form['name'].strip(),
                request.form.get('phone', '').strip(),
                request.form.get('address', '').strip(),
                request.form.get('notes', '').strip(),
                to_float(request.form.get('opening_credit')),
            ),
        )
        log_activity('create_client', 'client', None, request.form['name'].strip()); backup_database('create_client'); flash('Client ajouté avec succès.', 'success')
        return redirect(url_for('clients'))
    return render_template('client_new.html')


@app.route('/suppliers/new', methods=['GET', 'POST'])
@login_required
def new_supplier():
    if request.method == 'POST':
        execute_db(
            'INSERT INTO suppliers (name, phone, address, notes) VALUES (?, ?, ?, ?)',
            (
                request.form['name'].strip(),
                request.form.get('phone', '').strip(),
                request.form.get('address', '').strip(),
                request.form.get('notes', '').strip(),
            ),
        )
        log_activity('create_supplier', 'supplier', None, request.form['name'].strip()); backup_database('create_supplier'); flash('Fournisseur ajouté avec succès.', 'success')
        return redirect(url_for('suppliers'))
    return render_template('supplier_new.html')


@app.route('/catalog/new', methods=['GET', 'POST'])
@login_required
def new_catalog_item():
    kind = request.args.get('kind', 'raw')
    if request.method == 'POST':
        kind = request.form.get('kind', kind)
        item_name = resolve_catalog_item_name(request.form)
        if kind == 'raw':
            execute_db(
                'INSERT INTO raw_materials (name, unit, stock_qty, avg_cost, sale_price, alert_threshold) VALUES (?, ?, ?, ?, ?, ?)',
                (
                    item_name,
                    request.form['unit'].strip(),
                    to_float(request.form.get('stock_qty')),
                    to_float(request.form.get('avg_cost')),
                    to_float(request.form.get('sale_price')),
                    to_float(request.form.get('alert_threshold')),
                ),
            )
            log_activity('create_product', 'raw_material', None, item_name); backup_database('create_raw_material'); flash('Matière première ajoutée.', 'success')
        else:
            execute_db(
                'INSERT INTO finished_products (name, default_unit, stock_qty, sale_price, avg_cost) VALUES (?, ?, ?, ?, ?)',
                (
                    item_name,
                    request.form['unit'].strip(),
                    to_float(request.form.get('stock_qty')),
                    to_float(request.form.get('sale_price')),
                    to_float(request.form.get('avg_cost')),
                ),
            )
            log_activity('create_product', 'finished_product', None, item_name); backup_database('create_finished_product'); flash('Produit fini ajouté.', 'success')
        return redirect(url_for('catalog'))
    return render_template(
        'catalog_new.html',
        kind=kind,
        units=unit_choices(),
        **catalog_name_form_context(kind),
    )


@app.route('/purchases/new', methods=['GET', 'POST'])
@login_required
def new_purchase():
    if request.method == 'POST':
        supplier_id = request.form.get('supplier_id') or None
        raw_id = int(request.form['raw_material_id'])
        qty = to_float(request.form.get('quantity'))
        unit = request.form.get('unit', 'kg').strip()
        unit_price = to_float(request.form.get('unit_price'))
        purchase_date = request.form.get('purchase_date') or date.today().isoformat()
        notes = request.form.get('notes', '').strip()
        if qty <= 0:
            flash('La quantité doit être supérieure à zéro.', 'danger')
            return redirect(url_for('new_purchase'))
        try:
            purchase_id = create_purchase_record(supplier_id, raw_id, qty, unit_price, purchase_date, notes, unit)
        except Exception as exc:
            flash(str(exc), 'danger')
            return redirect(url_for('new_purchase'))
        log_activity('create_purchase', 'purchase', purchase_id, f"matière #{raw_id} qty={qty} {unit}"); backup_database('create_purchase'); flash('Achat enregistré avec succès.', 'success')
        if wants_print_after_submit():
            return redirect(url_for('print_document', doc_type='purchase', item_id=purchase_id))
        return redirect(url_for('purchases'))
    return render_template('purchase_new.html', suppliers=query_db('SELECT * FROM suppliers ORDER BY name'), raw_materials=query_db('SELECT * FROM raw_materials ORDER BY name'), units=unit_choices())


@app.route('/sales/new', methods=['GET', 'POST'])
@login_required
def new_sale():
    if request.method == 'POST':
        client_id = request.form.get('client_id') or None
        item_key = request.form['item_key']
        if ':' not in item_key:
            flash('Article de vente invalide.', 'danger')
            return redirect(url_for('new_sale'))
        item_kind, item_id_str = item_key.split(':', 1)
        item_id = int(item_id_str)
        qty = to_float(request.form.get('quantity'))
        unit = request.form.get('unit', 'kg').strip()
        unit_price = to_float(request.form.get('unit_price'))
        total = qty * unit_price
        sale_type = 'credit' if client_id else 'cash'
        sale_date = request.form.get('sale_date') or date.today().isoformat()
        notes = request.form.get('notes', '').strip()
        try:
            created_kind, created_sale_id = create_sale_record(client_id, item_kind, item_id, qty, unit, unit_price, sale_type, sale_date, notes, 0 if client_id else qty * unit_price)
        except Exception as exc:
            flash(str(exc), 'danger')
            return redirect(url_for('new_sale'))
        created_doc_type = 'sale_finished' if created_kind == 'finished' else 'sale_raw'
        log_activity('create_sale', 'sale', created_sale_id, f"{item_kind} #{item_id} qty={qty} {unit} total={total}"); backup_database('create_sale'); flash('Vente enregistrée avec bénéfice estimé.', 'success')
        if wants_print_after_submit():
            return redirect(url_for('print_document', doc_type=created_doc_type, item_id=created_sale_id))
        return redirect(url_for('sales'))
    sellable_items = []
    for p in query_db('SELECT id, name, default_unit AS unit, stock_qty, sale_price, avg_cost FROM finished_products ORDER BY name'):
        sellable_items.append({'key': f"finished:{p['id']}", 'label': f"{p['name']} · produit fini", 'unit': p['unit'], 'stock_qty': p['stock_qty'], 'sale_price': p['sale_price'], 'avg_cost': p['avg_cost']})
    for r in query_db('SELECT id, name, unit, stock_qty, sale_price, avg_cost FROM raw_materials ORDER BY name'):
        sellable_items.append({'key': f"raw:{r['id']}", 'label': f"{r['name']} · matière première", 'unit': r['unit'], 'stock_qty': r['stock_qty'], 'sale_price': r['sale_price'], 'avg_cost': r['avg_cost']})
    return render_template('sale_new.html', clients=query_db('SELECT * FROM clients ORDER BY name'), sellable_items=sellable_items, units=unit_choices())



def load_saved_recipes() -> list[dict[str, Any]]:
    recipes = [dict(r) for r in query_db('''
        SELECT sr.id, sr.finished_product_id, sr.name, COALESCE(sr.notes,'') AS notes,
               sr.created_at, fp.name AS finished_name
        FROM saved_recipes sr
        JOIN finished_products fp ON fp.id = sr.finished_product_id
        ORDER BY fp.name, sr.name
    ''')]
    if not recipes:
        return []
    item_rows = query_db('''
        SELECT sri.recipe_id, sri.raw_material_id, sri.quantity, sri.position,
               rm.name AS material_name, rm.stock_qty, rm.unit
        FROM saved_recipe_items sri
        JOIN raw_materials rm ON rm.id = sri.raw_material_id
        ORDER BY sri.recipe_id, sri.position, sri.id
    ''')
    grouped: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in item_rows:
        grouped[int(row['recipe_id'])].append({
            'raw_material_id': int(row['raw_material_id']),
            'quantity': float(row['quantity']),
            'material_name': row['material_name'],
            'stock_qty': float(row['stock_qty']),
            'unit': row['unit'],
        })
    for recipe in recipes:
        recipe['items'] = grouped.get(int(recipe['id']), [])
    return recipes


def save_recipe_definition(finished_id: int, recipe_name: str, notes: str, recipe_lines: list[dict[str, Any]], user_id: int | None = None) -> int | None:
    clean_name = (recipe_name or '').strip()
    if not clean_name or not recipe_lines:
        return None
    existing = query_db('SELECT id FROM saved_recipes WHERE finished_product_id = ? AND lower(name) = lower(?)', (finished_id, clean_name), one=True)
    if existing:
        recipe_id = int(existing['id'])
        execute_db('UPDATE saved_recipes SET notes = ?, updated_at = CURRENT_TIMESTAMP, created_by_user_id = COALESCE(created_by_user_id, ?) WHERE id = ?', (notes, user_id, recipe_id))
        execute_db('DELETE FROM saved_recipe_items WHERE recipe_id = ?', (recipe_id,))
    else:
        recipe_id = execute_db('INSERT INTO saved_recipes (finished_product_id, name, notes, created_by_user_id) VALUES (?, ?, ?, ?)', (finished_id, clean_name, notes, user_id))
    for pos, line in enumerate(recipe_lines, start=1):
        execute_db('INSERT INTO saved_recipe_items (recipe_id, raw_material_id, quantity, position) VALUES (?, ?, ?, ?)', (recipe_id, int(line['material']['id']), float(line['qty']), pos))
    return recipe_id


@app.route('/production/new', methods=['GET', 'POST'])
@login_required
def new_production():
    if request.method == 'POST':
        finished_id = int(request.form['finished_product_id'])
        output_qty = to_float(request.form.get('output_quantity'))
        production_date = request.form.get('production_date') or date.today().isoformat()
        notes = request.form.get('notes', '').strip()
        recipe_name = (request.form.get('recipe_name') or '').strip()
        save_recipe_flag = (request.form.get('save_recipe') or '1').strip() not in ('0', 'false', 'off')
        raw_ids = request.form.getlist('raw_material_id[]')
        quantities = request.form.getlist('quantity[]')

        if production_date > date.today().isoformat():
            flash('La date de production ne peut pas être dans le futur.', 'danger')
            return redirect(url_for('new_production'))
        if output_qty <= 0:
            flash('La quantité produite doit être supérieure à zéro.', 'danger')
            return redirect(url_for('new_production'))
        product = query_db('SELECT * FROM finished_products WHERE id = ?', (finished_id,), one=True)
        if not product:
            flash('Produit fini introuvable.', 'danger')
            return redirect(url_for('new_production'))
        recipe_lines = []
        total_cost = 0.0
        total_recipe_qty = 0.0
        for raw_id, qty_str in zip(raw_ids, quantities):
            if not raw_id:
                continue
            qty = to_float(qty_str)
            if qty <= 0:
                continue
            material = query_db('SELECT * FROM raw_materials WHERE id = ?', (int(raw_id),), one=True)
            if not material:
                flash('Une matière première sélectionnée est introuvable.', 'danger')
                return redirect(url_for('new_production'))
            if qty > float(material['stock_qty']):
                flash(f"Stock insuffisant pour {material['name']}.", 'danger')
                return redirect(url_for('new_production'))
            line_cost = qty * float(material['avg_cost'])
            recipe_lines.append({'material': material, 'qty': qty, 'unit_cost': float(material['avg_cost']), 'line_cost': line_cost})
            total_cost += line_cost
            total_recipe_qty += qty
        if not recipe_lines:
            flash('Ajoute au moins une matière première dans la recette.', 'danger')
            return redirect(url_for('new_production'))
        try:
            with db_transaction():
                batch_id = execute_db('INSERT INTO production_batches (finished_product_id, output_quantity, production_cost, unit_cost, production_date, notes) VALUES (?, ?, ?, ?, ?, ?)',
                    (finished_id, output_qty, total_cost, (total_cost / output_qty) if output_qty else 0, production_date, notes))
                for line in recipe_lines:
                    execute_db('INSERT INTO production_batch_items (batch_id, raw_material_id, quantity, unit_cost_snapshot, line_cost) VALUES (?, ?, ?, ?, ?)',
                        (batch_id, int(line['material']['id']), line['qty'], line['unit_cost'], line['line_cost']))
                    new_stock = float(line['material']['stock_qty']) - line['qty']
                    execute_db('UPDATE raw_materials SET stock_qty = ? WHERE id = ?', (new_stock, int(line['material']['id'])))
                current_stock = float(product['stock_qty'])
                current_value = current_stock * float(product['avg_cost'])
                new_value = current_value + total_cost
                new_stock = current_stock + output_qty
                new_avg = (new_value / new_stock) if new_stock > 0 else 0
                sale_price = float(product['sale_price']) if float(product['sale_price']) > 0 else new_avg * 1.15
                execute_db('UPDATE finished_products SET stock_qty = ?, avg_cost = ?, sale_price = ? WHERE id = ?', (new_stock, new_avg, sale_price, finished_id))
                recipe_id = None
                if save_recipe_flag:
                    recipe_id = save_recipe_definition(finished_id, recipe_name or f"Recette {product['name']}", notes, recipe_lines, int(g.user['id']) if getattr(g, 'user', None) else None)
        except Exception as exc:
            flash(str(exc), 'danger')
            return redirect(url_for('new_production'))
        log_activity('create_production', 'production', batch_id, f"produit #{finished_id} sortie={output_qty}kg coût={total_cost}")
        backup_database('create_production')
        remainder = output_qty - total_recipe_qty
        if recipe_id:
            flash(f"Production enregistrée. Recette sauvegardée ({recipe_name or 'Recette ' + product['name']}). Reste théorique: {remainder:.2f} kg.", 'success')
        else:
            flash(f"Production enregistrée avec recette et coût de revient. Reste théorique: {remainder:.2f} kg.", 'success')
        if wants_print_after_submit():
            return redirect(url_for('print_document', doc_type='production', item_id=batch_id))
        return redirect(url_for('production'))
    return render_template(
        'production_new.html',
        raw_materials=query_db('SELECT * FROM raw_materials ORDER BY name'),
        raw_materials_json=[dict(r) for r in query_db('SELECT * FROM raw_materials ORDER BY name')],
        products=query_db('SELECT * FROM finished_products ORDER BY name'),
        recipes=load_saved_recipes(),
    )



@app.route('/payments/new', methods=['GET', 'POST'])
@login_required
def new_payment():
    mode = request.args.get('mode') or request.form.get('payment_type') or 'versement'
    heading = 'Enregistrer une avance' if mode == 'avance' else 'Enregistrer un versement'
    button_label = "Enregistrer l'avance" if mode == 'avance' else 'Enregistrer le versement'
    if request.method == 'POST':
        payment_type = (request.form.get('payment_type') or mode or 'versement').strip() or 'versement'
        client_raw = (request.form.get('client_id') or '').strip()
        if not client_raw:
            flash('Choisis un client.', 'danger')
            return redirect(url_for('new_payment', mode=payment_type))
        client_id = int(client_raw)
        sale_link = request.form.get('sale_link') or ''
        amount = to_float(request.form.get('amount'))
        payment_date = request.form.get('payment_date') or date.today().isoformat()
        notes = request.form.get('notes', '').strip()
        if amount <= 0:
            flash('Le montant doit être supérieur à zéro.', 'danger')
            return redirect(url_for('new_payment', mode=payment_type))
        try:
            payment_id = create_payment_record(client_id, amount, payment_date, notes, sale_link, payment_type)
            log_activity('create_payment', 'payment', payment_id, f"client #{client_id} {payment_type} montant={amount}")
            backup_database('create_payment')
            flash('Avance enregistrée.' if payment_type == 'avance' else 'Versement enregistré.', 'success')
            if wants_print_after_submit():
                return redirect(url_for('print_document', doc_type='payment', item_id=payment_id))
            return redirect(url_for('transactions', type='payment'))
        except Exception as e:
            flash(str(e), 'danger')
            return redirect(url_for('new_payment', mode=payment_type))
    return render_template('payment_new.html',
                           clients=query_db('SELECT * FROM clients ORDER BY name'),
                           open_sales=get_open_credit_entries(),
                           payment_type=mode,
                           heading=heading,
                           button_label=button_label,
                           show_sale_link=(mode != 'avance'))


@app.route('/catalog')
@login_required
def catalog():
    raw_items = query_db("SELECT id, name, unit AS unit, stock_qty, avg_cost, sale_price, 'Matière première' AS kind FROM raw_materials ORDER BY name")
    finished_items = query_db("SELECT id, name, default_unit AS unit, stock_qty, avg_cost, sale_price, 'Produit fini' AS kind FROM finished_products ORDER BY name")
    all_products = sorted([dict(x) for x in raw_items] + [dict(x) for x in finished_items], key=lambda x: (x['kind'], x['name']))
    return render_template('catalog.html', raw_items=raw_items, finished_items=finished_items, all_products=all_products)


@app.route('/purchases', methods=['GET', 'POST'])
@login_required
def purchases():
    if request.method == 'POST':
        supplier_id = request.form.get('supplier_id') or None
        raw_id = int(request.form['raw_material_id'])
        qty = to_float(request.form.get('quantity'))
        unit = request.form.get('unit', 'kg').strip()
        unit_price = to_float(request.form.get('unit_price'))
        purchase_date = request.form.get('purchase_date') or date.today().isoformat()
        notes = request.form.get('notes', '').strip()
        if qty <= 0:
            flash('La quantité doit être supérieure à zéro.', 'danger')
            return redirect(url_for('purchases'))
        try:
            purchase_id = create_purchase_record(supplier_id, raw_id, qty, unit_price, purchase_date, notes, unit)
        except Exception as exc:
            flash(str(exc), 'danger')
            return redirect(url_for('purchases'))
        log_activity('create_purchase', 'purchase', purchase_id, f"matière #{raw_id} qty={qty} {unit}")
        backup_database('create_purchase')
        flash('Achat enregistré et stock mis à jour.', 'success')
        return redirect(url_for('purchases'))

    return render_template('purchases.html', purchases=query_db('''SELECT p.*, s.name AS supplier_name, r.name AS material_name, r.unit AS material_unit FROM purchases p LEFT JOIN suppliers s ON s.id = p.supplier_id JOIN raw_materials r ON r.id = p.raw_material_id ORDER BY p.id DESC'''), suppliers=query_db('SELECT * FROM suppliers ORDER BY name'), raw_materials=query_db('SELECT * FROM raw_materials ORDER BY name'), units=unit_choices())


@app.route('/production', methods=['GET', 'POST'])
@login_required
def production():
    if request.method == 'POST':
        finished_id = int(request.form['finished_product_id'])
        output_qty = to_float(request.form.get('output_quantity'))
        production_date = request.form.get('production_date') or date.today().isoformat()
        notes = request.form.get('notes', '').strip()
        raw_ids = request.form.getlist('raw_material_id[]')
        quantities = request.form.getlist('quantity[]')

        ingredients: list[tuple[int, float, sqlite3.Row]] = []
        total_cost = 0.0
        if output_qty <= 0:
            flash('La quantité produite doit être supérieure à zéro.', 'danger')
            return redirect(url_for('production'))

        for rid, qty_str in zip(raw_ids, quantities):
            if not rid:
                continue
            qty = to_float(qty_str)
            if qty <= 0:
                continue
            raw = query_db('SELECT * FROM raw_materials WHERE id = ?', (int(rid),), one=True)
            if not raw:
                flash('Une matière première est introuvable.', 'danger')
                return redirect(url_for('production'))
            if qty > float(raw['stock_qty']):
                flash(f"Stock insuffisant pour {raw['name']}.", 'danger')
                return redirect(url_for('production'))
            ingredients.append((int(rid), qty, raw))
            total_cost += qty * float(raw['avg_cost'])

        if not ingredients:
            flash('Ajoute au moins une matière première dans la recette.', 'danger')
            return redirect(url_for('production'))

        finished = query_db('SELECT * FROM finished_products WHERE id = ?', (finished_id,), one=True)
        if not finished:
            flash('Produit fini introuvable.', 'danger')
            return redirect(url_for('production'))

        unit_cost = total_cost / output_qty if output_qty else 0
        try:
            with db_transaction():
                batch_id = execute_db('INSERT INTO production_batches (finished_product_id, output_quantity, production_cost, unit_cost, production_date, notes) VALUES (?, ?, ?, ?, ?, ?)',
                    (finished_id, output_qty, total_cost, unit_cost, production_date, notes))

                for rid, qty, raw in ingredients:
                    execute_db('INSERT INTO production_batch_items (batch_id, raw_material_id, quantity, unit_cost_snapshot, line_cost) VALUES (?, ?, ?, ?, ?)',
                        (batch_id, rid, qty, float(raw['avg_cost']), qty * float(raw['avg_cost'])))
                    execute_db('UPDATE raw_materials SET stock_qty = stock_qty - ? WHERE id = ?', (qty, rid))

                current_stock = float(finished['stock_qty'])
                current_value = current_stock * float(finished['avg_cost'])
                new_stock = current_stock + output_qty
                new_avg_cost = (current_value + total_cost) / new_stock if new_stock > 0 else 0
                execute_db('UPDATE finished_products SET stock_qty = ?, avg_cost = ? WHERE id = ?', (new_stock, new_avg_cost, finished_id))
        except Exception as exc:
            flash(str(exc), 'danger')
            return redirect(url_for('production'))
        log_activity('create_production', 'production', batch_id, f"produit #{finished_id} sortie={output_qty}kg coût={total_cost}")
        backup_database('create_production')
        flash('Production multi-matières enregistrée avec coût de revient.', 'success')
        return redirect(url_for('production'))

    batches = query_db(
        '''
        SELECT pb.*, fp.name AS finished_name,
               GROUP_CONCAT(r.name || ' ' || CAST(pbi.quantity AS TEXT) || ' ' || r.unit, ' + ') AS recipe_text
        FROM production_batches pb
        JOIN finished_products fp ON fp.id = pb.finished_product_id
        LEFT JOIN production_batch_items pbi ON pbi.batch_id = pb.id
        LEFT JOIN raw_materials r ON r.id = pbi.raw_material_id
        GROUP BY pb.id
        ORDER BY pb.id DESC
        '''
    )
    return render_template('production.html', productions=batches, raw_materials=query_db('SELECT * FROM raw_materials ORDER BY name'), products=query_db('SELECT * FROM finished_products ORDER BY name'), units=unit_choices())


@app.route('/sales', methods=['GET', 'POST'])
@login_required
def sales():
    if request.method == 'POST':
        client_id = request.form.get('client_id') or None
        item_key = request.form.get('item_key', '')
        qty = to_float(request.form.get('quantity'))
        unit = request.form.get('unit', 'kg').strip()
        unit_price = to_float(request.form.get('unit_price'))
        sale_type = 'credit' if client_id else 'cash'
        sale_date = request.form.get('sale_date') or date.today().isoformat()
        notes = request.form.get('notes', '').strip()
        if ':' not in item_key:
            flash('Article de vente invalide.', 'danger')
            return redirect(url_for('sales'))
        item_kind, item_id_str = item_key.split(':', 1)
        item_id = int(item_id_str)
        total = qty * unit_price
        try:
            created_kind, created_sale_id = create_sale_record(client_id, item_kind, item_id, qty, unit, unit_price, sale_type, sale_date, notes, 0 if client_id else total)
        except Exception as exc:
            flash(str(exc), 'danger')
            return redirect(url_for('sales'))
        created_doc_type = 'sale_finished' if created_kind == 'finished' else 'sale_raw'
        log_activity('create_sale', 'sale', created_sale_id, f"{item_kind} #{item_id} qty={qty} {unit} total={total}"); backup_database('create_sale'); flash('Vente enregistrée avec bénéfice estimé.', 'success')
        if wants_print_after_submit():
            return redirect(url_for('print_document', doc_type=created_doc_type, item_id=created_sale_id))
        return redirect(url_for('sales'))

    sellable_items = []
    for p in query_db('SELECT id, name, default_unit AS unit, stock_qty, sale_price, avg_cost FROM finished_products ORDER BY name'):
        sellable_items.append({'key': f"finished:{p['id']}", 'label': f"{p['name']} · produit fini", 'unit': p['unit'], 'stock_qty': p['stock_qty'], 'sale_price': p['sale_price'], 'avg_cost': p['avg_cost']})
    for r in query_db('SELECT id, name, unit, stock_qty, sale_price, avg_cost FROM raw_materials ORDER BY name'):
        sellable_items.append({'key': f"raw:{r['id']}", 'label': f"{r['name']} · matière première", 'unit': r['unit'], 'stock_qty': r['stock_qty'], 'sale_price': r['sale_price'], 'avg_cost': r['avg_cost']})

    sales_rows = query_db(
        '''
        SELECT * FROM (
            SELECT s.id, s.sale_date, COALESCE(c.name, 'Comptoir') AS client_name, f.name AS item_name, s.quantity, s.unit, s.total, s.amount_paid, s.balance_due, s.profit_amount, 'Produit fini' AS item_kind, 'finished' AS row_kind
            FROM sales s
            LEFT JOIN clients c ON c.id = s.client_id
            JOIN finished_products f ON f.id = s.finished_product_id
            UNION ALL
            SELECT rs.id, rs.sale_date, COALESCE(c.name, 'Comptoir') AS client_name, r.name AS item_name, rs.quantity, rs.unit, rs.total, rs.amount_paid, rs.balance_due, rs.profit_amount, 'Matière première' AS item_kind, 'raw' AS row_kind
            FROM raw_sales rs
            LEFT JOIN clients c ON c.id = rs.client_id
            JOIN raw_materials r ON r.id = rs.raw_material_id
        ) x ORDER BY sale_date DESC
        '''
    )
    return render_template('sales.html', sales=sales_rows, clients=query_db('SELECT * FROM clients ORDER BY name'), sellable_items=sellable_items, sellable_json=json.dumps(sellable_items), units=unit_choices())


@app.route('/payments', methods=['GET', 'POST'])
@login_required
def payments():
    if request.method == 'POST':
        client_raw = (request.form.get('client_id') or '').strip()
        if not client_raw:
            flash('Choisis un client.', 'danger')
            return redirect(request.url)
        client_id = int(client_raw)
        sale_link = request.form.get('sale_link') or ''
        amount = to_float(request.form.get('amount'))
        payment_type = (request.form.get('payment_type') or 'versement').strip().lower()
        if payment_type not in {'versement', 'avance'}:
            payment_type = 'versement'
        payment_date = request.form.get('payment_date') or date.today().isoformat()
        notes = request.form.get('notes', '').strip()

        if amount <= 0:
            flash('Le montant doit être supérieur à zéro.', 'danger')
            return redirect(url_for('payments'))

        try:
            payment_id = create_payment_record(client_id, amount, payment_date, notes, sale_link, payment_type)
        except Exception as exc:
            flash(str(exc), 'danger')
            return redirect(url_for('payments'))

        log_activity('create_payment', 'payment', payment_id, f"client #{client_id} {payment_type} montant={amount}")
        backup_database('create_payment')
        flash('Paiement enregistré.', 'success')
        return redirect(url_for('payments'))

    rows = query_db(
        '''
        SELECT p.id, p.*, c.name AS client_name,
               CASE
                   WHEN p.sale_kind = 'finished' AND p.sale_id IS NOT NULL THEN 'Produit #' || p.sale_id
                   WHEN p.sale_kind = 'raw' AND p.raw_sale_id IS NOT NULL THEN 'Matière #' || p.raw_sale_id
                   ELSE '-'
               END AS sale_ref, p.payment_type
        FROM payments p
        JOIN clients c ON c.id = p.client_id
        ORDER BY p.id DESC
        '''
    )
    open_sales = get_open_credit_entries()
    return render_template('payments.html', payments=rows, clients=query_db('SELECT * FROM clients ORDER BY name'), open_sales=open_sales)




def build_print_payload(doc_type: str, item_id: int):
    if doc_type == 'purchase':
        row = query_db(
            """
            SELECT p.*, rm.name AS item_name, rm.unit AS base_unit, s.name AS partner_name
            FROM purchases p
            JOIN raw_materials rm ON rm.id = p.raw_material_id
            LEFT JOIN suppliers s ON s.id = p.supplier_id
            WHERE p.id = ?
            """,
            (item_id,), one=True
        )
        if not row:
            return None
        return {
            'title': "Bon d'achat",
            'subtitle': 'Achat matière première',
            'number': f"ACH-{row['id']:06d}",
            'date': row['purchase_date'],
            'partner_label': 'Fournisseur',
            'partner_name': row['partner_name'] or 'Non renseigné',
            'item_label': 'Matière',
            'item_name': row['item_name'],
            'quantity': kg_to_display(float(row['quantity']), row['base_unit']),
            'unit': row['base_unit'],
            'unit_price': float(row['unit_price']) * (50 if str(row['base_unit']).lower() == 'sac' else 100 if str(row['base_unit']).lower() in {'qt', 'quintal'} else 1),
            'total': row['total'],
            'paid': None,
            'due': None,
            'notes': row['notes'] or '',
        }
    if doc_type == 'sale_finished':
        row = query_db(
            """
            SELECT s.*, f.name AS item_name, COALESCE(c.name, 'Comptoir') AS partner_name
            FROM sales s
            JOIN finished_products f ON f.id = s.finished_product_id
            LEFT JOIN clients c ON c.id = s.client_id
            WHERE s.id = ?
            """,
            (item_id,), one=True
        )
        if not row:
            return None
        return {
            'title': 'Facture',
            'subtitle': 'Vente produit final',
            'number': f"VPF-{row['id']:06d}",
            'date': row['sale_date'],
            'partner_label': 'Client',
            'partner_name': row['partner_name'],
            'item_label': 'Article',
            'item_name': row['item_name'],
            'quantity': row['quantity'],
            'unit': row['unit'],
            'unit_price': row['unit_price'],
            'total': row['total'],
            'paid': row['amount_paid'],
            'due': row['balance_due'],
            'notes': row['notes'] or '',
        }
    if doc_type == 'sale_raw':
        row = query_db(
            """
            SELECT rs.*, r.name AS item_name, COALESCE(c.name, 'Comptoir') AS partner_name
            FROM raw_sales rs
            JOIN raw_materials r ON r.id = rs.raw_material_id
            LEFT JOIN clients c ON c.id = rs.client_id
            WHERE rs.id = ?
            """,
            (item_id,), one=True
        )
        if not row:
            return None
        return {
            'title': 'Facture',
            'subtitle': 'Vente matière première',
            'number': f"VMP-{row['id']:06d}",
            'date': row['sale_date'],
            'partner_label': 'Client',
            'partner_name': row['partner_name'],
            'item_label': 'Article',
            'item_name': row['item_name'],
            'quantity': row['quantity'],
            'unit': row['unit'],
            'unit_price': row['unit_price'],
            'total': row['total'],
            'paid': row['amount_paid'],
            'due': row['balance_due'],
            'notes': row['notes'] or '',
        }
    if doc_type == 'payment':
        row = query_db(
            """
            SELECT p.*, c.name AS partner_name
            FROM payments p
            JOIN clients c ON c.id = p.client_id
            WHERE p.id = ?
            """,
            (item_id,), one=True
        )
        if not row:
            return None
        label = 'Avance client' if row['payment_type'] == 'avance' else 'Versement client'
        return {
            'title': 'Reçu',
            'subtitle': label,
            'number': f"PAY-{row['id']:06d}",
            'date': row['payment_date'],
            'partner_label': 'Client',
            'partner_name': row['partner_name'],
            'item_label': 'Référence',
            'item_name': label,
            'quantity': None,
            'unit': '',
            'unit_price': None,
            'total': row['amount'],
            'paid': row['amount'],
            'due': 0,
            'notes': row['notes'] or '',
        }
    if doc_type == 'production':
        row = query_db(
            """
            SELECT pb.*, fp.name AS item_name,
                   GROUP_CONCAT(r.name || ' ' || CAST(pbi.quantity AS TEXT) || ' ' || r.unit, ' + ') AS recipe_text
            FROM production_batches pb
            JOIN finished_products fp ON fp.id = pb.finished_product_id
            LEFT JOIN production_batch_items pbi ON pbi.batch_id = pb.id
            LEFT JOIN raw_materials r ON r.id = pbi.raw_material_id
            WHERE pb.id = ?
            GROUP BY pb.id
            """,
            (item_id,), one=True
        )
        if not row:
            return None
        return {
            'title': 'Fiche de production',
            'subtitle': 'Production enregistrée',
            'number': f"PROD-{row['id']:06d}",
            'date': row['production_date'],
            'partner_label': 'Produit fini',
            'partner_name': row['item_name'],
            'item_label': 'Recette',
            'item_name': row['recipe_text'] or '—',
            'quantity': row['output_quantity'],
            'unit': 'kg',
            'unit_price': row['unit_cost'],
            'total': row['production_cost'],
            'paid': None,
            'due': None,
            'notes': row['notes'] or '',
        }
    return None


def _purchase_line_to_doc_line(row) -> dict[str, Any]:
    unit = row['display_unit'] or row['base_unit'] or 'kg'
    return {
        'item_name': row['item_name'],
        'quantity': row['display_quantity'],
        'unit': unit,
        'unit_price': row['display_unit_price'],
        'total': row['total'],
    }


def _sale_line_to_doc_line(row, item_name: str) -> dict[str, Any]:
    return {
        'item_name': item_name,
        'quantity': row['quantity'],
        'unit': row['unit'],
        'unit_price': row['unit_price'],
        'total': row['total'],
    }


def _sale_document_subtitle(lines: list[dict[str, Any]]) -> str:
    kinds = {str(line.get('kind') or '') for line in lines}
    if kinds == {'finished'}:
        return 'Vente produit final'
    if kinds == {'raw'}:
        return 'Vente matière première'
    return 'Vente multi-produits'


def build_print_payload(doc_type: str, item_id: int):
    if doc_type == 'purchase':
        pointer = query_db('SELECT id, document_id FROM purchases WHERE id = ?', (item_id,), one=True)
        if pointer and pointer['document_id']:
            return build_print_payload('purchase_document', int(pointer['document_id']))
        row = query_db(
            """
            SELECT p.*, rm.name AS item_name, rm.unit AS base_unit, COALESCE(p.unit, rm.unit, 'kg') AS display_unit,
                   CASE
                       WHEN lower(COALESCE(p.unit, rm.unit, 'kg')) = 'sac' THEN p.quantity / 50.0
                       WHEN lower(COALESCE(p.unit, rm.unit, 'kg')) IN ('qt', 'quintal') THEN p.quantity / 100.0
                       ELSE p.quantity
                   END AS display_quantity,
                   CASE
                       WHEN lower(COALESCE(p.unit, rm.unit, 'kg')) = 'sac' THEN p.unit_price * 50.0
                       WHEN lower(COALESCE(p.unit, rm.unit, 'kg')) IN ('qt', 'quintal') THEN p.unit_price * 100.0
                       ELSE p.unit_price
                   END AS display_unit_price,
                   s.name AS partner_name
            FROM purchases p
            JOIN raw_materials rm ON rm.id = p.raw_material_id
            LEFT JOIN suppliers s ON s.id = p.supplier_id
            WHERE p.id = ?
            """,
            (item_id,), one=True
        )
        if not row:
            return None
        lines = [_purchase_line_to_doc_line(row)]
        return {
            'title': "Bon d'achat",
            'subtitle': 'Achat matière première',
            'number': f"ACH-{row['id']:06d}",
            'date': row['purchase_date'],
            'partner_label': 'Fournisseur',
            'partner_name': row['partner_name'] or 'Non renseigné',
            'item_label': 'Matière',
            'item_name': row['item_name'],
            'quantity': row['display_quantity'],
            'unit': row['display_unit'],
            'unit_price': row['display_unit_price'],
            'total': row['total'],
            'paid': None,
            'due': None,
            'notes': row['notes'] or '',
            'lines': lines,
        }
    if doc_type == 'purchase_document':
        doc = query_db(
            """
            SELECT pd.*, s.name AS partner_name
            FROM purchase_documents pd
            LEFT JOIN suppliers s ON s.id = pd.supplier_id
            WHERE pd.id = ?
            """,
            (item_id,), one=True
        )
        if not doc:
            return None
        line_rows = query_db(
            """
            SELECT p.*, rm.name AS item_name, rm.unit AS base_unit, COALESCE(p.unit, rm.unit, 'kg') AS display_unit,
                   CASE
                       WHEN lower(COALESCE(p.unit, rm.unit, 'kg')) = 'sac' THEN p.quantity / 50.0
                       WHEN lower(COALESCE(p.unit, rm.unit, 'kg')) IN ('qt', 'quintal') THEN p.quantity / 100.0
                       ELSE p.quantity
                   END AS display_quantity,
                   CASE
                       WHEN lower(COALESCE(p.unit, rm.unit, 'kg')) = 'sac' THEN p.unit_price * 50.0
                       WHEN lower(COALESCE(p.unit, rm.unit, 'kg')) IN ('qt', 'quintal') THEN p.unit_price * 100.0
                       ELSE p.unit_price
                   END AS display_unit_price
            FROM purchases p
            JOIN raw_materials rm ON rm.id = p.raw_material_id
            WHERE p.document_id = ?
            ORDER BY p.id ASC
            """,
            (item_id,),
        )
        if not line_rows:
            return None
        lines = [_purchase_line_to_doc_line(row) for row in line_rows]
        return {
            'title': "Bon d'achat",
            'subtitle': 'Achat multi-produits',
            'number': f"ACH-{doc['id']:06d}",
            'date': doc['purchase_date'],
            'partner_label': 'Fournisseur',
            'partner_name': doc['partner_name'] or 'Non renseigné',
            'item_label': 'Matière',
            'item_name': f"{len(lines)} ligne(s)",
            'quantity': None,
            'unit': '',
            'unit_price': None,
            'total': doc['total'],
            'paid': None,
            'due': None,
            'notes': doc['notes'] or '',
            'lines': lines,
        }
    if doc_type == 'sale_finished':
        pointer = query_db('SELECT id, document_id FROM sales WHERE id = ?', (item_id,), one=True)
        if pointer and pointer['document_id']:
            return build_print_payload('sale_document', int(pointer['document_id']))
        row = query_db(
            """
            SELECT s.*, f.name AS item_name, COALESCE(c.name, 'Comptoir') AS partner_name
            FROM sales s
            JOIN finished_products f ON f.id = s.finished_product_id
            LEFT JOIN clients c ON c.id = s.client_id
            WHERE s.id = ?
            """,
            (item_id,), one=True
        )
        if not row:
            return None
        lines = [_sale_line_to_doc_line(row, row['item_name'])]
        return {
            'title': 'Facture',
            'subtitle': 'Vente produit final',
            'number': f"VPF-{row['id']:06d}",
            'date': row['sale_date'],
            'partner_label': 'Client',
            'partner_name': row['partner_name'],
            'item_label': 'Article',
            'item_name': row['item_name'],
            'quantity': row['quantity'],
            'unit': row['unit'],
            'unit_price': row['unit_price'],
            'total': row['total'],
            'paid': row['amount_paid'],
            'due': row['balance_due'],
            'notes': row['notes'] or '',
            'lines': lines,
        }
    if doc_type == 'sale_raw':
        pointer = query_db('SELECT id, document_id FROM raw_sales WHERE id = ?', (item_id,), one=True)
        if pointer and pointer['document_id']:
            return build_print_payload('sale_document', int(pointer['document_id']))
        row = query_db(
            """
            SELECT rs.*, r.name AS item_name, COALESCE(c.name, 'Comptoir') AS partner_name
            FROM raw_sales rs
            JOIN raw_materials r ON r.id = rs.raw_material_id
            LEFT JOIN clients c ON c.id = rs.client_id
            WHERE rs.id = ?
            """,
            (item_id,), one=True
        )
        if not row:
            return None
        lines = [_sale_line_to_doc_line(row, row['item_name'])]
        return {
            'title': 'Facture',
            'subtitle': 'Vente matière première',
            'number': f"VMP-{row['id']:06d}",
            'date': row['sale_date'],
            'partner_label': 'Client',
            'partner_name': row['partner_name'],
            'item_label': 'Article',
            'item_name': row['item_name'],
            'quantity': row['quantity'],
            'unit': row['unit'],
            'unit_price': row['unit_price'],
            'total': row['total'],
            'paid': row['amount_paid'],
            'due': row['balance_due'],
            'notes': row['notes'] or '',
            'lines': lines,
        }
    if doc_type == 'sale_document':
        doc = query_db(
            """
            SELECT sd.*, COALESCE(c.name, 'Comptoir') AS partner_name
            FROM sale_documents sd
            LEFT JOIN clients c ON c.id = sd.client_id
            WHERE sd.id = ?
            """,
            (item_id,), one=True
        )
        if not doc:
            return None
        line_rows = query_db(
            """
            SELECT * FROM (
                SELECT 'finished' AS kind, s.id AS line_id, s.quantity, s.unit, s.unit_price, s.total, f.name AS item_name
                FROM sales s
                JOIN finished_products f ON f.id = s.finished_product_id
                WHERE s.document_id = ?
                UNION ALL
                SELECT 'raw' AS kind, rs.id AS line_id, rs.quantity, rs.unit, rs.unit_price, rs.total, r.name AS item_name
                FROM raw_sales rs
                JOIN raw_materials r ON r.id = rs.raw_material_id
                WHERE rs.document_id = ?
            ) lines
            ORDER BY line_id ASC
            """,
            (item_id, item_id),
        )
        if not line_rows:
            return None
        lines = [_sale_line_to_doc_line(row, row['item_name']) | {'kind': row['kind']} for row in line_rows]
        subtitle = _sale_document_subtitle(lines)
        clean_lines = [{k: v for k, v in line.items() if k != 'kind'} for line in lines]
        return {
            'title': 'Facture',
            'subtitle': subtitle,
            'number': f"FAC-{doc['id']:06d}",
            'date': doc['sale_date'],
            'partner_label': 'Client',
            'partner_name': doc['partner_name'],
            'item_label': 'Article',
            'item_name': f"{len(clean_lines)} ligne(s)",
            'quantity': None,
            'unit': '',
            'unit_price': None,
            'total': doc['total'],
            'paid': doc['amount_paid'],
            'due': doc['balance_due'],
            'notes': doc['notes'] or '',
            'lines': clean_lines,
        }
    if doc_type == 'payment':
        row = query_db(
            """
            SELECT p.*, c.name AS partner_name
            FROM payments p
            JOIN clients c ON c.id = p.client_id
            WHERE p.id = ?
            """,
            (item_id,), one=True
        )
        if not row:
            return None
        label = 'Avance client' if row['payment_type'] == 'avance' else 'Versement client'
        lines = [{
            'item_name': label,
            'quantity': None,
            'unit': '',
            'unit_price': None,
            'total': row['amount'],
        }]
        return {
            'title': 'Reçu',
            'subtitle': label,
            'number': f"PAY-{row['id']:06d}",
            'date': row['payment_date'],
            'partner_label': 'Client',
            'partner_name': row['partner_name'],
            'item_label': 'Référence',
            'item_name': label,
            'quantity': None,
            'unit': '',
            'unit_price': None,
            'total': row['amount'],
            'paid': row['amount'],
            'due': 0,
            'notes': row['notes'] or '',
            'lines': lines,
        }
    if doc_type == 'production':
        row = query_db(
            """
            SELECT pb.*, fp.name AS item_name,
                   GROUP_CONCAT(r.name || ' ' || CAST(pbi.quantity AS TEXT) || ' ' || r.unit, ' + ') AS recipe_text
            FROM production_batches pb
            JOIN finished_products fp ON fp.id = pb.finished_product_id
            LEFT JOIN production_batch_items pbi ON pbi.batch_id = pb.id
            LEFT JOIN raw_materials r ON r.id = pbi.raw_material_id
            WHERE pb.id = ?
            GROUP BY pb.id
            """,
            (item_id,), one=True
        )
        if not row:
            return None
        lines = [{
            'item_name': row['recipe_text'] or '—',
            'quantity': row['output_quantity'],
            'unit': 'kg',
            'unit_price': row['unit_cost'],
            'total': row['production_cost'],
        }]
        return {
            'title': 'Fiche de production',
            'subtitle': 'Production enregistrée',
            'number': f"PROD-{row['id']:06d}",
            'date': row['production_date'],
            'partner_label': 'Produit fini',
            'partner_name': row['item_name'],
            'item_label': 'Recette',
            'item_name': row['recipe_text'] or '—',
            'quantity': row['output_quantity'],
            'unit': 'kg',
            'unit_price': row['unit_cost'],
            'total': row['production_cost'],
            'paid': None,
            'due': None,
            'notes': row['notes'] or '',
            'lines': lines,
        }
    return None


def _generate_invoice_pdf_legacy_old(doc: dict, printed_by: str) -> BytesIO | None:
    """Generate a PDF invoice in memory using ReportLab. Returns None if unavailable."""
    if not REPORTLAB_AVAILABLE:
        return None
    buf = BytesIO()
    page_doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=1.8*cm, bottomMargin=1.8*cm,
        leftMargin=2*cm, rightMargin=2*cm
    )
    styles = getSampleStyleSheet()
    DARK  = colors.HexColor('#1e293b')
    LIGHT = colors.HexColor('#f8fafc')
    MUTED = colors.HexColor('#64748b')
    ACCENT = colors.HexColor('#3b82f6')
    GREEN = colors.HexColor('#16a34a')
    RED   = colors.HexColor('#dc2626')
    story = []

    # ── Header ──────────────────────────────────────────────────────────
    logo_path = str(BASE_DIR / 'static' / 'fab_logo.png')
    import os as _os
    logo_cell = RLImage(logo_path, width=1.6*cm, height=1.6*cm) if _os.path.exists(logo_path) else Paragraph('')

    header_style = ParagraphStyle('hdr', parent=styles['Normal'],
                                  fontName='Helvetica-Bold', fontSize=18,
                                  textColor=DARK, leading=22)
    sub_style    = ParagraphStyle('sub', parent=styles['Normal'],
                                  fontName='Helvetica', fontSize=9,
                                  textColor=MUTED)
    num_style    = ParagraphStyle('num', parent=styles['Normal'],
                                  fontName='Helvetica-Bold', fontSize=10,
                                  textColor=DARK, alignment=TA_RIGHT)
    meta_style   = ParagraphStyle('meta', parent=styles['Normal'],
                                  fontName='Helvetica', fontSize=9,
                                  textColor=MUTED, alignment=TA_RIGHT)

    header_table = Table(
        [[logo_cell,
          [Paragraph('FABOuanes', header_style), Paragraph(doc.get('subtitle',''), sub_style)],
          [Paragraph(doc['title'], header_style),
           Paragraph(f"N° {doc['number']}", num_style),
           Paragraph(f"Date : {doc['date']}", meta_style)]],
        ],
        colWidths=[2*cm, 9*cm, 6*cm]
    )
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.3*cm))
    story.append(HRFlowable(width="100%", thickness=1.5, color=DARK))
    story.append(Spacer(1, 0.5*cm))

    # ── Partner / Info block ─────────────────────────────────────────────
    lbl_style  = ParagraphStyle('lbl', parent=styles['Normal'],
                                fontName='Helvetica-Bold', fontSize=8,
                                textColor=MUTED, spaceAfter=2)
    val_style  = ParagraphStyle('val', parent=styles['Normal'],
                                fontName='Helvetica-Bold', fontSize=11,
                                textColor=DARK)
    info_style = ParagraphStyle('info', parent=styles['Normal'],
                                fontName='Helvetica', fontSize=9, textColor=MUTED)

    partner_block = [
        Paragraph(doc.get('partner_label',''), lbl_style),
        Paragraph(doc.get('partner_name',''), val_style),
    ]
    printed_block = [
        Paragraph('Créé par', lbl_style),
        Paragraph(printed_by, val_style),
    ]
    info_tbl = Table(
        [[partner_block, printed_block]],
        colWidths=[8.5*cm, 8.5*cm]
    )
    info_tbl.setStyle(TableStyle([
        ('BOX', (0,0), (0,0), 0.5, colors.HexColor('#e2e8f0')),
        ('BOX', (1,0), (1,0), 0.5, colors.HexColor('#e2e8f0')),
        ('BACKGROUND', (0,0), (-1,-1), LIGHT),
        ('LEFTPADDING', (0,0), (-1,-1), 10),
        ('RIGHTPADDING', (0,0), (-1,-1), 10),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('ROUNDEDCORNERS', [4, 4, 4, 4]),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ── Items table ──────────────────────────────────────────────────────
    th_style = ParagraphStyle('th', parent=styles['Normal'],
                               fontName='Helvetica-Bold', fontSize=9,
                               textColor=colors.white)
    td_style = ParagraphStyle('td', parent=styles['Normal'],
                               fontName='Helvetica', fontSize=10, textColor=DARK)
    td_right = ParagraphStyle('tdr', parent=td_style, alignment=TA_RIGHT)

    qty_str = f"{doc['quantity']} {doc['unit']}" if doc.get('quantity') is not None else '—'
    pu_str  = _fmt_money_pdf(doc.get('unit_price')) if doc.get('unit_price') is not None else '—'
    tot_str = _fmt_money_pdf(doc.get('total', 0))

    data = [
        [Paragraph(doc.get('item_label','Article'), th_style),
         Paragraph('Quantité', th_style),
         Paragraph('P.U.', th_style),
         Paragraph('Total', th_style)],
        [Paragraph(str(doc.get('item_name','—')), td_style),
         Paragraph(qty_str, td_right),
         Paragraph(pu_str, td_right),
         Paragraph(tot_str, td_right)],
    ]
    items_tbl = Table(data, colWidths=[9*cm, 3*cm, 3*cm, 2.5*cm])
    items_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), DARK),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT]),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#e2e8f0')),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
    ]))
    story.append(items_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ── Totals block ─────────────────────────────────────────────────────
    totals = []
    if doc.get('paid') is not None:
        totals.append(('Payé', _fmt_money_pdf(doc['paid']), GREEN))
    if doc.get('due') is not None:
        totals.append(('Reste à payer', _fmt_money_pdf(doc['due']), RED))
    totals.append(('Montant total', _fmt_money_pdf(doc.get('total', 0)), ACCENT))

    for label, value, col in totals:
        row_tbl = Table(
            [[Paragraph(label, ParagraphStyle('tl', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=MUTED)),
              Paragraph(value, ParagraphStyle('tv', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=11, textColor=col, alignment=TA_RIGHT))]],
            colWidths=[14*cm, 4*cm]
        )
        row_tbl.setStyle(TableStyle([
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        story.append(row_tbl)

    # ── Notes ─────────────────────────────────────────────────────────────
    if doc.get('notes'):
        story.append(Spacer(1, 0.4*cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#e2e8f0')))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph('Observations', lbl_style))
        story.append(Paragraph(doc['notes'], info_style))

    # ── Footer ────────────────────────────────────────────────────────────
    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#e2e8f0')))
    story.append(Spacer(1, 0.2*cm))
    footer_style = ParagraphStyle('ft', parent=styles['Normal'], fontName='Helvetica',
                                  fontSize=8, textColor=MUTED, alignment=TA_CENTER)
    story.append(Paragraph('FABOuanes — Gestion commerciale', footer_style))

    page_doc.build(story)
    buf.seek(0)
    return buf


def _fmt_money_pdf(val) -> str:
    """Format a monetary value for PDF output."""
    try:
        return f"{float(val):,.2f} DA".replace(',', ' ')
    except (TypeError, ValueError):
        return '0,00 DA'


def _generate_invoice_pdf(doc: dict, printed_by: str) -> BytesIO | None:
    """Generate the shared print PDF with a sober, print-first layout."""
    if not REPORTLAB_AVAILABLE:
        return None

    buf = BytesIO()
    page_doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=0.95 * cm,
        bottomMargin=0.95 * cm,
        leftMargin=1.0 * cm,
        rightMargin=1.0 * cm,
    )
    styles = getSampleStyleSheet()
    dark = colors.HexColor('#111827')
    muted = colors.HexColor('#6b7280')
    light = colors.HexColor('#f8fafc')
    line = colors.HexColor('#d7dee8')
    story = []

    logo_path = str(BASE_DIR / 'static' / 'fab_logo.png')
    import os as _os

    logo_size = 2.6 * cm
    content_width = page_doc.width
    logo_cell = RLImage(logo_path, width=logo_size, height=logo_size) if _os.path.exists(logo_path) else Spacer(logo_size, logo_size)

    title_style = ParagraphStyle(
        'pdf_title_modern',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=16.5,
        leading=17.4,
        textColor=dark,
    )
    partner_style = ParagraphStyle(
        'pdf_partner_modern',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=13.4,
        leading=14.4,
        textColor=colors.HexColor('#1f2937'),
    )
    partner_label_style = ParagraphStyle(
        'pdf_partner_label_modern',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7.0,
        leading=8.2,
        textColor=muted,
    )
    subtitle_style = ParagraphStyle(
        'pdf_subtitle_modern',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.6,
        leading=10.8,
        textColor=muted,
    )
    ref_label_style = ParagraphStyle(
        'pdf_ref_label_modern',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7.1,
        leading=8.5,
        alignment=TA_RIGHT,
        textColor=muted,
    )
    ref_value_style = ParagraphStyle(
        'pdf_ref_value_modern',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9.2,
        leading=10.5,
        alignment=TA_RIGHT,
        textColor=dark,
    )
    card_label_style = ParagraphStyle(
        'pdf_card_label_modern',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7.0,
        leading=8.6,
        textColor=muted,
    )
    card_value_style = ParagraphStyle(
        'pdf_card_value_modern',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9.0,
        leading=11.0,
        textColor=dark,
    )
    cell_text_style = ParagraphStyle(
        'pdf_cell_text_modern',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.3,
        leading=10.1,
        textColor=dark,
    )
    cell_right_style = ParagraphStyle(
        'pdf_cell_right_modern',
        parent=cell_text_style,
        alignment=TA_RIGHT,
    )
    table_head_style = ParagraphStyle(
        'pdf_table_head_modern',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7.1,
        leading=8.5,
        textColor=colors.HexColor('#374151'),
    )
    footer_style = ParagraphStyle(
        'pdf_footer_modern',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        alignment=TA_CENTER,
        textColor=muted,
    )
    printed_date = str(doc.get('printed_date') or doc.get('date') or '')
    printed_time = str(doc.get('printed_time') or '')

    ref_box = Table(
        [
            [Paragraph('Reference', ref_label_style), Paragraph(str(doc.get('number', '')), ref_value_style)],
            [Paragraph('Date', ref_label_style), Paragraph(printed_date, ref_value_style)],
            [Paragraph('Heure', ref_label_style), Paragraph(printed_time or '-', ref_value_style)],
            [Paragraph('Total', ref_label_style), Paragraph(_fmt_money_pdf(doc.get('total', 0)), ref_value_style)],
        ],
        colWidths=[2.1 * cm, 3.05 * cm],
    )
    ref_box.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('BOX', (0, 0), (-1, -1), 0.7, line),
        ('INNERGRID', (0, 0), (-1, -1), 0.4, line),
        ('LEFTPADDING', (0, 0), (-1, -1), 7),
        ('RIGHTPADDING', (0, 0), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    partner_box = Table(
        [[[
            Paragraph(str(doc.get('partner_label', '')), partner_label_style),
            Spacer(1, 0.05 * cm),
            Paragraph(str(doc.get('partner_name', '')), partner_style),
        ]]],
        colWidths=[4.75 * cm],
    )
    partner_box.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('BOX', (0, 0), (-1, -1), 0.7, line),
        ('LEFTPADDING', (0, 0), (-1, -1), 7),
        ('RIGHTPADDING', (0, 0), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))

    party_column = [
        logo_cell,
        Spacer(1, 0.12 * cm),
        partner_box,
    ]
    brand_copy = [
        Paragraph(str(doc.get('title', '')), title_style),
        Spacer(1, 0.1 * cm),
        Paragraph(str(doc.get('subtitle', '')), subtitle_style),
    ]
    party_width = 4.95 * cm
    ref_width = 5.15 * cm
    header_table = Table(
        [[party_column, brand_copy, ref_box]],
        colWidths=[party_width, content_width - party_width - ref_width, ref_width],
    )
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.18 * cm))
    story.append(HRFlowable(width='100%', thickness=0.8, color=line))
    story.append(Spacer(1, 0.22 * cm))

    printed_block = [
        Paragraph('Prepare par', card_label_style),
        Spacer(1, 0.04 * cm),
        Paragraph(str(printed_by), card_value_style),
    ]
    number_block = [
        Paragraph('Document', card_label_style),
        Spacer(1, 0.04 * cm),
        Paragraph(str(doc.get('number', '')), card_value_style),
    ]
    info_tbl = Table(
        [[printed_block, number_block]],
        colWidths=[content_width / 2.0, content_width / 2.0],
    )
    info_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('BOX', (0, 0), (-1, -1), 0.7, line),
        ('INNERGRID', (0, 0), (-1, -1), 0.4, line),
        ('LEFTPADDING', (0, 0), (-1, -1), 7),
        ('RIGHTPADDING', (0, 0), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 0.22 * cm))

    lines = doc.get('lines') or [{
        'item_name': doc.get('item_name') or '-',
        'quantity': doc.get('quantity'),
        'unit': doc.get('unit') or '',
        'unit_price': doc.get('unit_price'),
        'total': doc.get('total', 0),
    }]
    items_data = [[
        Paragraph(str(doc.get('item_label', 'Article')), table_head_style),
        Paragraph('Quantite', table_head_style),
        Paragraph('PU', table_head_style),
        Paragraph('Total', table_head_style),
    ]]
    for line_item in lines:
        qty_str = f"{line_item['quantity']} {line_item['unit']}" if line_item.get('quantity') is not None else '-'
        unit_price_str = _fmt_money_pdf(line_item.get('unit_price')) if line_item.get('unit_price') is not None else '-'
        total_str = _fmt_money_pdf(line_item.get('total', 0))
        items_data.append([
            Paragraph(str(line_item.get('item_name') or '-'), cell_text_style),
            Paragraph(qty_str, cell_right_style),
            Paragraph(unit_price_str, cell_right_style),
            Paragraph(total_str, cell_right_style),
        ])

    items_tbl = Table(
        items_data,
        colWidths=[content_width * 0.54, content_width * 0.15, content_width * 0.15, content_width * 0.16],
        repeatRows=1,
    )
    items_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), light),
        ('TEXTCOLOR', (0, 0), (-1, -1), dark),
        ('BOX', (0, 0), (-1, -1), 0.7, line),
        ('INNERGRID', (0, 0), (-1, -1), 0.4, line),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 7),
        ('RIGHTPADDING', (0, 0), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(items_tbl)
    story.append(Spacer(1, 0.24 * cm))

    summary_rows = []
    if doc.get('paid') is not None:
        summary_rows.append([
            Paragraph('Paye', card_label_style),
            Paragraph(_fmt_money_pdf(doc['paid']), ref_value_style),
        ])
    if doc.get('due') is not None:
        summary_rows.append([
            Paragraph('Reste', card_label_style),
            Paragraph(_fmt_money_pdf(doc['due']), ref_value_style),
        ])
    summary_rows.append([
        Paragraph('Total', card_label_style),
        Paragraph(_fmt_money_pdf(doc.get('total', 0)), ref_value_style),
    ])
    summary_tbl = Table(summary_rows, colWidths=[3.1 * cm, 3.1 * cm], hAlign='RIGHT')
    summary_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('BOX', (0, 0), (-1, -1), 0.7, line),
        ('INNERGRID', (0, 0), (-1, -1), 0.4, line),
        ('LEFTPADDING', (0, 0), (-1, -1), 7),
        ('RIGHTPADDING', (0, 0), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    story.append(summary_tbl)

    if doc.get('notes'):
        notes_tbl = Table(
            [
                [Paragraph('Observations', card_label_style)],
                [Paragraph(str(doc['notes']), cell_text_style)],
            ],
            colWidths=[content_width],
        )
        notes_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ('BOX', (0, 0), (-1, -1), 0.7, line),
            ('INNERGRID', (0, 0), (-1, -1), 0.4, line),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ]))
        story.append(Spacer(1, 0.22 * cm))
        story.append(notes_tbl)

    story.append(Spacer(1, 0.45 * cm))
    story.append(HRFlowable(width='100%', thickness=0.7, color=line))
    story.append(Spacer(1, 0.12 * cm))
    story.append(Paragraph('FABOuanes - Document commercial', footer_style))

    page_doc.build(story)
    buf.seek(0)
    return buf


@login_required
def print_document(doc_type: str, item_id: int):
    payload = build_print_payload(doc_type, item_id)
    if not payload:
        flash('Document introuvable pour impression.', 'danger')
        return redirect(url_for('index'))
    printed_at = datetime.now()
    payload = {
        **payload,
        'printed_date': printed_at.strftime('%Y-%m-%d'),
        'printed_time': printed_at.strftime('%H:%M'),
    }

    # Toujours afficher d'abord la facture HTML avec bouton retour.
    # Le PDF n'est généré que sur demande explicite via ?format=pdf.
    want_pdf = request.args.get('format') == 'pdf'

    if want_pdf:
        pdf_buf = _generate_invoice_pdf(payload, g.user['username'])
        if pdf_buf:
            filename = f"{payload['number']}.pdf"
            return send_file(
                pdf_buf,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=filename
            )
        flash("Génération PDF indisponible. Affichage HTML utilisé à la place.", 'warning')

    return render_template('print_document.html', doc=payload, printed_by=g.user['username'])


@app.route('/clients/<int:client_id>/delete', methods=['POST'])
@login_required
def delete_client(client_id: int):
    has_ops = query_db('SELECT 1 FROM sales WHERE client_id=? UNION SELECT 1 FROM raw_sales WHERE client_id=? UNION SELECT 1 FROM payments WHERE client_id=? LIMIT 1', (client_id, client_id, client_id), one=True)
    if has_ops:
        flash('Impossible de supprimer un client avec historique.', 'danger')
    else:
        execute_db('DELETE FROM clients WHERE id=?', (client_id,))
        log_activity('delete_client', 'client', client_id, 'Suppression client'); backup_database('delete_client'); flash('Client supprimé.', 'success')
    return redirect(url_for('clients'))


@app.route('/suppliers/<int:supplier_id>/delete', methods=['POST'])
@login_required
def delete_supplier(supplier_id: int):
    execute_db('DELETE FROM suppliers WHERE id=?', (supplier_id,))
    log_activity('delete_supplier', 'supplier', supplier_id, 'Suppression fournisseur'); backup_database('delete_supplier'); flash('Fournisseur supprimé.', 'success')
    return redirect(url_for('suppliers'))


@app.route('/raw-materials/<int:material_id>/delete', methods=['POST'])
@login_required
def delete_raw_material(material_id: int):
    linked = query_db('SELECT 1 FROM purchases WHERE raw_material_id=? UNION SELECT 1 FROM raw_sales WHERE raw_material_id=? UNION SELECT 1 FROM production_batch_items WHERE raw_material_id=? LIMIT 1', (material_id, material_id, material_id), one=True)
    if linked:
        flash('Impossible de supprimer une matière avec historique.', 'danger')
    else:
        execute_db('DELETE FROM raw_materials WHERE id=?', (material_id,))
        log_activity('delete_raw_material', 'raw_material', material_id, 'Suppression matière'); backup_database('delete_raw_material'); flash('Matière première supprimée.', 'success')
    return redirect(url_for('raw_materials'))


@app.route('/products/<int:product_id>/delete', methods=['POST'])
@login_required
def delete_product(product_id: int):
    linked = query_db('SELECT 1 FROM sales WHERE finished_product_id=? UNION SELECT 1 FROM production_batches WHERE finished_product_id=? LIMIT 1', (product_id, product_id), one=True)
    if linked:
        flash('Impossible de supprimer un produit avec historique.', 'danger')
    else:
        execute_db('DELETE FROM finished_products WHERE id=?', (product_id,))
        log_activity('delete_product', 'finished_product', product_id, 'Suppression produit'); backup_database('delete_product'); flash('Produit fini supprimé.', 'success')
    return redirect(url_for('products'))


@app.route('/purchases/<int:purchase_id>/delete', methods=['POST'])
@login_required
def delete_purchase(purchase_id: int):
    if reverse_purchase(purchase_id):
        log_activity('delete_purchase', 'purchase', purchase_id, 'Suppression achat'); backup_database('delete_purchase'); flash('Achat supprimé et stock corrigé.', 'success')
    else:
        flash('Impossible de supprimer cet achat.', 'danger')
    return redirect(url_for('purchases'))



@app.post('/production/edit-notes')
@login_required
def edit_production_notes():
    batch_id      = int(request.form.get('batch_id', 0))
    production_date = request.form.get('production_date', '').strip()
    notes         = request.form.get('notes', '').strip()
    if not batch_id:
        flash('Identifiant manquant.', 'danger')
        return redirect(url_for('production'))
    updates = {}
    if production_date:
        updates['production_date'] = production_date
    updates['notes'] = notes
    sets = ', '.join(f"{k}=?" for k in updates)
    vals = list(updates.values()) + [batch_id]
    execute_db(f"UPDATE production_batches SET {sets} WHERE id=?", tuple(vals))
    log_activity('edit_production_notes', 'production', batch_id, f"date={production_date}")
    flash('Production mise à jour.', 'success')
    return redirect(url_for('production'))


@app.route('/production/<int:batch_id>/delete', methods=['POST'])
@login_required
def delete_production(batch_id: int):
    if reverse_production(batch_id):
        log_activity('delete_production', 'production', batch_id, 'Suppression production'); backup_database('delete_production'); flash('Production supprimée et stock corrigé.', 'success')
    else:
        flash('Impossible de supprimer cette production.', 'danger')
    return redirect(url_for('production'))


@app.route('/sales/<kind>/<int:row_id>/delete', methods=['POST'])
@login_required
def delete_sale(kind: str, row_id: int):
    if reverse_sale(kind, row_id):
        log_activity('delete_sale', 'sale', row_id, f"Suppression vente {kind}"); backup_database('delete_sale'); flash('Vente supprimée et stock corrigé.', 'success')
    else:
        flash('Vente introuvable.', 'danger')
    return redirect(url_for('sales'))


@app.route('/payments/<int:payment_id>/delete', methods=['POST'])
@login_required
def delete_payment(payment_id: int):
    payment = query_db('SELECT * FROM payments WHERE id = ?', (payment_id,), one=True)
    if not payment:
        flash('Transaction introuvable.', 'danger')
        return redirect(url_for('transactions', type='payment'))
    with db_transaction():
        reverse_payment_allocations(payment)
        execute_db('DELETE FROM payments WHERE id = ?', (payment_id,))
    log_activity('delete_payment', 'payment', payment_id, 'Suppression transaction client'); backup_database('delete_payment'); flash('Transaction client supprimée.', 'success')
    return redirect(url_for('transactions', type='payment'))


def create_purchase_record(supplier_id, raw_id: int, qty: float, unit_price: float, purchase_date: str, notes: str, unit: str = 'kg', document_id: int | None = None, custom_item_name: str = '') -> int:
    with db_transaction():
        material = query_db('SELECT * FROM raw_materials WHERE id = ?', (raw_id,), one=True)
        if not material:
            raise ValueError('Matière première introuvable.')
        if purchase_date and purchase_date > date.today().isoformat():
            raise ValueError("La date d'achat ne peut pas être dans le futur.")
        total = qty * unit_price
        qty_kg = qty_to_kg(qty, unit)
        unit_price_kg = unit_price_to_kg(unit_price, unit)
        purchase_id = execute_db('INSERT INTO purchases (supplier_id, document_id, raw_material_id, quantity, unit, unit_price, total, purchase_date, notes) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (supplier_id, document_id, raw_id, qty_kg, unit, unit_price_kg, total, purchase_date, notes))
        new_stock = float(material['stock_qty']) + qty_kg
        current_value = float(material['stock_qty']) * float(material['avg_cost'])
        added_value = qty_kg * unit_price_kg
        avg_cost = (current_value + added_value) / new_stock if new_stock > 0 else 0
        sale_price = float(material['sale_price']) or unit_price
        execute_db('UPDATE raw_materials SET stock_qty = ?, avg_cost = ?, sale_price = ? WHERE id = ?', (new_stock, avg_cost, sale_price, raw_id))
        recalc_purchase_document_totals(document_id)
        return purchase_id


def create_sale_record(client_id, item_kind: str, item_id: int, qty: float, unit: str, unit_price: float, sale_type: str, sale_date: str, notes: str, amount_paid_input: float = 0, document_id: int | None = None, custom_item_name: str = '') -> tuple[str, int]:
    qty_kg = qty_to_kg(qty, unit)
    unit_price_kg = unit_price_to_kg(unit_price, unit)
    total = qty * unit_price
    requested_sale_type = (sale_type or '').strip().lower()
    if requested_sale_type not in {'cash', 'credit'}:
        requested_sale_type = 'credit' if client_id else 'cash'
    if requested_sale_type == 'credit' and not client_id:
        raise ValueError('Une vente à crédit nécessite un client.')
    effective_sale_type = requested_sale_type
    if effective_sale_type == 'cash':
        amount_paid = total
    else:
        amount_paid = max(0.0, min(float(amount_paid_input or 0), total))
    balance_due = round(total - amount_paid, 2)
    if qty <= 0:
        raise ValueError('La quantité doit être supérieure à zéro.')
    # Validation date — refuser les dates futures
    if sale_date and sale_date > date.today().isoformat():
        raise ValueError('La date de vente ne peut pas être dans le futur.')
    with db_transaction():
        if item_kind == 'finished':
            item = query_db('SELECT * FROM finished_products WHERE id = ?', (item_id,), one=True)
            if not item:
                raise ValueError('Produit fini introuvable.')
            if qty_kg > float(item['stock_qty']):
                raise ValueError('Stock produit insuffisant.')
            cost_snapshot = float(item['avg_cost'])
            below_cost_warning = unit_price_kg < cost_snapshot * 0.97 and cost_snapshot > 0
            profit_amount = total - qty_kg * cost_snapshot
            row_id = execute_db('INSERT INTO sales (client_id, finished_product_id, quantity, unit, unit_price, total, sale_type, amount_paid, balance_due, cost_price_snapshot, profit_amount, sale_date, notes) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                (client_id, item_id, qty, unit, unit_price, total, effective_sale_type, amount_paid, balance_due, cost_snapshot, profit_amount, sale_date, notes))
            execute_db('UPDATE finished_products SET stock_qty = stock_qty - ? WHERE id = ?', (qty_kg, item_id))
            if amount_paid > 0 and client_id:
                execute_db('INSERT INTO payments (client_id, sale_id, sale_kind, payment_type, amount, payment_date, notes) VALUES (?, ?, ?, ?, ?, ?, ?)', (client_id, row_id, 'finished', 'versement', amount_paid, sale_date, 'Paiement initial vente'))
            if below_cost_warning:
                flash(f"⚠️ Vente sous coût : {unit_price_kg:.2f} DA/kg < coût de revient {cost_snapshot:.2f} DA/kg — vérifiez le prix.", 'warning')
            return 'finished', row_id
        item = query_db('SELECT * FROM raw_materials WHERE id = ?', (item_id,), one=True)
        if not item:
            raise ValueError('Matière première introuvable.')
        if qty_kg > float(item['stock_qty']):
            raise ValueError('Stock matière insuffisant.')
        cost_snapshot = float(item['avg_cost'])
        below_cost_warning = unit_price_kg < cost_snapshot * 0.97 and cost_snapshot > 0
        profit_amount = total - qty_kg * cost_snapshot
        row_id = execute_db('INSERT INTO raw_sales (client_id, raw_material_id, quantity, unit, unit_price, total, sale_type, amount_paid, balance_due, cost_price_snapshot, profit_amount, sale_date, notes) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (client_id, item_id, qty, unit, unit_price, total, effective_sale_type, amount_paid, balance_due, cost_snapshot, profit_amount, sale_date, notes))
        execute_db('UPDATE raw_materials SET stock_qty = stock_qty - ? WHERE id = ?', (qty_kg, item_id))
        if amount_paid > 0 and client_id:
            execute_db('INSERT INTO payments (client_id, raw_sale_id, sale_kind, payment_type, amount, payment_date, notes) VALUES (?, ?, ?, ?, ?, ?, ?)', (client_id, row_id, 'raw', 'versement', amount_paid, sale_date, 'Paiement initial vente'))
        if below_cost_warning:
            flash(f"⚠️ Vente sous coût : {unit_price_kg:.2f} DA/kg < coût de revient {cost_snapshot:.2f} DA/kg — vérifiez le prix.", 'warning')
        return 'raw', row_id


def create_purchase_record(supplier_id, raw_id: int, qty: float, unit_price: float, purchase_date: str, notes: str, unit: str = 'kg', document_id: int | None = None, custom_item_name: str = '') -> int:
    with db_transaction():
        material = query_db('SELECT * FROM raw_materials WHERE id = ?', (raw_id,), one=True)
        if not material:
            raise ValueError('Matière première introuvable.')
        if purchase_date and purchase_date > date.today().isoformat():
            raise ValueError("La date d'achat ne peut pas être dans le futur.")
        custom_item_name = str(custom_item_name or '').strip()
        if is_other_operation_name(material['name']):
            unit = OTHER_OPERATION_UNIT
            if not custom_item_name:
                raise ValueError('Precise le nom du produit pour la ligne AUTRE.')
        else:
            custom_item_name = ''
        total = qty * unit_price
        qty_kg = qty_to_kg(qty, unit)
        unit_price_kg = unit_price_to_kg(unit_price, unit)
        purchase_id = execute_db(
            'INSERT INTO purchases (supplier_id, document_id, raw_material_id, quantity, unit, unit_price, total, purchase_date, notes, custom_item_name) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (supplier_id, document_id, raw_id, qty_kg, unit, unit_price_kg, total, purchase_date, notes, custom_item_name),
        )
        new_stock = float(material['stock_qty']) + qty_kg
        current_value = float(material['stock_qty']) * float(material['avg_cost'])
        added_value = qty_kg * unit_price_kg
        avg_cost = (current_value + added_value) / new_stock if new_stock > 0 else 0
        sale_price = float(material['sale_price']) or unit_price
        execute_db('UPDATE raw_materials SET stock_qty = ?, avg_cost = ?, sale_price = ? WHERE id = ?', (new_stock, avg_cost, sale_price, raw_id))
        recalc_purchase_document_totals(document_id)
        return purchase_id


def create_sale_record(client_id, item_kind: str, item_id: int, qty: float, unit: str, unit_price: float, sale_type: str, sale_date: str, notes: str, amount_paid_input: float = 0, document_id: int | None = None, custom_item_name: str = '') -> tuple[str, int]:
    total = qty * unit_price
    requested_sale_type = (sale_type or '').strip().lower()
    if requested_sale_type not in {'cash', 'credit'}:
        requested_sale_type = 'credit' if client_id else 'cash'
    if requested_sale_type == 'credit' and not client_id:
        raise ValueError('Une vente à crédit nécessite un client.')
    effective_sale_type = requested_sale_type
    if effective_sale_type == 'cash':
        amount_paid = total
    else:
        amount_paid = max(0.0, min(float(amount_paid_input or 0), total))
    balance_due = round(total - amount_paid, 2)
    if qty <= 0:
        raise ValueError('La quantité doit être supérieure à zéro.')
    if sale_date and sale_date > date.today().isoformat():
        raise ValueError('La date de vente ne peut pas être dans le futur.')
    with db_transaction():
        if item_kind == 'finished':
            custom_item_name = ''
            qty_kg = qty_to_kg(qty, unit)
            unit_price_kg = unit_price_to_kg(unit_price, unit)
            item = query_db('SELECT * FROM finished_products WHERE id = ?', (item_id,), one=True)
            if not item:
                raise ValueError('Produit fini introuvable.')
            if qty_kg > float(item['stock_qty']):
                raise ValueError('Stock produit insuffisant.')
            cost_snapshot = float(item['avg_cost'])
            below_cost_warning = unit_price_kg < cost_snapshot * 0.97 and cost_snapshot > 0
            profit_amount = total - qty_kg * cost_snapshot
            row_id = execute_db(
                'INSERT INTO sales (client_id, document_id, finished_product_id, quantity, unit, unit_price, total, sale_type, amount_paid, balance_due, cost_price_snapshot, profit_amount, sale_date, notes) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                (client_id, document_id, item_id, qty, unit, unit_price, total, effective_sale_type, amount_paid, balance_due, cost_snapshot, profit_amount, sale_date, notes),
            )
            execute_db('UPDATE finished_products SET stock_qty = stock_qty - ? WHERE id = ?', (qty_kg, item_id))
            if amount_paid > 0 and client_id:
                execute_db('INSERT INTO payments (client_id, sale_id, sale_kind, payment_type, amount, payment_date, notes) VALUES (?, ?, ?, ?, ?, ?, ?)', (client_id, row_id, 'finished', 'versement', amount_paid, sale_date, 'Paiement initial vente'))
            recalc_sale_document_totals(document_id)
            if below_cost_warning:
                flash(f"Vente sous coût : {unit_price_kg:.2f} DA/kg < coût de revient {cost_snapshot:.2f} DA/kg — vérifiez le prix.", 'warning')
            return 'finished', row_id

        item = query_db('SELECT * FROM raw_materials WHERE id = ?', (item_id,), one=True)
        if not item:
            raise ValueError('Matière première introuvable.')
        custom_item_name = str(custom_item_name or '').strip()
        if is_other_operation_name(item['name']):
            unit = OTHER_OPERATION_UNIT
            if not custom_item_name:
                raise ValueError('Precise le nom du produit pour la ligne AUTRE.')
        else:
            custom_item_name = ''
        qty_kg = qty_to_kg(qty, unit)
        unit_price_kg = unit_price_to_kg(unit_price, unit)
        if qty_kg > float(item['stock_qty']):
            raise ValueError('Stock matière insuffisant.')
        cost_snapshot = float(item['avg_cost'])
        below_cost_warning = unit_price_kg < cost_snapshot * 0.97 and cost_snapshot > 0
        profit_amount = total - qty_kg * cost_snapshot
        row_id = execute_db(
            'INSERT INTO raw_sales (client_id, document_id, raw_material_id, quantity, unit, unit_price, total, sale_type, amount_paid, balance_due, cost_price_snapshot, profit_amount, sale_date, notes, custom_item_name) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (client_id, document_id, item_id, qty, unit, unit_price, total, effective_sale_type, amount_paid, balance_due, cost_snapshot, profit_amount, sale_date, notes, custom_item_name),
        )
        execute_db('UPDATE raw_materials SET stock_qty = stock_qty - ? WHERE id = ?', (qty_kg, item_id))
        if amount_paid > 0 and client_id:
            execute_db('INSERT INTO payments (client_id, raw_sale_id, sale_kind, payment_type, amount, payment_date, notes) VALUES (?, ?, ?, ?, ?, ?, ?)', (client_id, row_id, 'raw', 'versement', amount_paid, sale_date, 'Paiement initial vente'))
        recalc_sale_document_totals(document_id)
        if below_cost_warning:
            flash(f"Vente sous coût : {unit_price_kg:.2f} DA/kg < coût de revient {cost_snapshot:.2f} DA/kg — vérifiez le prix.", 'warning')
        return 'raw', row_id


def create_payment_record(client_id: int, amount: float, payment_date: str, notes: str, sale_link: str = '', payment_type: str = 'versement') -> int:
    if amount <= 0:
        raise ValueError('Le montant doit être supérieur à zéro.')
    with db_transaction():
        client = query_db('SELECT id FROM clients WHERE id = ?', (client_id,), one=True)
        if not client:
            raise ValueError('Client introuvable.')
        if payment_type == 'avance':
            return execute_db(
                'INSERT INTO payments (client_id, sale_id, raw_sale_id, sale_kind, payment_type, allocation_meta, amount, payment_date, notes) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
                (client_id, None, None, None, 'avance', None, amount, payment_date, notes or 'Avance client')
            )

        sale_id = None
        raw_sale_id = None
        sale_kind = None
        allocations: list[dict[str, Any]] = []
        applied = 0.0

        if sale_link and ':' in sale_link:
            sale_kind, id_str = sale_link.split(':', 1)
            row_id = int(id_str)
            entry = query_db('SELECT client_id FROM sales WHERE id = ?' if sale_kind == 'finished' else 'SELECT client_id FROM raw_sales WHERE id = ?', (row_id,), one=True)
            if entry and int(entry['client_id'] or 0) != client_id:
                raise ValueError('Cette créance ne correspond pas au client choisi.')
            applied = apply_payment_to_entry(sale_kind, row_id, amount)
            if applied <= 0:
                raise ValueError('Aucune créance ouverte à solder pour ce client.')
            allocations = [{'kind': sale_kind, 'id': row_id, 'amount': applied}]
            if sale_kind == 'finished':
                sale_id = row_id
            else:
                raw_sale_id = row_id
        else:
            remaining = amount
            for entry in get_open_credit_entries(client_id):
                if remaining <= 0:
                    break
                paid = apply_payment_to_entry(entry['item_kind'], entry['id'], remaining)
                if paid > 0:
                    allocations.append({'kind': entry['item_kind'], 'id': int(entry['id']), 'amount': paid})
                    applied += paid
                    remaining -= paid

            current_balance = client_balance(client_id)
            if applied <= 0 and current_balance <= 0:
                raise ValueError('Aucune dette ouverte pour ce client.')

        return execute_db(
            'INSERT INTO payments (client_id, sale_id, raw_sale_id, sale_kind, payment_type, allocation_meta, amount, payment_date, notes) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (client_id, sale_id, raw_sale_id, sale_kind, 'versement', json.dumps(allocations) if allocations else None, amount, payment_date, notes or 'Versement client')
        )


@login_required
def contacts():
    rows = query_db(
        """
        SELECT * FROM (
            SELECT 'Client' AS contact_type, c.id, c.name, c.phone, c.address, c.notes,
                   c.opening_credit
                   + COALESCE((SELECT SUM(total) FROM sales s WHERE s.client_id = c.id AND s.sale_type = 'credit'), 0)
                   + COALESCE((SELECT SUM(total) FROM raw_sales rs WHERE rs.client_id = c.id AND rs.sale_type = 'credit'), 0)
                   - COALESCE((SELECT SUM(amount) FROM payments p WHERE p.client_id = c.id AND p.payment_type = 'versement'), 0)
                   + COALESCE((SELECT SUM(amount) FROM payments p WHERE p.client_id = c.id AND p.payment_type = 'avance'), 0) AS current_balance,
                   COALESCE((SELECT SUM(total) FROM sales s WHERE s.client_id = c.id), 0)
                   + COALESCE((SELECT SUM(total) FROM raw_sales rs WHERE rs.client_id = c.id), 0) AS total_amount,
                   COALESCE((SELECT SUM(amount) FROM payments p WHERE p.client_id = c.id AND p.payment_type = 'versement'), 0) AS total_paid
            FROM clients c
            UNION ALL
            SELECT 'Fournisseur' AS contact_type, s.id, s.name, s.phone, s.address, s.notes,
                   0 AS current_balance,
                   COALESCE((SELECT SUM(total) FROM purchases p WHERE p.supplier_id = s.id), 0) AS total_amount,
                   0 AS total_paid
            FROM suppliers s
        ) x ORDER BY contact_type, name
        """
    )
    filter_type = request.args.get('type', 'all')
    filter_name = (request.args.get('name') or '').strip().lower()
    filtered_rows = []
    for row in rows:
        if filter_type == 'client' and row['contact_type'] != 'Client':
            continue
        if filter_type == 'supplier' and row['contact_type'] != 'Fournisseur':
            continue
        hay = f"{row['name']} {row['phone'] or ''} {row['address'] or ''}".lower()
        if filter_name and filter_name not in hay:
            continue
        filtered_rows.append(row)
    return render_template('contacts.html', contacts=filtered_rows, filter_type=filter_type, filter_name=request.args.get('name', ''))


@login_required
def transactions():
    rows = query_db(
        """
        SELECT * FROM (
            SELECT 'Achat' AS tx_type, 'purchase' AS tx_kind, p.id, p.purchase_date AS tx_date,
                   COALESCE(s.name, '-') AS partner_name, r.name AS designation,
                   CASE
                       WHEN lower(COALESCE(p.unit, r.unit, 'kg')) = 'sac' THEN p.quantity / 50.0
                       WHEN lower(COALESCE(p.unit, r.unit, 'kg')) IN ('qt', 'quintal') THEN p.quantity / 100.0
                       ELSE p.quantity
                   END AS quantity,
                   COALESCE(p.unit, r.unit, 'kg') AS unit,
                   CASE
                       WHEN lower(COALESCE(p.unit, r.unit, 'kg')) = 'sac' THEN p.unit_price * 50.0
                       WHEN lower(COALESCE(p.unit, r.unit, 'kg')) IN ('qt', 'quintal') THEN p.unit_price * 100.0
                       ELSE p.unit_price
                   END AS unit_price,
                   p.total, NULL AS paid, NULL AS due, p.document_id AS document_id
            FROM purchases p
            LEFT JOIN suppliers s ON s.id = p.supplier_id
            JOIN raw_materials r ON r.id = p.raw_material_id
            UNION ALL
            SELECT 'Vente' AS tx_type,
                   CASE WHEN x.row_kind='finished' THEN 'sale_finished' ELSE 'sale_raw' END AS tx_kind,
                   x.id, x.sale_date AS tx_date, COALESCE(x.client_name, '-') AS partner_name, x.item_name AS designation,
                   x.quantity, x.unit, x.unit_price, x.total, x.amount_paid AS paid, x.balance_due AS due, x.document_id AS document_id
            FROM (
                SELECT s.id, s.document_id, 'finished' AS row_kind, s.sale_date, c.name AS client_name, f.name AS item_name, s.quantity, s.unit, s.unit_price, s.total, s.amount_paid, s.balance_due
                FROM sales s LEFT JOIN clients c ON c.id = s.client_id JOIN finished_products f ON f.id = s.finished_product_id
                UNION ALL
                SELECT rs.id, rs.document_id, 'raw' AS row_kind, rs.sale_date, c.name AS client_name, r.name AS item_name, rs.quantity, rs.unit, rs.unit_price, rs.total, rs.amount_paid, rs.balance_due
                FROM raw_sales rs LEFT JOIN clients c ON c.id = rs.client_id JOIN raw_materials r ON r.id = rs.raw_material_id
            ) x
            UNION ALL
            SELECT CASE WHEN p.payment_type='avance' THEN 'Avance' ELSE 'Versement' END AS tx_type, 'payment' AS tx_kind, p.id, p.payment_date AS tx_date,
                   c.name AS partner_name,
                   CASE
                       WHEN p.sale_kind = 'finished' AND p.sale_id IS NOT NULL THEN 'Versement vente #' || p.sale_id
                       WHEN p.sale_kind = 'raw' AND p.raw_sale_id IS NOT NULL THEN 'Versement vente matière #' || p.raw_sale_id
                       ELSE CASE WHEN p.payment_type='avance' THEN 'Avance client' ELSE 'Versement client' END
                   END AS designation,
                   NULL AS quantity, NULL AS unit, NULL AS unit_price, p.amount AS total, p.amount AS paid, NULL AS due, NULL AS document_id
            FROM payments p JOIN clients c ON c.id = p.client_id
        ) t
        ORDER BY tx_date DESC, id DESC
        """
    )
    filter_type = request.args.get('type', 'all')
    filter_name = (request.args.get('name') or '').strip().lower()
    filter_date = (request.args.get('date') or '').strip()
    filter_operation = (request.args.get('operation') or '').strip().lower()

    filtered_rows = []
    for row in rows:
        if filter_type == 'purchase' and row['tx_type'] != 'Achat':
            continue
        if filter_type == 'sale' and row['tx_type'] != 'Vente':
            continue
        if filter_type == 'payment' and row['tx_kind'] != 'payment':
            continue
        if filter_name and filter_name not in f"{row['partner_name']} {row['designation']}".lower():
            continue
        if filter_date and str(row['tx_date']) != filter_date:
            continue
        if filter_operation and filter_operation not in str(row['tx_type']).lower():
            continue
        filtered_rows.append(row)

    return render_template(
        'transactions.html',
        transactions=filtered_rows,
        filter_type=filter_type,
        filter_name=request.args.get('name', ''),
        filter_date=request.args.get('date', ''),
        filter_operation=request.args.get('operation', '')
    )


@login_required
def supplier_detail(supplier_id: int):
    supplier = query_db('SELECT * FROM suppliers WHERE id = ?', (supplier_id,), one=True)
    if not supplier:
        flash('Fournisseur introuvable.', 'danger')
        return redirect(url_for('contacts', type='supplier'))
    purchases_rows = query_db(
        '''SELECT p.id, p.document_id, p.purchase_date AS event_date, r.name AS designation, p.quantity, r.unit, p.unit_price, p.total, p.notes
           FROM purchases p JOIN raw_materials r ON r.id = p.raw_material_id
           WHERE p.supplier_id = ? ORDER BY p.purchase_date DESC, p.id DESC''', (supplier_id,))
    total_amount = sum(float(x['total']) for x in purchases_rows)
    return render_template('supplier_detail.html', supplier=supplier, purchases=purchases_rows, total_amount=total_amount)


@app.route('/purchases/<int:purchase_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_purchase(purchase_id: int):
    purchase = query_db('SELECT * FROM purchases WHERE id = ?', (purchase_id,), one=True)
    if not purchase:
        flash('Achat introuvable.', 'danger')
        return redirect(url_for('transactions', type='purchase'))
    if request.method == 'POST':
        supplier_id = request.form.get('supplier_id') or None
        raw_id = int(request.form['raw_material_id'])
        qty = to_float(request.form.get('quantity'))
        unit_price = to_float(request.form.get('unit_price'))
        purchase_date = request.form.get('purchase_date') or date.today().isoformat()
        notes = request.form.get('notes', '').strip()
        if qty <= 0:
            flash('La quantité doit être supérieure à zéro.', 'danger')
            return redirect(request.url)
        try:
            with db_transaction():
                if not reverse_purchase(purchase_id):
                    raise ValueError("Impossible de modifier cet achat car le stock ne permet pas de l’annuler.")
                create_purchase_record(supplier_id, raw_id, qty, unit_price, purchase_date, notes, request.form.get('unit', 'kg'))
            log_activity('update_purchase', 'purchase', purchase_id, f"matière #{raw_id} qty={qty}"); backup_database('update_purchase'); flash('Achat modifié.', 'success')
            return redirect(url_for('transactions', type='purchase'))
        except Exception as e:
            flash(str(e), 'danger')
            return redirect(request.url)
    return render_template('purchase_edit.html', purchase=purchase, suppliers=query_db('SELECT * FROM suppliers ORDER BY name'), raw_materials=query_db('SELECT * FROM raw_materials ORDER BY name'), units=unit_choices())


@app.route('/sales/<kind>/<int:row_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_sale(kind: str, row_id: int):
    if kind == 'finished':
        row = query_db("SELECT s.*, 'finished' AS row_kind, 'finished:' || s.finished_product_id AS item_key FROM sales s WHERE s.id = ?", (row_id,), one=True)
    else:
        row = query_db("SELECT rs.*, 'raw' AS row_kind, 'raw:' || rs.raw_material_id AS item_key FROM raw_sales rs WHERE rs.id = ?", (row_id,), one=True)
    if not row:
        flash('Vente introuvable.', 'danger')
        return redirect(url_for('transactions', type='sale'))
    if request.method == 'POST':
        client_id = request.form.get('client_id') or None
        item_key = request.form['item_key']
        item_kind, item_id_str = item_key.split(':', 1)
        qty = to_float(request.form.get('quantity'))
        unit = request.form.get('unit', 'kg').strip()
        unit_price = to_float(request.form.get('unit_price'))
        sale_date = request.form.get('sale_date') or date.today().isoformat()
        notes = request.form.get('notes', '').strip()
        sale_type = 'credit' if client_id else 'cash'
        amount_paid = 0 if client_id else qty * unit_price
        try:
            with db_transaction():
                if not reverse_sale(kind, row_id):
                    raise ValueError('Impossible de modifier cette vente.')
                create_sale_record(client_id, item_kind, int(item_id_str), qty, unit, unit_price, sale_type, sale_date, notes, amount_paid)
            log_activity('update_sale', 'sale', row_id, f"{item_kind} #{item_id_str} qty={qty} {unit}"); backup_database('update_sale'); flash('Vente modifiée.', 'success')
            return redirect(url_for('transactions', type='sale'))
        except Exception as e:
            flash(str(e), 'danger')
            return redirect(request.url)
    sellable_items = [
        {'key': f"raw:{r['id']}", 'label': f"{r['name']} (matière première)", 'unit': r['unit'], 'sale_price': r['sale_price'], 'stock_qty': r['stock_qty'], 'avg_cost': r['avg_cost']}
        for r in query_db('SELECT * FROM raw_materials ORDER BY name')
    ] + [
        {'key': f"finished:{p['id']}", 'label': f"{p['name']} (produit fini)", 'unit': p['default_unit'], 'sale_price': p['sale_price'], 'stock_qty': p['stock_qty'], 'avg_cost': p['avg_cost']}
        for p in query_db('SELECT * FROM finished_products ORDER BY name')
    ]
    sellable_items = sorted(sellable_items, key=lambda x: x['label'])
    return render_template('sale_edit.html', sale=row, clients=query_db('SELECT * FROM clients ORDER BY name'), sellable_items=sellable_items, units=unit_choices())


@app.route('/payments/<int:payment_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_payment(payment_id: int):
    payment = query_db('SELECT * FROM payments WHERE id = ?', (payment_id,), one=True)
    if not payment:
        flash('Versement introuvable.', 'danger')
        return redirect(url_for('transactions', type='payment'))
    if request.method == 'POST':
        client_id = int(request.form['client_id'])
        sale_link = request.form.get('sale_link') or ''
        amount = to_float(request.form.get('amount'))
        payment_date = request.form.get('payment_date') or date.today().isoformat()
        notes = request.form.get('notes', '').strip()
        try:
            with db_transaction():
                reverse_payment_allocations(payment)
                execute_db('DELETE FROM payments WHERE id = ?', (payment_id,))
                create_payment_record(client_id, amount, payment_date, notes, sale_link, request.form.get('payment_type', 'versement'))
            log_activity('update_payment', 'payment', payment_id, f"client #{client_id} {request.form.get('payment_type', 'versement')} montant={amount}"); backup_database('update_payment'); flash('Transaction client modifiée.', 'success')
            return redirect(url_for('transactions', type='payment'))
        except Exception as e:
            flash(str(e), 'danger')
            return redirect(request.url)
    current_link = ''
    if payment['sale_kind'] == 'finished' and payment['sale_id']:
        current_link = f"finished:{payment['sale_id']}"
    elif payment['sale_kind'] == 'raw' and payment['raw_sale_id']:
        current_link = f"raw:{payment['raw_sale_id']}"
    open_sales = list(get_open_credit_entries())
    existing_keys = [f"{s['item_kind']}:{s['id']}" for s in open_sales]
    if current_link and current_link not in existing_keys:
        if payment['sale_kind'] == 'finished' and payment['sale_id']:
            s = query_db('SELECT s.id, s.client_id, c.name AS client_name, f.name AS item_name, s.balance_due + ? AS balance_due, s.sale_date, s.total FROM sales s JOIN clients c ON c.id=s.client_id JOIN finished_products f ON f.id=s.finished_product_id WHERE s.id=?', (payment['amount'], payment['sale_id']), one=True)
            if s:
                open_sales.append(dict(item_kind='finished', id=s['id'], client_id=s['client_id'], client_name=s['client_name'], item_name=s['item_name'], balance_due=s['balance_due'], sale_date=s['sale_date'], total=s['total']))
        elif payment['sale_kind'] == 'raw' and payment['raw_sale_id']:
            s = query_db('SELECT rs.id, rs.client_id, c.name AS client_name, r.name AS item_name, rs.balance_due + ? AS balance_due, rs.sale_date, rs.total FROM raw_sales rs JOIN clients c ON c.id=rs.client_id JOIN raw_materials r ON r.id=rs.raw_material_id WHERE rs.id=?', (payment['amount'], payment['raw_sale_id']), one=True)
            if s:
                open_sales.append(dict(item_kind='raw', id=s['id'], client_id=s['client_id'], client_name=s['client_name'], item_name=s['item_name'], balance_due=s['balance_due'], sale_date=s['sale_date'], total=s['total']))
    return render_template('payment_edit.html', payment=payment, current_link=current_link, clients=query_db('SELECT * FROM clients ORDER BY name'), open_sales=open_sales)



@app.errorhandler(Exception)
def handle_unexpected_error(exc):
    if isinstance(exc, HTTPException):
        return exc
    try:
        log_error(exc, request.path)
    except Exception:
        pass
    return render_template('error.html', error_message="Une erreur inattendue s'est produite."), 500


def log_server_start() -> None:
    ensure_runtime_dirs()
    conn = connect_database(DATABASE_URL, DB_PATH)
    conn.execute('CREATE TABLE IF NOT EXISTS system_logs (id INTEGER PRIMARY KEY AUTOINCREMENT, level TEXT NOT NULL, message TEXT NOT NULL, created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP)')
    conn.execute('INSERT INTO system_logs (level, message, created_at) VALUES (?, ?, CURRENT_TIMESTAMP)', ('info', 'Démarrage du serveur'))
    conn.commit()
    conn.close()
    write_text_log('system.log', 'INFO | Démarrage du serveur')


def log_server_stop() -> None:
    try:
        conn = connect_database(DATABASE_URL, DB_PATH)
        conn.execute('INSERT INTO system_logs (level, message, created_at) VALUES (?, ?, CURRENT_TIMESTAMP)', ('warning', 'Arrêt du serveur'))
        conn.commit()
        conn.close()
        write_text_log('system.log', 'WARNING | Arrêt du serveur')
    except Exception:
        pass


atexit.register(log_server_stop)

def main() -> None:
    ensure_runtime_dirs()
    init_db()
    log_server_start()
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)


if __name__ == '__main__':
    main()
