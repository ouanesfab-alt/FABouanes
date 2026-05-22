from __future__ import annotations
from unittest.mock import patch
import pandas as pd
import pytest
from datetime import date
from app.services.excel_import_service import (
    parse_client_history_excel,
    _to_float,
    _classify_operation,
    parse_flexible_date,
    parse_flexible_amount,
    parse_excel_client_history,
    parse_excel_client_file,
)


def test_parse_client_history_excel_success():
    # Mock data for Excel file
    # First iloc[0, 1] is client name.
    # Ligne 2 (index 2) is headers: Date | Designation | Montant a Payer | Versement | Reste a Payer
    mock_df_raw = pd.DataFrame([
        [None, "Achour Boudjmaa", None, "TEL :", None],
        [None, None, None, None, None],
        ["Date", "Designation", "Montant a Payer", "Versement", "Reste a Payer"],
        ["01/01/2026", "Ancien solde", "0", "0", "10000"],
        ["02/01/2026", "Achat 1", "5000", "0", "15000"],
        ["03/01/2026", "Versement 1", "0", "4000", "11000"],
    ])

    with patch("pandas.read_excel") as mock_read_excel:
        # Mock first read (raw) and second read (with skiprows)
        mock_read_excel.side_effect = [
            mock_df_raw,  # First call (header=None)
            pd.DataFrame({
                "date_raw": ["01/01/2026", "02/01/2026", "03/01/2026"],
                "designation": ["Ancien solde", "Achat 1", "Versement 1"],
                "montant_achat_raw": ["0", "5000", "0"],
                "montant_verse_raw": ["0", "0", "4000"],
                "solde_cumule_raw": ["10000", "15000", "11000"],
            })  # Second call (with skiprows=2)
        ]

        result = parse_client_history_excel("dummy_path.xlsx")
        
        assert result["client_name"] == "Achour Boudjmaa"
        assert result["solde_final"] == 11000.0
        assert result["total_achats"] == 5000.0
        assert result["total_verses"] == 4000.0
        assert result["nb_lignes"] == 3
        
        # Verify rows classification
        rows = result["rows"]
        assert rows[0]["type_operation"] == "ouverture"
        assert rows[1]["type_operation"] == "achat"
        assert rows[2]["type_operation"] == "versement"


def test_parse_client_history_excel_invalid_shape():
    mock_df_raw = pd.DataFrame([
        [None, "Achour"]
    ])
    with patch("pandas.read_excel", return_value=mock_df_raw):
        with pytest.raises(ValueError, match="Le fichier Excel ne respecte pas le format attendu"):
            parse_client_history_excel("dummy.xlsx")


def test_parse_client_history_excel_missing_client_name():
    # Client name cell is empty or nan
    mock_df_raw = pd.DataFrame([
        [None, None, None],
        [None, None, None],
        ["Date", "Designation", "Montant a Payer", "Versement", "Reste a Payer"],
        ["01/01/2026", "Ancien solde", "0", "0", "10000"]
    ])
    with patch("pandas.read_excel") as mock_read_excel:
        mock_read_excel.side_effect = [
            mock_df_raw,
            pd.DataFrame({
                "date_raw": ["01/01/2026"],
                "designation": ["Ancien solde"],
                "montant_achat_raw": ["0"],
                "montant_verse_raw": ["0"],
                "solde_cumule_raw": ["10000"]
            })
        ]
        result = parse_client_history_excel("dummy.xlsx")
        assert result["client_name"] == "Client inconnu"


def test_parse_client_history_excel_no_rows():
    # No rows with valid dates
    mock_df_raw = pd.DataFrame([
        [None, "Client name", None],
        [None, None, None],
        ["Date", "Designation", "Montant a Payer", "Versement", "Reste a Payer"],
        [None, None, None, None, None]
    ])
    with patch("pandas.read_excel") as mock_read_excel:
        mock_read_excel.side_effect = [
            mock_df_raw,
            pd.DataFrame({
                "date_raw": [None],
                "designation": [None],
                "montant_achat_raw": [None],
                "montant_verse_raw": [None],
                "solde_cumule_raw": [None]
            })
        ]
        with pytest.raises(ValueError, match="Aucune ligne de données valide trouvée"):
            parse_client_history_excel("dummy.xlsx")


def test_to_float_conversions():
    assert _to_float("12 500,50") == 12500.50
    assert _to_float("12\xa0500.50") == 12500.50
    assert _to_float("nan") == 0.0
    assert _to_float(None) == 0.0
    assert _to_float("invalid") == 0.0


def test_classify_operations():
    # Mixed and immediate
    assert _classify_operation("Achat", 100.0, 100.0, 1) == "immediat"
    assert _classify_operation("Achat", 100.0, 50.0, 1) == "mixte"
    assert _classify_operation("Achat", 0.0, 0.0, 1) == "achat"


def test_parse_flexible_date():
    from datetime import datetime, date
    today_iso = date.today().isoformat()
    
    assert parse_flexible_date(datetime(2026, 5, 22, 12, 0)) == "2026-05-22"
    assert parse_flexible_date(date(2026, 5, 22)) == "2026-05-22"
    assert parse_flexible_date("22/05/2026") == "2026-05-22"
    assert parse_flexible_date("22-05-2026") == "2026-05-22"
    assert parse_flexible_date("2026-05-22") == "2026-05-22"
    assert parse_flexible_date("invalid_date") == today_iso
    assert parse_flexible_date(None) == today_iso


def test_parse_flexible_amount():
    assert parse_flexible_amount(1500) == 1500.0
    assert parse_flexible_amount(1500.75) == 1500.75
    assert parse_flexible_amount("1 500 DA") == 1500.0
    assert parse_flexible_amount("1,500.75 da") == 1500.75
    assert parse_flexible_amount("invalid") == 0.0
    assert parse_flexible_amount(None) == 0.0


@pytest.fixture
def mock_openpyxl_workbook():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    ws["B1"] = "Achour Boudjmaa"
    ws["D1"] = "TEL : 0550000000"
    
    ws["A3"] = "Date"
    ws["B3"] = "Designation"
    ws["C3"] = "Montant a Payer"
    ws["D3"] = "Versement"
    ws["E3"] = "Reste a Payer"
    
    # Ancien solde
    ws["A4"] = "01/01/2026"
    ws["B4"] = "Ancien solde"
    ws["E4"] = "10000"
    
    # Achat
    ws["A5"] = "02/01/2026"
    ws["B5"] = "Achat 1"
    ws["C5"] = "5000"
    ws["E5"] = "15000"
    
    return wb


def test_parse_excel_client_history(tmp_path, mock_openpyxl_workbook):
    file_path = tmp_path / "test_history.xlsx"
    mock_openpyxl_workbook.save(file_path)
    
    res = parse_excel_client_history(str(file_path))
    assert res["last_date"] == "2026-01-02"
    assert res["last_balance"] == 15000.0


def test_parse_excel_client_file(tmp_path, mock_openpyxl_workbook):
    file_path = tmp_path / "test_file.xlsx"
    mock_openpyxl_workbook.save(file_path)
    
    res = parse_excel_client_file(str(file_path))
    assert res["name"] == "Achour Boudjmaa"
    assert res["phone"] == "0550000000"
    assert res["opening_credit"] == 10000.0
    assert res["history_count"] == 2
