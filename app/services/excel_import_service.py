from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any
import pandas as pd
from decimal import Decimal, InvalidOperation


def parse_client_history_excel(file_path: str) -> dict:
    """
    Parse un grand livre client au format Excel FABOuanes.

    Structure attendue du fichier :
      - iloc[0, 1] : nom du client (ou recherché dans la première ligne)
      - Ligne d'en-tête (Date | Designation | Montant a Payer | Versement | Reste a Payer) détectée dynamiquement
      - Lignes de données suivantes

    Retourne un dict avec :
      {
        "client_name": str,
        "rows": [
          {
            "ordre_import": int,
            "date": str,                # format ISO "YYYY-MM-DD"
            "designation": str,
            "montant_achat": float,
            "montant_verse": float,
            "solde_cumule": float,
            "type_operation": str,      # "achat"|"versement"|"mixte"|"immediat"|"ouverture"
          },
          ...
        ],
        "solde_final": float,
        "total_achats": float,
        "total_verses": float,
        "nb_lignes": int,
        "history_count": int,
        "nb_dates_hors_ordre": int,
      }
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]
    except Exception as e:
        raise ValueError(f"Impossible de lire le fichier Excel : {e}")

    try:
        # 1. Read first row to find client name
        row_iter = sheet.iter_rows(values_only=True)
        first_row = next(row_iter, None)
        if not first_row or len(first_row) < 2:
            raise ValueError("Le fichier Excel ne respecte pas le format attendu (trop petit).")

        client_name = str(first_row[1]).strip() if first_row[1] is not None else ""
        if client_name.lower() in ("nan", "none", ""):
            # Fallback : chercher la première cellule non vide sur la première ligne
            for val in first_row:
                if val is not None:
                    val_str = str(val).strip()
                    if val_str and val_str.lower() not in ("nan", "none"):
                        client_name = val_str
                        break
            else:
                client_name = "Client inconnu"

        # Enlever les infos de téléphone de la chaîne du nom si présente
        if "tel" in client_name.lower():
            client_name = re.split(r"(?i)\s+tel\s*", client_name)[0].strip()

        # Détection dynamique de la ligne d'en-tête
        header_row_idx = None
        header_row_iter = sheet.iter_rows(values_only=True)
        for idx, row in enumerate(header_row_iter):
            row_vals = [str(val).strip().lower() for val in row if val is not None]
            if any("designation" in val for val in row_vals):
                header_row_idx = idx
                break
        else:
            header_row_idx = 2  # Fallback par défaut

        rows = []
        ordre = 0

        # Read data rows starting from the row after the header (1-based index)
        data_start_row = header_row_idx + 2
        for row in sheet.iter_rows(min_row=data_start_row, max_col=5, values_only=True):
            row_vals = list(row)
            while len(row_vals) < 5:
                row_vals.append(None)

            date_raw_val = row_vals[0]
            if date_raw_val is None:
                continue
            date_str = str(date_raw_val).strip()

            # Ignorer les lignes sans date valide ou lignes d'en-tête répétées
            if not date_str or date_str.lower() in ("nan", "none", "date", ""):
                continue

            # Parser la date (format JJ/MM/AAAA) avec détection de fautes (ex: 203 -> 2023)
            date_parsed = parse_flexible_date(date_str, fallback_to_today=False)
            if not date_parsed:
                continue  # ignorer les lignes avec date illisible

            montant_achat = _to_float(row_vals[2])
            montant_verse = _to_float(row_vals[3])
            solde_cumule  = _to_float(row_vals[4])

            designation   = str(row_vals[1] or "").strip()
            if designation.lower() in ("nan", "none"):
                designation = ""

            # Ignorer les lignes avec Montant a Payer = 0 ET Versement = 0 (sauf si ouverture / solde d'ouverture)
            if montant_achat == 0 and montant_verse == 0 and "ancien" not in designation.lower():
                continue

            # Déterminer le type d'opération
            type_op = _classify_operation(
                designation, montant_achat, montant_verse, ordre
            )

            rows.append({
                "ordre_import":  ordre,
                "date":          date_parsed,
                "designation":   designation,
                "montant_achat": montant_achat,
                "montant_verse": montant_verse,
                "solde_cumule":  solde_cumule,
                "type_operation": type_op,
            })
            ordre += 1

        if not rows:
            raise ValueError("Aucune ligne de données valide trouvée dans le fichier Excel.")

        # Compter les dates hors ordre
        nb_hors_ordre = 0
        for i in range(1, len(rows)):
            if rows[i]["date"] < rows[i-1]["date"]:
                nb_hors_ordre += 1

        return {
            "client_name":        client_name,
            "rows":               rows,
            "solde_final":        rows[-1]["solde_cumule"],
            "total_achats":       sum(r["montant_achat"] for r in rows),
            "total_verses":       sum(r["montant_verse"] for r in rows),
            "nb_lignes":          len(rows),
            "history_count":      len(rows),
            "nb_dates_hors_ordre": nb_hors_ordre,
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _to_float(val) -> float:
    """Convertit une valeur brute Excel en float, 0.0 si vide/invalide."""
    try:
        s = str(val or "").strip().replace(" ", "").replace("\xa0", "").replace(",", ".")
        if s.lower() in ("nan", "none", ""):
            return 0.0
        return float(Decimal(s))
    except (InvalidOperation, ValueError):
        return 0.0


def _classify_operation(
    designation: str,
    montant_achat: float,
    montant_verse: float,
    ordre: int
) -> str:
    """
    Classifie le type d'opération pour l'affichage.

    Règles :
    - ordre == 0 ET designation contient "ancien" → "ouverture"
    - purchase > 0 ET verse == 0  → "achat"
    - purchase == 0 ET verse > 0  → "versement"
    - purchase > 0 ET verse > 0 ET purchase == verse → "immediat" (payé cash)
    - purchase > 0 ET verse > 0 ET purchase != verse → "mixte"
    """
    if ordre == 0 and "ancien" in designation.lower():
        return "ouverture"
    if montant_achat > 0 and montant_verse == 0:
        return "achat"
    if montant_achat == 0 and montant_verse > 0:
        return "versement"
    if montant_achat > 0 and montant_verse > 0:
        if abs(montant_achat - montant_verse) < 0.01:
            return "immediat"
        return "mixte"
    return "achat"  # fallback


# === FONCTIONS ORIGINALES RESTAURÉES ===

def parse_flexible_date(value, fallback_to_today: bool = True) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip() if value is not None else ""
    if not text or text.lower() in {"none", "nan", ""}:
        return date.today().isoformat() if fallback_to_today else None

    # Standardisation des séparateurs
    text = text.replace("-", "/").replace(".", "/").replace("\\", "/")

    for fmt in ("%d/%m/%Y", "%Y/%m/%d", "%d/%m/%y", "%Y/%m/%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except Exception:
            pass

    # Gestion de l'année sur 3 chiffres (ex: 27/06/203 -> 27/06/2023)
    match = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})(?:\s+.*)?$", text)
    if match:
        day, month, year = match.groups()
        if len(year) == 3:
            year = "202" + year[2]
        try:
            return date(int(year), int(month), int(day)).isoformat()
        except Exception:
            pass

    try:
        return datetime.fromisoformat(text).date().isoformat()
    except Exception:
        pass

    try:
        return pd.to_datetime(text, dayfirst=True, errors="raise").date().isoformat()
    except Exception:
        pass

    return date.today().isoformat() if fallback_to_today else None


def parse_flexible_amount(value) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\xa0", "").replace(" ", "").replace("DA", "").replace("da", "")
    text = text.replace(",", ".") if text.count(",") == 1 and text.count(".") == 0 else text.replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else 0.0




def parse_excel_client_file(file_path) -> dict:
    try:
        import openpyxl
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("Le module openpyxl est requis pour l'import Excel.") from exc

    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]

        def cell_str(value: Any) -> str:
            return str(value).strip() if value is not None else ""

        top_rows = [[cell_str(value) for value in row] for row in sheet.iter_rows(min_row=1, max_row=min(6, getattr(sheet, "max_row", 6) or 6), values_only=True)]
        client_name = ""
        phone = ""
        for row in top_rows:
            cleaned = [value for value in row if value]
            if len(cleaned) >= 2 and not client_name:
                for value in cleaned:
                    upper = value.upper()
                    if value and "TEL" not in upper and "DATE" not in upper and "DESIGNATION" not in upper:
                        client_name = value
                        break
            for index, value in enumerate(row):
                upper = value.upper()
                if "TEL" not in upper:
                    continue
                inline = value.split(":", 1)[1].strip() if ":" in value else ""
                if inline:
                    phone = inline
                elif index + 1 < len(row):
                    phone = row[index + 1].strip()

        header_row = None
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [cell_str(value).lower() for value in row]
            joined = " | ".join(values)
            if "designation" in joined and ("montant a payer" in joined or "reste a payer" in joined):
                header_row = row_index
                break

        history_count = 0
        opening_credit = 0.0
        final_balance = 0.0
        if header_row:
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                raw_date, designation, amount, payment, balance = (list(row) + [None] * 5)[:5]
                designation_text = cell_str(designation)
                if not any(value is not None and str(value).strip() for value in (raw_date, designation, amount, payment, balance)):
                    continue
                history_count += 1
                balance_num = parse_flexible_amount(balance)
                amount_num = parse_flexible_amount(amount)
                final_balance = balance_num or final_balance
                if designation_text and "ancien solde" in designation_text.lower():
                    opening_credit = balance_num or amount_num or opening_credit

        if opening_credit <= 0 and final_balance > 0:
            opening_credit = final_balance
        if not client_name:
            client_name = Path(file_path).stem.replace("_", " ").strip()

        return {
            "name": client_name.strip(),
            "phone": phone.strip(),
            "address": "",
            "notes": f"Importe depuis Excel ({Path(file_path).name}). Lignes detectees: {history_count}.",
            "opening_credit": round(float(opening_credit or 0.0), 2),
            "history_count": history_count,
            "source_file": Path(file_path).name,
        }
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def parse_excel_bulk_clients(file_path: str) -> list[dict]:
    try:
        import openpyxl
    except Exception as exc:
        raise RuntimeError("Le module openpyxl est requis pour l'import Excel.") from exc

    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]

        def cell_str(value: Any) -> str:
            return str(value).strip() if value is not None else ""

        header_row = None
        headers_map = {}

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [cell_str(value).lower() for value in row]
            if any("nom" in val or "client" in val or "raison" in val for val in values):
                header_row = row_index
                for col_idx, val in enumerate(values):
                    if not val:
                        continue
                    if "nom" in val or "client" in val or "raison" in val:
                        headers_map["name"] = col_idx
                    elif "tel" in val or "tél" in val or "phone" in val or "mobile" in val or "gsm" in val:
                        headers_map["phone"] = col_idx
                    elif "adresse" in val or "ville" in val or "location" in val:
                        headers_map["address"] = col_idx
                    elif "solde" in val or "credit" in val or "crédit" in val or "initial" in val:
                        headers_map["opening_credit"] = col_idx
                    elif "note" in val or "remarque" in val or "obs" in val:
                        headers_map["notes"] = col_idx
                break

        if header_row is None:
            header_row = 1
            headers_map = {
                "name": 0,
                "phone": 1,
                "address": 2,
                "opening_credit": 3,
                "notes": 4
            }

        clients = []
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            row_vals = list(row)
            if not row_vals or not any(val is not None and str(val).strip() for val in row_vals):
                continue

            name_idx = headers_map.get("name")
            name = cell_str(row_vals[name_idx]) if name_idx is not None and name_idx < len(row_vals) else ""
            if not name or name.lower() in ["total", "somme"]:
                continue

            phone_idx = headers_map.get("phone")
            phone = cell_str(row_vals[phone_idx]) if phone_idx is not None and phone_idx < len(row_vals) else ""

            addr_idx = headers_map.get("address")
            address = cell_str(row_vals[addr_idx]) if addr_idx is not None and addr_idx < len(row_vals) else ""

            credit_idx = headers_map.get("opening_credit")
            credit_val = row_vals[credit_idx] if credit_idx is not None and credit_idx < len(row_vals) else 0.0
            opening_credit = parse_flexible_amount(credit_val)

            notes_idx = headers_map.get("notes")
            notes = cell_str(row_vals[notes_idx]) if notes_idx is not None and notes_idx < len(row_vals) else ""

            clients.append({
                "name": name,
                "phone": phone,
                "address": address,
                "opening_credit": round(opening_credit, 2),
                "notes": notes or f"Importé en masse depuis {Path(file_path).name}"
            })

        return clients
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def parse_excel_bulk_products(file_path: str) -> list[dict]:
    try:
        import openpyxl
    except Exception as exc:
        raise RuntimeError("Le module openpyxl est requis pour l'import Excel.") from exc

    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]

        def cell_str(value: Any) -> str:
            return str(value).strip() if value is not None else ""

        header_row = None
        headers_map = {}

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [cell_str(value).lower() for value in row]
            if any("nom" in val or "designation" in val or "désignation" in val or "produit" in val or "article" in val for val in values):
                header_row = row_index
                for col_idx, val in enumerate(values):
                    if not val:
                        continue
                    if "nom" in val or "designation" in val or "désignation" in val or "produit" in val or "article" in val:
                        headers_map["name"] = col_idx
                    elif "unite" in val or "unité" in val or "unit" in val or "mesure" in val:
                        headers_map["unit"] = col_idx
                    elif "stock" in val or "quantite" in val or "quantité" in val or "qte" in val or "qté" in val:
                        headers_map["stock_qty"] = col_idx
                    elif "prix" in val or "vente" in val or "tarif" in val or "sale" in val:
                        headers_map["sale_price"] = col_idx
                    elif "cout" in val or "coût" in val or "achat" in val or "cost" in val:
                        headers_map["avg_cost"] = col_idx
                    elif "seuil" in val or "alerte" in val or "alert" in val:
                        headers_map["alert_threshold"] = col_idx
                break

        if header_row is None:
            header_row = 1
            headers_map = {
                "name": 0,
                "unit": 1,
                "stock_qty": 2,
                "sale_price": 3,
                "avg_cost": 4,
                "alert_threshold": 5
            }

        products = []
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            row_vals = list(row)
            if not row_vals or not any(val is not None and str(val).strip() for val in row_vals):
                continue

            name_idx = headers_map.get("name")
            name = cell_str(row_vals[name_idx]) if name_idx is not None and name_idx < len(row_vals) else ""
            if not name or name.lower() in ["total", "somme"]:
                continue

            unit_idx = headers_map.get("unit")
            unit = cell_str(row_vals[unit_idx]) if unit_idx is not None and unit_idx < len(row_vals) else "kg"

            qty_idx = headers_map.get("stock_qty")
            qty_val = row_vals[qty_idx] if qty_idx is not None and qty_idx < len(row_vals) else 0.0
            stock_qty = parse_flexible_amount(qty_val)

            price_idx = headers_map.get("sale_price")
            price_val = row_vals[price_idx] if price_idx is not None and price_idx < len(row_vals) else 0.0
            sale_price = parse_flexible_amount(price_val)

            cost_idx = headers_map.get("avg_cost")
            cost_val = row_vals[cost_idx] if cost_idx is not None and cost_idx < len(row_vals) else 0.0
            avg_cost = parse_flexible_amount(cost_val)

            alert_idx = headers_map.get("alert_threshold")
            alert_val = row_vals[alert_idx] if alert_idx is not None and alert_idx < len(row_vals) else 0.0
            alert_threshold = parse_flexible_amount(alert_val)

            products.append({
                "name": name,
                "unit": unit or "kg",
                "stock_qty": round(stock_qty, 2),
                "sale_price": round(sale_price, 2),
                "avg_cost": round(avg_cost, 2),
                "alert_threshold": round(alert_threshold, 2)
            })

        return products
    finally:
        try:
            workbook.close()
        except Exception:
            pass

