"""Routes web du module Rapports & Statistiques."""
from __future__ import annotations

import csv
import io

from fastapi import APIRouter, Request, Depends
from fastapi.responses import RedirectResponse, StreamingResponse

from app.modules.reports.service import ReportsService
from app.web.deps import require_permission, template_context, templates
from app.core.perf_cache import cached_result

router = APIRouter()

def get_reports_service() -> ReportsService:
    return ReportsService()


@router.get("/reports", name="reports_dashboard")
async def reports_page(
    request: Request,
    reports_service: ReportsService = Depends(get_reports_service)
):
    denied = require_permission(request, "reports.read")
    if denied:
        return denied
    date_from = request.query_params.get("date_from", "")
    date_to = request.query_params.get("date_to", "")
    
    ctx = cached_result(
        ("dashboard", "reports", date_from or "", date_to or ""),
        lambda: reports_service.build_reports_context(date_from or None, date_to or None).dict(),
        ttl_seconds=300.0,
    )
    
    fmt = request.query_params.get("format", "").lower()
    if fmt == "csv":
        import io, csv
        from datetime import date
        output = io.StringIO()
        writer = csv.writer(output, delimiter=";")
        writer.writerow(["=== RÉSUMÉ GLOBAL ==="])
        writer.writerow(["Indicateur", "Valeur"])
        writer.writerow(["Total Ventes", ctx["summary"]["total_sales"]])
        writer.writerow(["Profit Brut", ctx["summary"]["total_profit"]])
        writer.writerow(["Total Dépenses", ctx["expenses_total"]])
        writer.writerow(["Bénéfice Net", ctx["net_profit"]])
        writer.writerow(["Total Achats", ctx["summary"]["total_purchases"]])
        writer.writerow(["Total Versements", ctx["summary"]["total_payments"]])
        writer.writerow([])
        
        writer.writerow(["=== BALANCES CLIENTS / RETARDS ==="])
        writer.writerow(["Client", "Total Dû", "Dépassement", "Dernier Versement", "Retard (jours)"])
        for c in ctx.get("client_debts", []):
            writer.writerow([c["name"], c["debt"], c["over_threshold"], c["last_payment_date"] or "-", c["overdue_days"]])
        writer.writerow([])

        writer.writerow(["=== TOP PRODUITS ==="])
        writer.writerow(["Produit", "CA", "Profit", "Quantité"])
        for p in ctx["top_products"]:
            writer.writerow([p["name"], p["revenue"], p["profit"], p["qty"]])
        writer.writerow([])

        writer.writerow(["=== TOP CLIENTS ==="])
        writer.writerow(["Client", "CA", "Profit", "Nombre"])
        for c in ctx["top_clients"]:
            writer.writerow([c["name"], c["revenue"], c["profit"], c["count"]])

        output.seek(0)
        filename = f"report_{date.today().isoformat()}.csv"
        return StreamingResponse(
            iter(["\ufeff" + output.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    elif fmt == "xlsx":
        import io, openpyxl
        from openpyxl.styles import Font
        from datetime import date
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport"
        
        ws.append(["RAPPORT DÉCISIONNEL FABOUANES"])
        ws.append([])
        
        ws.append(["=== RÉSUMÉ GLOBAL ==="])
        ws.append(["Indicateur", "Valeur"])
        ws.append(["Total Ventes", ctx["summary"]["total_sales"]])
        ws.append(["Profit Brut", ctx["summary"]["total_profit"]])
        ws.append(["Total Dépenses", ctx["expenses_total"]])
        ws.append(["Bénéfice Net", ctx["net_profit"]])
        ws.append(["Total Achats", ctx["summary"]["total_purchases"]])
        ws.append(["Total Versements", ctx["summary"]["total_payments"]])
        ws.append([])
        
        ws.append(["=== BALANCES CLIENTS / RETARDS ==="])
        ws.append(["Client", "Total Dû", "Dépassement", "Dernier Versement", "Retard (jours)"])
        for c in ctx.get("client_debts", []):
            ws.append([c["name"], c["debt"], c["over_threshold"], c["last_payment_date"] or "-", c["overdue_days"]])
        ws.append([])

        ws.append(["=== TOP PRODUITS ==="])
        ws.append(["Produit", "CA", "Profit", "Quantité"])
        for p in ctx["top_products"]:
            ws.append([p["name"], p["revenue"], p["profit"], p["qty"]])
        ws.append([])

        ws.append(["=== TOP CLIENTS ==="])
        ws.append(["Client", "CA", "Profit", "Nombre"])
        for c in ctx["top_clients"]:
            ws.append([c["name"], c["revenue"], c["profit"], c["count"]])

        ws.freeze_panes = "A4"
        
        bold_font = Font(bold=True)
        # Bold on major headings/sections
        sections = [1, 3, 12, 12 + len(ctx.get("client_debts", [])) + 2, 12 + len(ctx.get("client_debts", [])) + 2 + len(ctx["top_products"]) + 2]
        for idx in sections:
            try:
                for cell in ws[idx]:
                    cell.font = bold_font
            except Exception:
                pass

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        filename = f"report_{date.today().isoformat()}.xlsx"
        out_buf = io.BytesIO()
        wb.save(out_buf)
        out_buf.seek(0)
        return StreamingResponse(
            out_buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    
    return templates.TemplateResponse("reports_dashboard.html", template_context(
        request, title="Rapports & Statistiques", **ctx,
    ))


@router.get("/reports/export-csv", name="reports_export_csv")
async def export_csv(
    request: Request,
    reports_service: ReportsService = Depends(get_reports_service)
):
    denied = require_permission(request, "reports.read")
    if denied:
        return denied
    date_from = request.query_params.get("date_from", "")
    date_to = request.query_params.get("date_to", "")
    
    ctx = cached_result(
        ("dashboard", "reports", date_from or "", date_to or ""),
        lambda: reports_service.build_reports_context(date_from or None, date_to or None).dict(),
        ttl_seconds=300.0,
    )

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")

    # Résumé
    writer.writerow(["=== RÉSUMÉ GLOBAL ==="])
    writer.writerow(["Indicateur", "Valeur"])
    writer.writerow(["Total Ventes", ctx["summary"]["total_sales"]])
    writer.writerow(["Profit Brut", ctx["summary"]["total_profit"]])
    writer.writerow(["Total Dépenses", ctx["expenses_total"]])
    writer.writerow(["Bénéfice Net", ctx["net_profit"]])
    writer.writerow(["Total Achats", ctx["summary"]["total_purchases"]])
    writer.writerow(["Total Versements", ctx["summary"]["total_payments"]])
    writer.writerow([])

    # Top produits
    writer.writerow(["=== TOP PRODUITS ==="])
    writer.writerow(["Produit", "CA", "Profit", "Quantité"])
    for p in ctx["top_products"]:
        writer.writerow([p["name"], p["revenue"], p["profit"], p["qty"]])
    writer.writerow([])

    # Top clients
    writer.writerow(["=== TOP CLIENTS ==="])
    writer.writerow(["Client", "CA", "Profit", "Nombre"])
    for c in ctx["top_clients"]:
        writer.writerow([c["name"], c["revenue"], c["profit"], c["count"]])

    output.seek(0)
    bom = "\ufeff"
    return StreamingResponse(
        iter([bom + output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=rapport_fabouanes.csv"},
    )
