from __future__ import annotations

from datetime import datetime

from fastapi import APIRouter, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse

from app.services.print_service import COMPANY_INFO, PRINT_LAYOUT, build_print_payload, generate_invoice_pdf
from app.services.transactions_service import transactions_context, update_production_notes
from app.web.deps import csrf_protect, flash, get_current_user, login_redirect, require_permission, template_context, templates
from app.core.permissions import PERMISSION_OPERATIONS_READ, PERMISSION_PRODUCTION_WRITE, PERMISSION_OPERATIONS_WRITE


router = APIRouter()


def _print_not_found_response(message: str = "Bon introuvable.") -> HTMLResponse:
    return HTMLResponse(
        f"""<!doctype html>
<html lang="fr">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"></head>
<body style="margin:0;font-family:Arial,sans-serif;background:#f8fafc;color:#111827;display:grid;place-items:center;min-height:100vh;">
  <main style="max-width:520px;padding:24px;text-align:center;">
    <h1 style="font-size:1.25rem;margin:0 0 8px;">{message}</h1>
    <p style="margin:0;color:#64748b;">Le document demandé n'existe plus ou n'est pas disponible.</p>
  </main>
</body>
</html>""",
        status_code=404,
    )


@router.get("/operations", name="operations")
@router.get("/transactions", name="transactions")
async def operations_page(request: Request):
    denied = require_permission(request, PERMISSION_OPERATIONS_READ)
    if denied:
        return denied
    
    fmt = request.query_params.get("format", "").lower()
    if fmt in ("csv", "xlsx"):
        large_args = dict(request.query_params)
        large_args["page_size"] = 1000000
        context = transactions_context(
            filter_type=str(request.query_params.get("type", "all") or "all"),
            filter_name=str(request.query_params.get("name", "") or ""),
            filter_date=str(request.query_params.get("date", "") or ""),
            filter_operation=str(request.query_params.get("operation", "") or ""),
            args=large_args,
            path=request.url.path,
        )
        data = context["transactions"]
        
        if fmt == "csv":
            import io, csv
            from datetime import date
            output = io.StringIO()
            fieldnames = ["tx_type", "tx_date", "partner_name", "designation", "quantity", "unit", "unit_price", "total", "paid", "due"]
            writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter=";", extrasaction="ignore")
            writer.writerow({f: f.upper() for f in fieldnames})
            for row in data:
                writer.writerow(row)
            output.seek(0)
            filename = f"transactions_{date.today().isoformat()}.csv"
            return StreamingResponse(
                iter(["\ufeff" + output.getvalue()]),
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
            
        elif fmt == "xlsx":
            import io, openpyxl
            from openpyxl.styles import Font
            from datetime import date
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Transactions"
            
            headers = ["TYPE", "DATE", "TIERS", "DÉSIGNATION", "QUANTITÉ", "UNITÉ", "PRIX UNITAIRE", "TOTAL", "PAYÉ", "DU"]
            ws.append(headers)
            
            fieldnames = ["tx_type", "tx_date", "partner_name", "designation", "quantity", "unit", "unit_price", "total", "paid", "due"]
            for row in data:
                ws.append([str(row.get(f) or "") if row.get(f) is not None else "" for f in fieldnames])
                
            ws.freeze_panes = "A2"
            bold_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = bold_font
                
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
                
            filename = f"transactions_{date.today().isoformat()}.xlsx"
            out_buf = io.BytesIO()
            wb.save(out_buf)
            out_buf.seek(0)
            return StreamingResponse(
                out_buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )

    context = transactions_context(
        filter_type=str(request.query_params.get("type", "all") or "all"),
        filter_name=str(request.query_params.get("name", "") or ""),
        filter_date=str(request.query_params.get("date", "") or ""),
        filter_operation=str(request.query_params.get("operation", "") or ""),
        args=request.query_params,
        path=request.url.path,
    )
    return templates.TemplateResponse("transactions.html", template_context(request, **context))


@router.get("/operations/new", name="new_operation")
async def new_operation_page(request: Request):
    denied = require_permission(request, PERMISSION_OPERATIONS_WRITE)
    if denied:
        return denied
    
    from app.services.purchase_service import purchase_form_context
    from app.services.sale_service import sale_form_context
    from app.services.payment_service import new_payment_context
    from app.core.db_access import query_db

    p_ctx = purchase_form_context()
    s_ctx = sale_form_context()
    pay_ctx = new_payment_context()
    
    context = {}
    context.update(p_ctx)
    context.update(s_ctx)
    context.update(pay_ctx)
    
    context["clients"] = query_db("SELECT * FROM clients ORDER BY name")
    context["suppliers"] = query_db("SELECT * FROM suppliers ORDER BY name")
    context["mode"] = request.query_params.get("mode", "achat")
    
    return templates.TemplateResponse("operation_new.html", template_context(request, **context))



@router.get("/print/{doc_type}/{item_id}", name="print_document")
async def print_document_page(request: Request, doc_type: str, item_id: int):
    user = get_current_user(request)
    if not user:
        return login_redirect()
    payload = build_print_payload(doc_type, item_id)
    if not payload:
        return _print_not_found_response("Document introuvable pour impression.")
    printed_at = datetime.now()
    payload = {
        **payload,
        "printed_date": printed_at.strftime("%Y-%m-%d"),
        "printed_time": printed_at.strftime("%H:%M"),
    }
    if request.query_params.get("format") == "pdf":
        pdf_buf = generate_invoice_pdf(payload, user["username"])
        if pdf_buf:
            headers = {"Content-Disposition": f'attachment; filename="{payload["number"]}.pdf"'}
            return StreamingResponse(pdf_buf, media_type="application/pdf", headers=headers)
        flash(request, "Génération PDF indisponible. Affichage HTML utilisé à la place.", "warning")
    return templates.TemplateResponse(
        "print_document.html",
        template_context(
            request,
            doc=payload,
            company=payload.get("company") or COMPANY_INFO,
            printed_by=user["username"],
            print_layout=PRINT_LAYOUT,
        ),
    )


@router.post("/production/edit-notes", name="edit_production_notes")
@router.post("/production/notes")
async def edit_production_notes(request: Request):
    denied = require_permission(request, PERMISSION_PRODUCTION_WRITE)
    if denied:
        return denied
    await csrf_protect(request)
    form = await request.form()
    try:
        batch_id = int(str(form.get("batch_id", "") or "0"))
    except ValueError:
        batch_id = 0
    try:
        update_production_notes(
            batch_id=batch_id,
            production_date=str(form.get("production_date", "") or "").strip(),
            notes=str(form.get("notes", "") or "").strip(),
        )
        flash(request, "Notes de production mises à jour.", "success")
    except Exception as exc:
        from app.core.exceptions import get_friendly_error_message
        flash(request, get_friendly_error_message(exc), "danger")
    return RedirectResponse("/production", status_code=303)
